from __future__ import annotations

import calendar
import csv
import json
from collections import Counter, defaultdict
from datetime import date
from functools import lru_cache
from io import BytesIO, StringIO
from urllib.parse import urlencode
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

from shift_core.models import (
    PreviousShiftRecord,
    ShiftAssignment,
    ShiftPeriod,
    ShiftRequest,
    Staff,
    StaffSkill,
    WorkType,
)


def period_matrix(period: ShiftPeriod):
    staff_list = list(Staff.objects.filter(is_active=True).order_by("employee_number"))
    works = list(WorkType.objects.filter(active=True).order_by("display_order", "name"))
    assignments = {
        (item.staff_id, item.day): item
        for item in period.assignments.select_related("work_type")
    }
    request_map = {
        (item.staff_id, item.day): item
        for item in ShiftRequest.objects.filter(
            day__gte=period.month,
            day__lte=period.month.replace(
                day=calendar.monthrange(period.month.year, period.month.month)[1]
            ),
        )
    }
    day_dates = sorted({item.day for item in period.assignments.all()})
    days = [_day_context(day) for day in day_dates]
    skill_map = {
        (skill.staff_id, skill.work_type_id): skill
        for skill in StaffSkill.objects.select_related("level").filter(
            staff__in=staff_list,
            work_type__in=works,
        )
    }
    previous_days = _previous_day_contexts(period)
    previous_map = {
        (item.staff_id, item.day): item
        for item in PreviousShiftRecord.objects.select_related("work_type").filter(
            day__in=[day["date"] for day in previous_days],
            staff__in=staff_list,
        )
    }
    rows = []
    for staff in staff_list:
        cells = []
        work_summary = []
        work_counts = Counter()
        work_count = 0
        public_holiday_count = 0
        paid_leave_count = 0
        standby_count = 0
        for day in days:
            assignment = assignments.get((staff.id, day["date"]))
            if assignment and assignment.status == ShiftAssignment.Status.WORK:
                work_count += 1
                work_counts[assignment.work_type_id] += 1
            elif assignment and assignment.status == ShiftAssignment.Status.PUBLIC_HOLIDAY:
                public_holiday_count += 1
            elif assignment and assignment.status == ShiftAssignment.Status.PAID_LEAVE:
                paid_leave_count += 1
            elif assignment and assignment.status == ShiftAssignment.Status.STANDBY:
                standby_count += 1
            skill = (
                skill_map.get((staff.id, assignment.work_type_id))
                if assignment and assignment.work_type_id
                else None
            )
            request = request_map.get((staff.id, day["date"]))
            cells.append(
                {
                    "day": day,
                    "assignment": assignment,
                    "label": _label(assignment),
                    "work": assignment.work_type if assignment and assignment.work_type else None,
                    "field_name": f"assignment_{assignment.id}" if assignment else "",
                    "cell_class": _cell_class(assignment),
                    "rest_kind": _rest_kind(assignment),
                    "rest_label": _label(assignment),
                    "source_class": assignment.source if assignment else "",
                    "source_label": assignment.get_source_display() if assignment else "",
                    "request_label": _request_label(request),
                    "request_kind": request.kind if request else "",
                    "request_conflict": _request_conflict(request, assignment),
                    "assignment_is_trainee": bool(skill and skill.level.trainee),
                    "assignment_is_instructor": bool(skill and skill.level.instructor),
                }
            )
        for work in works:
            work_summary.append({"work": work, "count": work_counts[work.id]})
        expected_total = work_count + public_holiday_count + paid_leave_count + standby_count
        public_holiday_target = staff.public_holiday_count
        public_holiday_status = (
            "ok"
            if public_holiday_count == public_holiday_target
            else "under"
            if public_holiday_count < public_holiday_target
            else "over"
        )
        rows.append(
            {
                "staff": staff,
                "cells": cells,
                "previous_cells": _previous_cells(staff, previous_days, previous_map),
                "work_count": work_count,
                "standby_count": standby_count,
                "public_holiday_count": public_holiday_count,
                "public_holiday_target": public_holiday_target,
                "public_holiday_status": public_holiday_status,
                "public_holiday_status_label": "不足"
                if public_holiday_status == "under"
                else "超過"
                if public_holiday_status == "over"
                else "",
                "paid_leave_count": paid_leave_count,
                "expected_total": expected_total,
                "expected_total_status": "ok"
                if expected_total == len(days)
                else "under"
                if expected_total < len(days)
                else "over",
                "work_summary": work_summary,
            }
        )
    daily_work_stats = _daily_work_stats(days, works, assignments, skill_map)
    return days, rows, works, previous_days, daily_work_stats


def build_shift_workbook(period: ShiftPeriod) -> bytes:
    days, rows, _works, _previous_days, _daily_work_stats = period_matrix(period)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "シフト表"
    sheet.append([f"{period.month:%Y年%m月} シフト表"])
    sheet.append(["社員番号", "氏名", *[f"{day['number']}日" for day in days], "公休", "有休", "出勤"])
    header_fill = PatternFill("solid", fgColor="E0F2FE")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_data in rows:
        public_holidays = sum(1 for cell in row_data["cells"] if cell["assignment"] and cell["assignment"].status == ShiftAssignment.Status.PUBLIC_HOLIDAY)
        paid_leaves = sum(1 for cell in row_data["cells"] if cell["assignment"] and cell["assignment"].status == ShiftAssignment.Status.PAID_LEAVE)
        work_days = sum(1 for cell in row_data["cells"] if cell["assignment"] and cell["assignment"].status in {ShiftAssignment.Status.WORK, ShiftAssignment.Status.STANDBY})
        sheet.append([
            row_data["staff"].employee_number,
            row_data["staff"].name,
            *[cell["label"] for cell in row_data["cells"]],
            public_holidays,
            paid_leaves,
            work_days,
        ])
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "C3"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_shift_csv(period: ShiftPeriod) -> str:
    days, rows, _works, _previous_days, _daily_work_stats = period_matrix(period)
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow([f"{period.month:%Y年%m月} シフト表"])
    writer.writerow(["社員番号", "氏名", *[f"{day['number']}日" for day in days]])
    for row_data in rows:
        writer.writerow([
            row_data["staff"].employee_number,
            row_data["staff"].name,
            *[cell["label"] for cell in row_data["cells"]],
        ])
    return stream.getvalue()


def _day_context(day):
    holiday = _holiday_for_day(day)
    return {
        "date": day,
        "number": day.day,
        "weekday": _weekday_label(day),
        "holiday": holiday,
        "is_holiday": bool(holiday),
        "is_saturday": day.weekday() == 5,
        "is_non_workday": day.weekday() == 6 or bool(holiday),
    }


def _weekday_label(day):
    return "月火水木金土日"[day.weekday()]


def _holiday_for_day(day: date) -> str:
    return _holidays_for_month(day.replace(day=1)).get(day, "")


def _koyomi_api_url(month: date) -> str:
    day_count = calendar.monthrange(month.year, month.month)[1]
    params = {
        "mode": "d",
        "cnt": day_count,
        "targetyyyy": month.year,
        "targetmm": f"{month.month:02d}",
        "targetdd": "01",
    }
    return "https://koyomi.zingsystem.com/api/?" + urlencode(params)


@lru_cache(maxsize=24)
def _holidays_for_month(month: date) -> dict[date, str]:
    try:
        with urlopen(_koyomi_api_url(month), timeout=5) as response:
            data = json.load(response)
    except Exception:
        return {}

    holidays = {}
    for day_text, info in data.get("datelist", {}).items():
        holiday_name = info.get("holiday", "")
        if holiday_name:
            holidays[date.fromisoformat(day_text)] = holiday_name
    return holidays


def _previous_day_contexts(period):
    previous_dates = list(
        PreviousShiftRecord.objects.filter(day__lt=period.month)
        .order_by("-day")
        .values_list("day", flat=True)
        .distinct()[:7]
    )
    return [_day_context(day) for day in sorted(previous_dates)]


def _previous_cells(staff, previous_days, previous_map):
    cells = []
    for day in previous_days:
        record = previous_map.get((staff.id, day["date"]))
        work = record.work_type if record and record.work_type else None
        cells.append(
            {
                "day": day,
                "label": _previous_label(record),
                "work": work,
                "status": record.status if record else PreviousShiftRecord.Status.BLANK,
                "cell_class": _previous_cell_class(record),
            }
        )
    return cells


def _daily_work_stats(days, works, assignments, skill_map):
    result = []
    for work in works:
        counts = []
        total = 0
        for day in days:
            count = 0
            for assignment in assignments.values():
                if (
                    assignment.day != day["date"]
                    or assignment.status != ShiftAssignment.Status.WORK
                    or assignment.work_type_id != work.id
                ):
                    continue
                skill = skill_map.get((assignment.staff_id, assignment.work_type_id))
                if skill and skill.level.trainee:
                    continue
                count += 1
            counts.append(count)
            total += count
        result.append({"work": work, "counts": counts, "total": total})
    return result


def _cell_class(assignment):
    if not assignment:
        return "rest"
    return "assigned" if assignment.status == ShiftAssignment.Status.WORK else "rest"


def _rest_kind(assignment):
    if not assignment:
        return "assignment-pending"
    if assignment.status == ShiftAssignment.Status.PUBLIC_HOLIDAY:
        return "inserted-public-holiday"
    if assignment.status == ShiftAssignment.Status.PAID_LEAVE:
        return "paid-leave"
    if assignment.status == ShiftAssignment.Status.STANDBY:
        return "assignment-pending"
    return ""


def _previous_label(record):
    if not record:
        return "—"
    if record.status == PreviousShiftRecord.Status.WORK:
        return record.work_type.name if record.work_type else "勤務"
    if record.status == PreviousShiftRecord.Status.PUBLIC_HOLIDAY:
        return "公休"
    if record.status == PreviousShiftRecord.Status.PAID_LEAVE:
        return "有休"
    if record.status == PreviousShiftRecord.Status.STANDBY:
        return "余剰"
    return "—"


def _previous_cell_class(record):
    if not record:
        return "rest blank"
    return "assigned" if record.status == PreviousShiftRecord.Status.WORK else "rest"


def _label(assignment):
    if not assignment:
        return ""
    if assignment.status == ShiftAssignment.Status.WORK:
        return assignment.work_type.name if assignment.work_type else "勤務"
    if assignment.status == ShiftAssignment.Status.PUBLIC_HOLIDAY:
        return "公休"
    if assignment.status == ShiftAssignment.Status.PAID_LEAVE:
        return "有休"
    if assignment.status == ShiftAssignment.Status.STANDBY:
        return "余剰"
    return ""


def _request_label(request):
    if not request:
        return ""
    if request.kind == ShiftRequest.Kind.PUBLIC_HOLIDAY:
        return "希望公休"
    if request.kind == ShiftRequest.Kind.PAID_LEAVE:
        return "希望有休"
    if request.kind == ShiftRequest.Kind.UNAVAILABLE:
        return "勤務不可"
    return request.get_kind_display()


def _request_conflict(request, assignment):
    if not request or not assignment:
        return False
    if request.kind == ShiftRequest.Kind.PUBLIC_HOLIDAY:
        return assignment.status != ShiftAssignment.Status.PUBLIC_HOLIDAY
    if request.kind == ShiftRequest.Kind.PAID_LEAVE:
        return assignment.status != ShiftAssignment.Status.PAID_LEAVE
    if request.kind == ShiftRequest.Kind.UNAVAILABLE:
        return assignment.status == ShiftAssignment.Status.WORK
    return False
