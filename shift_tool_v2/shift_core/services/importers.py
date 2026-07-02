from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import from_excel

from shift_core.models import (
    PreviousShiftRecord,
    ShiftRequest,
    SkillLevel,
    Staff,
    StaffSkill,
    WorkType,
)


PUBLIC_HOLIDAY_TOKENS = {"休", "公休", "公", "休日"}
PAID_LEAVE_TOKENS = {"有休", "有給", "年休", "有"}
UNAVAILABLE_TOKENS = {"不可", "勤務不可", "出勤不可", "×", "✕", "NG", "ＮＧ"}
STANDBY_TOKENS = {"余剰", "応援", "社員"}
BLANK_TOKENS = {"", "-", "ー", "－", "未入力"}


def import_master_workbook(file_obj) -> dict[str, int]:
    workbook = load_workbook(file_obj, data_only=True)
    _ensure_default_skill_levels()
    level_count = _import_skill_levels(workbook)
    work_count = _import_works(workbook)
    staff_count = _import_staff_and_skills(workbook)
    previous_count = _import_previous_records(workbook)
    request_count = _import_shift_requests(workbook)
    return {
        "staff": staff_count,
        "works": work_count,
        "levels": level_count,
        "previous": previous_count,
        "requests": request_count,
    }


def build_template_workbook(sample: bool = False) -> Workbook:
    workbook = Workbook()
    skill_sheet = workbook.active
    skill_sheet.title = "スキル表"

    works_sheet = workbook.create_sheet("業務マスタ")
    works_sheet.append(["業務名", "必要人数", "有効"])

    levels_sheet = workbook.create_sheet("スキル区分")
    levels_sheet.append(["記号", "意味", "優先度", "配置可", "指導可能", "研修中"])
    for row in (
        ["◎", "指導可能", 1, "可", "はい", "いいえ"],
        ["○", "配置可", 3, "可", "いいえ", "いいえ"],
        ["△", "研修中", 8, "可", "いいえ", "はい"],
        ["×", "不可", 99, "不可", "いいえ", "いいえ"],
    ):
        levels_sheet.append(row)

    previous_sheet = workbook.create_sheet("先月シフト実績")
    previous_sheet.append(["社員番号", "氏名", "6/20", "6/21", "6/22", "6/23", "6/24", "6/25", "6/26", "6/27", "6/28", "6/29", "6/30"])
    request_sheet = workbook.create_sheet("シフト提出")
    request_sheet.append(["社員番号", "氏名", "7/1", "7/2", "7/3", "7/4", "7/5", "7/6", "7/7"])

    if sample:
        works = [
            ("A", 4, "有効", "2563EB"),
            ("B", 3, "有効", "16A34A"),
            ("C", 3, "有効", "F97316"),
            ("D", 2, "有効", "9333EA"),
            ("E", 2, "有効", "DC2626"),
            ("F", 2, "有効", "0F766E"),
            ("G", 2, "有効", "0891B2"),
            ("H", 1, "有効", "7C3AED"),
        ]
        for name, required, active, color in works:
            works_sheet.append([name, required, active])
            works_sheet.cell(row=works_sheet.max_row, column=1).fill = PatternFill("solid", fgColor=color)
        skill_sheet.append(["社員番号", "氏名", "公休数", *[item[0] for item in works]])
        symbols = ["◎", "○", "○", "△", "○", "×", "○", "○"]
        previous_values = ["A", "B", "C", "公休", "D", "E", "余剰", "公休", "F", "G", "H"]
        for index in range(1, 31):
            code = f"S{index:03}"
            name = f"スタッフ{index:02}"
            shift = (index - 1) % len(symbols)
            row_symbols = symbols[shift:] + symbols[:shift]
            skill_sheet.append([code, name, 6 + (1 if index % 5 == 0 else 0), *row_symbols])
            previous_shift = (index - 1) % len(previous_values)
            previous_sheet.append([code, name, *previous_values[previous_shift:], *previous_values[:previous_shift]])
            request_values = ["", "公休", "", "有休", "", "不可", ""]
            request_shift = (index - 1) % len(request_values)
            request_sheet.append([code, name, *request_values[request_shift:], *request_values[:request_shift]])
    else:
        works_sheet.append(["A", 3, "有効"])
        skill_sheet.append(["社員番号", "氏名", "公休数", "A"])
        skill_sheet.append(["S001", "山田 太郎", 8, "○"])
        previous_sheet.append(["S001", "山田 太郎", "A", "公休", "A", "A", "余剰", "公休", "A", "A", "A", "公休", "A"])
        request_sheet.append(["S001", "山田 太郎", "公休", "", "有休", "", "", "不可", ""])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = 14
    return workbook


def _ensure_default_skill_levels() -> None:
    defaults = [
        ("◎", "指導可能", 1, True, True, False),
        ("○", "配置可", 3, True, False, False),
        ("△", "研修中", 8, True, False, True),
        ("×", "不可", 99, False, False, False),
    ]
    for symbol, label, priority, assignable, instructor, trainee in defaults:
        SkillLevel.objects.update_or_create(
            symbol=symbol,
            defaults={
                "label": label,
                "priority": priority,
                "assignable": assignable,
                "instructor": instructor,
                "trainee": trainee,
            },
        )


def _import_skill_levels(workbook) -> int:
    sheet = _sheet(workbook, "スキル区分")
    if not sheet:
        return 0
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        symbol = _text(row[0] if len(row) > 0 else "")
        if not symbol:
            continue
        SkillLevel.objects.update_or_create(
            symbol=symbol,
            defaults={
                "label": _text(row[1] if len(row) > 1 else ""),
                "priority": max(1, _int(row[2] if len(row) > 2 else 5, 5)),
                "assignable": _yes(row[3] if len(row) > 3 else True),
                "instructor": _yes(row[4] if len(row) > 4 else False),
                "trainee": _yes(row[5] if len(row) > 5 else False),
            },
        )
        count += 1
    return count


def _import_works(workbook) -> int:
    sheet = _sheet(workbook, "業務マスタ", "業務")
    if not sheet:
        return 0
    count = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        name = _text(row[0].value)
        if not name:
            continue
        color = _cell_color(row[0])
        WorkType.objects.update_or_create(
            name=name,
            defaults={
                "required_staff_per_day": max(1, _int(row[1].value, 1)),
                "active": _active(row[2].value if len(row) > 2 else "有効"),
                "display_order": row_index,
                "color": color,
            },
        )
        count += 1
    return count


def _import_staff_and_skills(workbook) -> int:
    skill_sheet = _sheet(workbook, "スキル表")
    staff_sheet = _sheet(workbook, "スタッフ")
    separate_skills_sheet = _sheet(workbook, "スキル")
    if not skill_sheet and not staff_sheet:
        return 0

    count = 0
    source_sheet = skill_sheet or staff_sheet
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        code = _text(row[0] if len(row) > 0 else "")
        name = _text(row[1] if len(row) > 1 else "")
        if not code or not name:
            continue
        Staff.objects.update_or_create(
            employee_number=code,
            defaults={
                "name": name,
                "public_holiday_count": max(0, _int(row[2] if len(row) > 2 else 8, 8)),
                "note": "",
                "is_active": True,
            },
        )
        count += 1

    skills_sheet = skill_sheet or separate_skills_sheet
    if skills_sheet:
        start_col = 3 if skill_sheet else 2
        work_names = [_text(cell.value) for cell in skills_sheet[1][start_col:]]
        work_map = {work.name: work for work in WorkType.objects.all()}
        level_map = {level.symbol: level for level in SkillLevel.objects.all()}
        for row in skills_sheet.iter_rows(min_row=2, values_only=True):
            code = _text(row[0] if len(row) > 0 else "")
            staff = Staff.objects.filter(employee_number=code).first()
            if not staff:
                continue
            for offset, work_name in enumerate(work_names, start=start_col):
                work = work_map.get(work_name)
                symbol = _text(row[offset] if len(row) > offset else "")
                level = level_map.get(symbol)
                if work and level:
                    StaffSkill.objects.update_or_create(
                        staff=staff,
                        work_type=work,
                        defaults={"level": level},
                    )
    return count


def _import_previous_records(workbook) -> int:
    sheet = _sheet(workbook, "先月シフト実績", "前月実績")
    if not sheet:
        return 0
    headers = [cell.value for cell in sheet[1]]
    work_map = {work.name: work for work in WorkType.objects.all()}
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        staff = Staff.objects.filter(employee_number=_text(row[0] if row else "")).first()
        if not staff:
            continue
        for index, header in enumerate(headers[2:], start=2):
            day = _parse_day(header)
            if not day:
                continue
            raw = _text(row[index] if len(row) > index else "")
            status, work = _classify_shift_value(raw, work_map)
            PreviousShiftRecord.objects.update_or_create(
                staff=staff,
                day=day,
                defaults={"status": status, "work_type": work, "raw_value": raw},
            )
            count += 1
    return count


def _import_shift_requests(workbook) -> int:
    sheet = _sheet(workbook, "シフト提出", "シフト希望", "公休申請")
    if not sheet:
        return 0
    headers = [cell.value for cell in sheet[1]]
    count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        staff = Staff.objects.filter(employee_number=_text(row[0] if row else "")).first()
        if not staff:
            continue
        for index, header in enumerate(headers[2:], start=2):
            day = _parse_day(header)
            if not day:
                continue
            raw = _text(row[index] if len(row) > index else "")
            kind = _classify_request_value(raw)
            if not kind:
                ShiftRequest.objects.filter(staff=staff, day=day).delete()
                continue
            ShiftRequest.objects.update_or_create(
                staff=staff,
                day=day,
                defaults={"kind": kind, "raw_value": raw},
            )
            count += 1
    return count


def _classify_request_value(value: str):
    compact = value.replace(" ", "").replace("　", "")
    if compact in BLANK_TOKENS:
        return None
    if compact in PUBLIC_HOLIDAY_TOKENS:
        return ShiftRequest.Kind.PUBLIC_HOLIDAY
    if compact in PAID_LEAVE_TOKENS:
        return ShiftRequest.Kind.PAID_LEAVE
    if compact in UNAVAILABLE_TOKENS:
        return ShiftRequest.Kind.UNAVAILABLE
    return None


def _classify_shift_value(value: str, work_map: dict[str, WorkType]):
    compact = value.replace(" ", "").replace("　", "")
    if compact in BLANK_TOKENS:
        return PreviousShiftRecord.Status.BLANK, None
    if compact in PUBLIC_HOLIDAY_TOKENS:
        return PreviousShiftRecord.Status.PUBLIC_HOLIDAY, None
    if compact in PAID_LEAVE_TOKENS:
        return PreviousShiftRecord.Status.PAID_LEAVE, None
    if compact in STANDBY_TOKENS:
        return PreviousShiftRecord.Status.STANDBY, None
    work = work_map.get(value) or work_map.get(compact)
    if work:
        return PreviousShiftRecord.Status.WORK, work
    raise ValueError(f"前月実績の「{value}」は業務名・公休・有休・余剰のどれにも一致しません。")


def workbook_response(workbook: Workbook, filename: str):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue(), filename


def _parse_day(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%m/%d", "%m月%d日"):
        try:
            parsed = datetime.strptime(text, fmt)
            year = date.today().year if "%Y" not in fmt else parsed.year
            return date(year, parsed.month, parsed.day)
        except ValueError:
            continue
    return None


def _cell_color(cell) -> str:
    color = cell.fill.fgColor.rgb
    if not color or color == "00000000":
        return ""
    return f"#{color[-6:]}"


def _active(value) -> bool:
    return _text(value) not in {"無効", "停止", "0", "false", "False"}


def _yes(value) -> bool:
    return _text(value).lower() not in {"", "不可", "いいえ", "no", "false", "0"}


def _sheet(workbook, *names):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def _int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _text(value) -> str:
    return str(value or "").strip()
