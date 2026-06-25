import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from shifts.domain.import_data import (
    ImportedSkillLevel,
    ImportedSkillMap,
    ImportedStaffRow,
    ImportedWorkType,
)


class SkillMapReadError(ValueError):
    pass


def _read_tabular_file(filename: str, file_obj, error_class):
    extension = Path(filename).suffix.lower()
    raw = file_obj.read()
    if extension == ".xlsx":
        return _read_xlsx(raw, error_class)
    if extension == ".xls":
        return _read_xls(raw, error_class)
    if extension == ".csv":
        return _read_csv(raw, error_class)
    raise error_class(".xlsx、.xls、.csvのいずれかを選択してください。")


def _read_xlsx(raw, error_class):
    try:
        sheet = load_workbook(BytesIO(raw), data_only=True, read_only=True).active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    except Exception as exc:
        raise error_class("Excelファイルを読み込めませんでした。") from exc


def _read_xls(raw, error_class):
    try:
        import xlrd

        sheet = xlrd.open_workbook(file_contents=raw).sheet_by_index(0)
        return [sheet.row_values(index) for index in range(sheet.nrows)]
    except Exception as exc:
        raise error_class("旧形式Excelファイルを読み込めませんでした。") from exc


def _read_csv(raw, error_class):
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return list(csv.reader(StringIO(raw.decode(encoding))))
        except UnicodeDecodeError:
            continue
    raise error_class("CSVの文字コードを判定できませんでした。")


class SkillMapFileReader:
    """xlsx/xls/csvを共通のDomain取込データへ変換する。"""

    def read(self, filename: str, file_obj) -> ImportedSkillMap:
        extension = Path(filename).suffix.lower()
        if extension == ".xlsx":
            rows, skill_level_rows, work_type_rows = self._read_xlsx_workbook(
                file_obj.read()
            )
            return self._convert(rows, skill_level_rows, work_type_rows)

        rows = _read_tabular_file(filename, file_obj, SkillMapReadError)
        return self._convert(rows, (), ())

    @staticmethod
    def _read_xlsx_workbook(raw):
        try:
            workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
            skill_sheet = (
                workbook["スキル表"] if "スキル表" in workbook.sheetnames else workbook.active
            )
            level_sheet = next(
                (
                    workbook[name]
                    for name in ("スキル区分", "マークの意味", "スキル記号")
                    if name in workbook.sheetnames
                ),
                None,
            )
            work_sheet = next(
                (
                    workbook[name]
                    for name in ("業務マスタ", "業務", "業務一覧")
                    if name in workbook.sheetnames
                ),
                None,
            )
            skill_rows = [list(row) for row in skill_sheet.iter_rows(values_only=True)]
            level_rows = (
                [list(row) for row in level_sheet.iter_rows(values_only=True)]
                if level_sheet
                else ()
            )
            work_rows = (
                SkillMapFileReader._work_rows_with_fill_colors(work_sheet)
                if work_sheet
                else ()
            )
            return skill_rows, level_rows, work_rows
        except Exception as exc:
            raise SkillMapReadError("Excelファイルを読み込めませんでした。") from exc

    @staticmethod
    def _work_rows_with_fill_colors(sheet):
        rows = list(sheet.iter_rows())
        if not rows:
            return ()

        header_values = [cell.value for cell in rows[0]]
        headers = [str(value or "").strip() for value in header_values]
        if "業務名" not in headers:
            return [list(row) for row in sheet.iter_rows(values_only=True)]

        name_index = headers.index("業務名")
        color_index = next(
            (
                headers.index(header)
                for header in ("色", "表示色", "カラー", "色コード")
                if header in headers
            ),
            None,
        )
        if color_index is None:
            color_index = len(header_values)
            header_values.append("色")

        converted_rows = [header_values]
        for raw_cells in rows[1:]:
            values = [cell.value for cell in raw_cells]
            values += [""] * max(0, color_index + 1 - len(values))
            current_color = SkillMapFileReader._color_cell(values[color_index])
            fill_color = ""
            if color_index < len(raw_cells):
                fill_color = SkillMapFileReader._excel_fill_color(raw_cells[color_index])
            if not fill_color and name_index < len(raw_cells):
                fill_color = SkillMapFileReader._excel_fill_color(raw_cells[name_index])
            values[color_index] = current_color or fill_color
            converted_rows.append(values)
        return converted_rows

    @staticmethod
    def _excel_fill_color(cell):
        fill = getattr(cell, "fill", None)
        if not fill or not fill.fill_type:
            return ""
        color = fill.fgColor or fill.start_color
        if getattr(color, "type", "") != "rgb" or not color.rgb:
            return ""
        rgb = str(color.rgb)
        if len(rgb) == 8:
            rgb = rgb[-6:]
        if len(rgb) == 6 and all(char in "0123456789abcdefABCDEF" for char in rgb):
            return f"#{rgb.lower()}"
        return ""

    @staticmethod
    def _convert(rows, skill_level_rows=(), work_type_rows=()):
        if not rows:
            return ImportedSkillMap(
                (),
                SkillMapFileReader._convert_skill_levels(skill_level_rows),
                SkillMapFileReader._convert_work_types(work_type_rows),
            )
        headers = [str(value or "").strip() for value in rows[0]]
        required = {"社員番号", "氏名"}
        if not required.issubset(headers):
            raise SkillMapReadError("見出しに「社員番号」と「氏名」が必要です。")
        employee_index, name_index = headers.index("社員番号"), headers.index("氏名")
        note_index = headers.index("備考") if "備考" in headers else None
        public_holiday_index = next(
            (
                headers.index(header)
                for header in ("公休数", "月公休数", "月の公休数")
                if header in headers
            ),
            None,
        )
        staff_master_columns = {
            "社員番号",
            "氏名",
            "備考",
            "公休数",
            "月公休数",
            "月の公休数",
            "希望上限",
            "希望上限日数",
            "休み希望上限",
        }
        work_columns = [
            (index, header)
            for index, header in enumerate(headers)
            if header and header not in staff_master_columns
        ]
        result = []
        for raw_row in rows[1:]:
            row = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
            employee_number, name = (
                str(row[employee_index] or "").strip(),
                str(row[name_index] or "").strip(),
            )
            if not employee_number and not name:
                continue
            if not employee_number or not name:
                raise SkillMapReadError("社員番号または氏名が空の行があります。")
            skills = {
                work: str(row[index] or "").strip()
                for index, work in work_columns
                if str(row[index] or "").strip()
            }
            note = str(row[note_index] or "").strip() if note_index is not None else ""
            monthly_public_holidays = SkillMapFileReader._integer_cell(
                row[public_holiday_index] if public_holiday_index is not None else "",
                default=8,
            )
            result.append(
                ImportedStaffRow(
                    employee_number,
                    name,
                    note,
                    skills,
                    max(0, monthly_public_holidays),
                )
            )
        return ImportedSkillMap(
            tuple(result),
            SkillMapFileReader._convert_skill_levels(skill_level_rows),
            SkillMapFileReader._convert_work_types(work_type_rows),
        )

    @staticmethod
    def _convert_skill_levels(rows):
        if not rows:
            return ()
        headers = [str(value or "").strip() for value in rows[0]]
        if "記号" not in headers:
            return ()

        symbol_index = headers.index("記号")
        meaning_index = headers.index("意味") if "意味" in headers else None
        priority_index = headers.index("優先度") if "優先度" in headers else None
        assignable_index = (
            headers.index("アサイン可") if "アサイン可" in headers else None
        )
        result = []
        for raw_row in rows[1:]:
            row = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
            symbol = SkillMapFileReader._cell(row[symbol_index])
            if not symbol:
                continue

            meaning = (
                SkillMapFileReader._cell(row[meaning_index])
                if meaning_index is not None
                else symbol
            )
            priority = SkillMapFileReader._integer_cell(
                row[priority_index] if priority_index is not None else "",
                default=99,
            )
            assignable = SkillMapFileReader._boolean_cell(
                row[assignable_index] if assignable_index is not None else "",
                default=symbol not in {"×", "-"},
            )
            result.append(
                ImportedSkillLevel(
                    symbol=symbol,
                    meaning=meaning or symbol,
                    priority=priority,
                    assignable=assignable,
                )
            )
        return tuple(result)

    @staticmethod
    def _convert_work_types(rows):
        if not rows:
            return ()
        headers = [str(value or "").strip() for value in rows[0]]
        if "業務名" not in headers:
            return ()

        name_index = headers.index("業務名")
        minimum_index = next(
            (
                headers.index(header)
                for header in ("最低必要人数", "最低必要人数/日", "必要人数", "必要人数/日")
                if header in headers
            ),
            None,
        )
        active_index = headers.index("有効") if "有効" in headers else None
        color_index = next(
            (
                headers.index(header)
                for header in ("色", "表示色", "カラー", "色コード")
                if header in headers
            ),
            None,
        )

        result = []
        for raw_row in rows[1:]:
            row = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
            name = SkillMapFileReader._cell(row[name_index])
            if not name:
                continue
            minimum_staff = SkillMapFileReader._integer_cell(
                row[minimum_index] if minimum_index is not None else "",
                default=1,
            )
            active = SkillMapFileReader._boolean_cell(
                row[active_index] if active_index is not None else "",
                default=True,
            )
            color = SkillMapFileReader._color_cell(
                row[color_index] if color_index is not None else ""
            )
            result.append(
                ImportedWorkType(
                    name=name,
                    minimum_staff_per_day=max(1, minimum_staff),
                    active=active,
                    color=color,
                )
            )
        return tuple(result)

    @staticmethod
    def _cell(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _integer_cell(value, default):
        text = SkillMapFileReader._cell(value)
        if not text:
            return default
        try:
            return int(float(text))
        except ValueError:
            return default

    @staticmethod
    def _boolean_cell(value, default):
        text = SkillMapFileReader._cell(value).lower()
        if not text:
            return default
        if text in {"1", "true", "yes", "y", "可能", "可", "はい", "○", "ok"}:
            return True
        if text in {"0", "false", "no", "n", "不可", "いいえ", "×", "ng"}:
            return False
        return default

    @staticmethod
    def _color_cell(value):
        text = SkillMapFileReader._cell(value)
        if not text:
            return ""
        if not text.startswith("#"):
            text = f"#{text}"
        hex_part = text[1:]
        if len(text) == 7 and all(char in "0123456789abcdefABCDEF" for char in hex_part):
            return f"#{hex_part.lower()}"
        return ""
