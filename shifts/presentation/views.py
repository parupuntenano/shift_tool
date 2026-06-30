import calendar
import csv
import json
import re
from collections import defaultdict
from io import BytesIO, StringIO
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import urlopen

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shifts.application.availability_suggestions import (
    RestRuleInput,
    build_staff_rest_constraint_notes,
    build_suggested_off_day_map,
)
from shifts.application.use_cases import GenerateMonthlyShift, ImportSkillMap
from shifts.domain.rest_patterns import (
    next_pattern_index_from_previous,
    previous_work_statuses,
    work_rest_pattern_from_text,
)
from shifts.infrastructure.importers import SkillMapFileReader, SkillMapReadError
from shifts.infrastructure.master_repository import DjangoMasterRepository
from shifts.infrastructure.models import (
    AvailabilityDay,
    AvailabilitySubmission,
    CompanyMembership,
    ConstraintType,
    GenerationWarning,
    ImportJob,
    IndividualConstraint,
    PreviousMonthShiftDay,
    ShiftAssignment,
    ShiftLeaveRequest,
    ShiftPeriod,
    SkillLevel,
    Staff,
    StaffSkill,
    WorkType,
)
from shifts.infrastructure.repositories import DjangoShiftRepository
from .company import admin_required, current_membership, staff_required
from .forms import (
    AvailabilityImportForm,
    BulkDesiredOffLimitForm,
    ConstraintForm,
    ConstraintTypeForm,
    GenerateForm,
    ImportForm,
    PreviousShiftImportForm,
    SkillLevelForm,
    StaffForm,
    WorkTypeForm,
)


def _month(value=None) -> date:
    if value:
        try:
            return date.fromisoformat(f"{value[:7]}-01")
        except ValueError:
            pass
    today = timezone.localdate()
    return today.replace(day=1)


def _add_months(month: date, offset: int) -> date:
    year = month.year + (month.month - 1 + offset) // 12
    month_number = (month.month - 1 + offset) % 12 + 1
    return month.replace(year=year, month=month_number, day=1)


def _next_month() -> date:
    return _add_months(timezone.localdate().replace(day=1), 1)


def _koyomi_api_url(month: date) -> str:
    # 祝日名は暦APIから取得する。
    # 取得した祝日名は _calendar_days() の holiday に入り、
    # submit.html / my_shift.html / shift_detail.html で表示される。
    day_count = calendar.monthrange(month.year, month.month)[1]
    params = {
        "mode": "d",
        "cnt": day_count,
        "targetyyyy": month.year,
        "targetmm": f"{month.month:02d}",
        "targetdd": "01",
    }
    return "https://koyomi.zingsystem.com/api/?" + urlencode(params)


def _holidays_for_month(month: date) -> dict[date, str]:
    # API通信に失敗してもシフト画面は表示したいので、失敗時は祝日なしで続行する。
    try:
        with urlopen(_koyomi_api_url(month), timeout=5) as response:
            data = json.load(response)
    except Exception:
        return {}

    holidays = {}
    for day_text, info in data.get("datelist", {}).items():
        holiday_name = info.get("holiday", "")
        if holiday_name:
            holidays[date.fromisoformat(day_text)] = holiday_name
    return holidays


def _calendar_days(month: date) -> list[dict]:
    # 3つの画面で共通利用する「日付表示用データ」をここで作る。
    #
    # - submit.html: スタッフのシフト希望提出カード
    # - my_shift.html: スタッフのシフト確認カード
    # - shift_detail.html: 管理者の月間シフト表ヘッダー/セル
    #
    # is_non_workday / is_saturday はテンプレートで class 名になり、
    # static/shifts/app.css の .non-workday / .saturday に干渉する。
    day_count = calendar.monthrange(month.year, month.month)[1]
    holidays = _holidays_for_month(month)
    days = []
    for number in range(1, day_count + 1):
        current = month.replace(day=number)
        holiday_name = holidays.get(current, "")
        days.append(
            {
                "number": number,
                "date": current,
                "holiday": holiday_name,
                "is_holiday": bool(holiday_name),
                "is_saturday": current.weekday() == 5,
                "is_weekend": current.weekday() >= 5,
                "is_non_workday": current.weekday() >= 5 or bool(holiday_name),
            }
        )
    return days


def _delete_confirmation(request, obj, label, cancel_url, success_url, on_deleted=None):
    if request.method == "POST":
        try:
            obj.delete()
        except ProtectedError:
            messages.error(
                request,
                f"{label}はシフトやスキル設定で使用中のため削除できません。先に関連データを削除するか、無効にしてください。",
            )
        else:
            if on_deleted:
                on_deleted()
            messages.success(request, f"{label}を削除しました。")
        return redirect(success_url)
    return render(
        request,
        "shifts/confirm_delete.html",
        {"target": obj, "label": label, "cancel_url": cancel_url},
    )


@login_required
def home(request):
    membership = current_membership(request)
    if not membership:
        return render(request, "shifts/no_company.html")
    if membership.role == CompanyMembership.Role.ADMIN:
        return redirect("manager_dashboard")
    messages.error(request, "このローカル版は管理者のみ利用できます。")
    return render(request, "shifts/no_company.html")


@login_required
@admin_required
def manager_dashboard(request):
    # ダッシュボードは ?month=YYYY-MM で表示月を切り替える。
    # templates/shifts/manager_dashboard.html の「翌月のダッシュボードへ」ボタンが
    # next_month を使ってこのviewに戻ってくる。
    month = _month(request.GET.get("month"))
    current_month = timezone.localdate().replace(day=1)
    prev_month = _add_months(month, -1)
    next_month = _add_months(month, 1)
    submission_target_filter = {
        "staff__company": request.company,
        "staff__active": True,
        "staff__is_employee": False,
    }
    submitted = AvailabilitySubmission.objects.filter(
        **submission_target_filter,
        month=month,
        status=AvailabilitySubmission.Status.SUBMITTED,
    ).count()
    staff_count = Staff.objects.filter(
        company=request.company,
        active=True,
        is_employee=False,
    ).count()
    rate = round(submitted / staff_count * 100) if staff_count else 0
    latest_period = ShiftPeriod.objects.filter(
        company=request.company, month=month
    ).first()
    pending_leave_requests = ShiftLeaveRequest.objects.filter(
        period__company=request.company,
        status=ShiftLeaveRequest.Status.PENDING,
    ).select_related("period", "staff", "work_type")
    context = {
        "month": month,
        "current_month": current_month,
        "prev_month": prev_month,
        "next_month": next_month,
        "staff_count": staff_count,
        "submitted_count": submitted,
        "submission_rate": rate,
        "missing_count": max(staff_count - submitted, 0),
        "warning_count": latest_period.warning_count if latest_period else 0,
        "shift_status": (
            latest_period.get_status_display() if latest_period else "未生成"
        ),
        "latest_period": latest_period,
        "pending_leave_request_count": pending_leave_requests.count(),
        "pending_leave_requests": pending_leave_requests[:8],
        "recent_periods": ShiftPeriod.objects.filter(company=request.company).annotate(
            assignment_count=Count("assignments")
        )[:5],
    }
    return render(request, "shifts/manager_dashboard.html", context)


@login_required
@admin_required
def missing_submissions(request):
    # シフト提出対象は、通常スタッフのみ。
    # 社員タグの人は後から補助投入する想定なので、提出率・未提出一覧から除外する。
    month = _month(request.GET.get("month"))
    current_month = timezone.localdate().replace(day=1)
    prev_month = _add_months(month, -1)
    next_month = _add_months(month, 1)
    import_form = AvailabilityImportForm(initial={"month": month})
    bulk_limit_form = BulkDesiredOffLimitForm(
        initial={"desired_off_limit": request.company.default_desired_off_limit}
    )
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "bulk_limit":
            bulk_limit_form = BulkDesiredOffLimitForm(request.POST)
            if bulk_limit_form.is_valid():
                limit = bulk_limit_form.cleaned_data["desired_off_limit"]
                request.company.default_desired_off_limit = limit
                request.company.save(update_fields=["default_desired_off_limit"])
                count = Staff.objects.filter(company=request.company).update(
                    desired_off_limit=limit
                )
                messages.success(
                    request,
                    f"公休・有休の申請上限を{limit}日に変更しました。（{count}名へ反映）",
                )
                return redirect(f"{reverse('missing_submissions')}?month={month:%Y-%m}")
        else:
            import_form = AvailabilityImportForm(request.POST, request.FILES)
        if action != "bulk_limit" and import_form.is_valid():
            month = import_form.cleaned_data["month"].replace(day=1)
            try:
                result = _import_availability_submissions(
                    request.company,
                    month,
                    import_form.cleaned_data["file"],
                )
            except ValueError as exc:
                import_form.add_error("file", str(exc))
            else:
                messages.success(
                    request,
                    f"{month.year}年{month.month}月の公休申請を{result['staff']}名分取り込みました。"
                    f"公休{result['preferred_off']}件、有給{result['paid_leave']}件を反映しました。",
                )
                return redirect(f"{reverse('missing_submissions')}?month={month:%Y-%m}")
    target_staff = list(
        Staff.objects.filter(
            company=request.company,
            active=True,
            is_employee=False,
        ).order_by("employee_number", "name")
    )
    submissions = {
        submission.staff_id: submission
        for submission in AvailabilitySubmission.objects.filter(
            staff__company=request.company,
            staff__active=True,
            staff__is_employee=False,
            month=month,
        )
    }
    submitted_staff_ids = {
        submission.staff_id
        for submission in submissions.values()
        if submission.status == AvailabilitySubmission.Status.SUBMITTED
    }
    missing_staff_rows = [
        {
            "staff": staff,
            "submission": submissions.get(staff.id),
        }
        for staff in target_staff
        if staff.id not in submitted_staff_ids
    ]

    return render(
        request,
        "shifts/missing_submissions.html",
        {
            "month": month,
            "current_month": current_month,
            "prev_month": prev_month,
            "next_month": next_month,
            "target_staff_count": len(target_staff),
            "submitted_count": len(submitted_staff_ids),
            "missing_staff_rows": missing_staff_rows,
            "missing_count": len(missing_staff_rows),
            "import_form": import_form,
            "bulk_limit_form": bulk_limit_form,
        },
    )


@login_required
@admin_required
def staff_manage(request):
    form = StaffForm()
    if request.method == "POST":
        form = StaffForm(request.POST)
        if form.is_valid():
            form.save_for_company(request.company)
            messages.success(request, "スタッフを登録しました。")
            return redirect("staff_manage")
    return render(
        request,
        "shifts/staff_manage.html",
        {
            "form": form,
            "items": Staff.objects.filter(company=request.company).select_related(
                "user"
            ),
        },
    )


@login_required
@admin_required
def staff_edit(request, pk):
    item = get_object_or_404(Staff, pk=pk, company=request.company)
    form = StaffForm(request.POST or None, instance=item)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "スタッフ情報を更新しました。")
        return redirect("staff_manage")
    return render(
        request,
        "shifts/staff_manage.html",
        {
            "form": form,
            "items": Staff.objects.filter(company=request.company).select_related(
                "user"
            ),
            "editing_item": item,
        },
    )


@login_required
@admin_required
def staff_delete(request, pk):
    item = get_object_or_404(Staff, pk=pk, company=request.company)
    user_id = item.user_id

    def delete_staff_membership():
        CompanyMembership.objects.filter(
            company=request.company,
            user_id=user_id,
            role=CompanyMembership.Role.STAFF,
        ).delete()

    return _delete_confirmation(
        request,
        item,
        f"スタッフ「{item.name}」",
        reverse("staff_manage"),
        "staff_manage",
        delete_staff_membership if user_id else None,
    )


@login_required
@admin_required
def staff_bulk_delete(request):
    staff_qs = Staff.objects.filter(company=request.company)
    staff_count = staff_qs.count()
    if request.method == "POST":
        user_ids = list(
            staff_qs.exclude(user=None).values_list("user_id", flat=True)
        )
        with transaction.atomic():
            # スタッフに紐づく作成済みシフトは、割当がStaff削除を保護するため先に削除する。
            ShiftPeriod.objects.filter(company=request.company).delete()
            CompanyMembership.objects.filter(
                company=request.company,
                user_id__in=user_ids,
                role=CompanyMembership.Role.STAFF,
            ).delete()
            deleted_count, _details = staff_qs.delete()
        messages.success(
            request,
            f"スタッフ情報を全削除しました。（関連データを含む削除件数：{deleted_count}件）",
        )
        return redirect("staff_manage")

    return render(
        request,
        "shifts/confirm_delete.html",
        {
            "label": f"スタッフ情報 全{staff_count}名分（作成済みシフト・提出データ・スキル・制約を含む）",
            "cancel_url": reverse("staff_manage"),
        },
    )


@login_required
@admin_required
def work_manage(request):
    form = WorkTypeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "業務を登録しました。")
        return redirect("work_manage")
    return render(
        request,
        "shifts/work_manage.html",
        {"form": form, "items": WorkType.objects.filter(company=request.company)},
    )


@login_required
@admin_required
def work_edit(request, pk):
    item = get_object_or_404(WorkType, pk=pk, company=request.company)
    form = WorkTypeForm(request.POST or None, instance=item)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "業務を更新しました。")
        return redirect("work_manage")
    return render(
        request,
        "shifts/work_manage.html",
        {
            "form": form,
            "items": WorkType.objects.filter(company=request.company),
            "editing_item": item,
        },
    )


@login_required
@admin_required
def work_delete(request, pk):
    item = get_object_or_404(WorkType, pk=pk, company=request.company)
    return _delete_confirmation(
        request, item, f"業務「{item.name}」", reverse("work_manage"), "work_manage"
    )


@login_required
@admin_required
def skill_manage(request):
    form = SkillLevelForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "スキル区分を登録しました。")
        return redirect("skill_manage")
    return render(
        request,
        "shifts/skill_manage.html",
        {"form": form, "items": SkillLevel.objects.filter(company=request.company)},
    )


@login_required
@admin_required
def skill_edit(request, pk):
    item = get_object_or_404(SkillLevel, pk=pk, company=request.company)
    form = SkillLevelForm(request.POST or None, instance=item)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "スキル区分を更新しました。")
        return redirect("skill_manage")
    return render(
        request,
        "shifts/skill_manage.html",
        {
            "form": form,
            "items": SkillLevel.objects.filter(company=request.company),
            "editing_item": item,
        },
    )


@login_required
@admin_required
def skill_delete(request, pk):
    item = get_object_or_404(SkillLevel, pk=pk, company=request.company)
    return _delete_confirmation(
        request,
        item,
        f"スキル区分「{item.symbol}」",
        reverse("skill_manage"),
        "skill_manage",
    )


def _constraint_items(company, query=""):
    # 勤務ルールの検索は、編集したいスタッフをすぐ見つけるために
    # 名前・社員番号だけを対象にする。
    items = IndividualConstraint.objects.filter(company=company).select_related(
        "staff", "related_staff", "rule_type", "work_type_a", "work_type_b"
    )
    if query:
        items = items.filter(
            Q(staff__name__icontains=query)
            | Q(staff__employee_number__icontains=query)
        )
    return items


@login_required
@admin_required
def constraint_manage(request):
    form = ConstraintForm(request.POST or None, company=request.company)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "勤務ルールを登録しました。")
        return redirect("constraint_manage")
    query = request.GET.get("q", "").strip()
    items = _constraint_items(request.company, query)
    return render(
        request,
        "shifts/constraint_manage.html",
        {
            "form": form,
            "items": items,
            "query": query,
            "has_rule_types": ConstraintType.objects.filter(
                company=request.company, active=True
            ).exists(),
        },
    )


@login_required
@admin_required
def constraint_edit(request, pk):
    item = get_object_or_404(IndividualConstraint, pk=pk, company=request.company)
    form = ConstraintForm(request.POST or None, instance=item, company=request.company)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "勤務ルールを更新しました。")
        return redirect("constraint_manage")
    query = request.GET.get("q", "").strip()
    items = _constraint_items(request.company, query)
    return render(
        request,
        "shifts/constraint_manage.html",
        {
            "form": form,
            "items": items,
            "editing_item": item,
            "query": query,
            "has_rule_types": ConstraintType.objects.filter(
                company=request.company, active=True
            ).exists(),
        },
    )


@login_required
@admin_required
def constraint_delete(request, pk):
    item = get_object_or_404(IndividualConstraint, pk=pk, company=request.company)
    return _delete_confirmation(
        request,
        item,
        f"勤務ルール「{item.name}」",
        reverse("constraint_manage"),
        "constraint_manage",
    )


@login_required
@admin_required
def constraint_type_manage(request):
    form = ConstraintTypeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "ルール種別を登録しました。")
        return redirect("constraint_type_manage")
    return render(
        request,
        "shifts/constraint_type_manage.html",
        {"form": form, "items": ConstraintType.objects.filter(company=request.company)},
    )


@login_required
@admin_required
def constraint_type_edit(request, pk):
    item = get_object_or_404(ConstraintType, pk=pk, company=request.company)
    form = ConstraintTypeForm(request.POST or None, instance=item)
    if request.method == "POST" and form.is_valid():
        form.save_for_company(request.company)
        messages.success(request, "ルール種別を更新しました。")
        return redirect("constraint_type_manage")
    return render(
        request,
        "shifts/constraint_type_manage.html",
        {
            "form": form,
            "items": ConstraintType.objects.filter(company=request.company),
            "editing_item": item,
        },
    )


@login_required
@admin_required
def constraint_type_delete(request, pk):
    item = get_object_or_404(ConstraintType, pk=pk, company=request.company)
    return _delete_confirmation(
        request,
        item,
        f"ルール種別「{item.name}」",
        reverse("constraint_type_manage"),
        "constraint_type_manage",
    )


@login_required
@admin_required
def skill_map(request):
    works = list(WorkType.objects.filter(company=request.company, active=True))
    skill_levels = list(SkillLevel.objects.filter(company=request.company))
    matrix_query = request.GET.get("matrix_q", "").strip()

    if request.method == "POST" and request.POST.get("action") == "update_matrix":
        staff_ids = {
            staff_id
            for staff_id in Staff.objects.filter(
                company=request.company, active=True
            ).values_list("id", flat=True)
        }
        work_ids = {work.id for work in works}
        level_ids = {level.id for level in skill_levels}
        updated_count = 0
        deleted_count = 0

        for staff_id in staff_ids:
            for work_id in work_ids:
                field_name = f"skill_{staff_id}_{work_id}"
                if field_name not in request.POST:
                    continue

                level_id = request.POST.get(field_name)
                current_skill = StaffSkill.objects.filter(
                    staff_id=staff_id,
                    staff__company=request.company,
                    work_type_id=work_id,
                )

                if not level_id:
                    deleted_count += current_skill.count()
                    current_skill.delete()
                    continue

                if not level_id.isdigit() or int(level_id) not in level_ids:
                    continue

                StaffSkill.objects.update_or_create(
                    staff_id=staff_id,
                    work_type_id=work_id,
                    defaults={"level_id": int(level_id)},
                )
                updated_count += 1

        messages.success(
            request,
            f"スキルマップを更新しました。（更新{updated_count}件・未設定{deleted_count}件）",
        )
        redirect_url = reverse("skill_map")
        if request.GET:
            redirect_url = f"{redirect_url}?{request.GET.urlencode()}"
        return redirect(redirect_url)

    staff_qs = Staff.objects.filter(company=request.company, active=True)
    if matrix_query:
        staff_qs = staff_qs.filter(
            Q(employee_number__icontains=matrix_query)
            | Q(name__icontains=matrix_query)
        )

    staff_rows = []
    for staff in staff_qs:
        levels = {
            item.work_type_id: item.level
            for item in staff.work_skills.select_related("level")
        }
        staff_rows.append(
            {
                "staff": staff,
                "cells": [
                    {
                        "field_name": f"skill_{staff.id}_{work.id}",
                        "level": levels.get(work.id),
                        "work": work,
                    }
                    for work in works
                ],
            }
        )
    return render(
        request,
        "shifts/skill_map.html",
        {
            "works": works,
            "staff_rows": staff_rows,
            "matrix_query": matrix_query,
            "skill_levels": skill_levels,
        },
    )


@login_required
@admin_required
def staff_skill_delete(request, pk):
    item = get_object_or_404(
        StaffSkill.objects.select_related("staff", "work_type"),
        pk=pk,
        staff__company=request.company,
    )
    return _delete_confirmation(
        request,
        item,
        f"{item.staff.name}の「{item.work_type.name}」スキル",
        reverse("skill_map"),
        "skill_map",
    )


@login_required
@admin_required
@require_POST
def staff_skill_bulk_delete(request):
    selected_ids = request.POST.getlist("skill_ids")
    selected_ids = [int(value) for value in selected_ids if value.isdigit()]
    items = (
        StaffSkill.objects.filter(pk__in=selected_ids, staff__company=request.company)
        .select_related("staff", "work_type", "level")
        .order_by("staff__employee_number", "work_type__display_order")
    )
    if not items.exists():
        messages.error(request, "削除するスキルを選択してください。")
        return redirect("skill_map")
    if request.POST.get("confirmed") == "1":
        count = items.count()
        items.delete()
        messages.success(request, f"スタッフスキルを{count}件削除しました。")
        return redirect("skill_map")
    return render(
        request,
        "shifts/bulk_delete_confirm.html",
        {"items": items, "selected_ids": selected_ids},
    )


def _skill_import_template_workbook(company):
    works = list(WorkType.objects.filter(company=company, active=True))
    skill_levels = list(SkillLevel.objects.filter(company=company))
    if not skill_levels:
        skill_levels = [
            {"symbol": "◎", "meaning": "リーダー", "priority": 1, "assignable": True},
            {"symbol": "○", "meaning": "対応可能", "priority": 2, "assignable": True},
            {"symbol": "△", "meaning": "訓練中", "priority": 3, "assignable": True},
            {"symbol": "×", "meaning": "対応不可", "priority": 99, "assignable": False},
        ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "スキル表"
    work_names = [work.name for work in works] or ["業務A", "業務B", "業務C"]
    example_work_names = work_names[:3]
    headers = ["社員番号", "氏名", "公休数", "備考", *example_work_names]
    sample_rows = [
        ["S001", "スタッフ01", 8, "4勤不可;単休不可", "◎", "○", "×"][: len(headers)],
        ["S002", "スタッフ02", 8, "業務Aと業務B交互;業務A連続不可", "○", "◎", "△"][: len(headers)],
        ["S003", "スタッフ03", 9, "業務B禁止", "△", "○", "◎"][: len(headers)],
    ]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for sample in sample_rows:
        sheet.append(sample)

    for cell in sheet["D"][1:]:
        cell.fill = note_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if len(headers) > 4:
        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=len(headers)):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="EAF4FF")
                cell.alignment = Alignment(horizontal="center")

    for column_index, header in enumerate(headers, start=1):
        width = 14
        if header == "氏名":
            width = 18
        elif header == "公休数":
            width = 12
        elif header == "備考":
            width = 42
        elif column_index >= 5:
            width = 14
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"

    level_sheet = workbook.create_sheet("スキル区分")
    level_headers = ["記号", "意味", "優先度", "アサイン可"]
    level_sheet.append(level_headers)
    for cell in level_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for index, level in enumerate(skill_levels, start=1):
        if isinstance(level, dict):
            symbol = level["symbol"]
            meaning = level["meaning"]
            priority = level["priority"]
            assignable = level["assignable"]
        else:
            symbol = level.symbol
            meaning = level.meaning
            priority = level.priority
            assignable = level.assignable
        level_sheet.append([symbol, meaning, priority, "可" if assignable else "不可"])
    for column_index, width in enumerate((12, 28, 12, 14), start=1):
        level_sheet.column_dimensions[get_column_letter(column_index)].width = width

    work_sheet = workbook.create_sheet("業務マスタ")
    work_headers = ["業務名", "必要人数", "有効"]
    work_sheet.append(work_headers)
    for cell in work_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    if works:
        for work in works:
            work_sheet.append(
                [
                    work.name,
                    work.required_staff_per_day,
                    "有効" if work.active else "無効",
                ]
            )
            _apply_work_name_fill(work_sheet.cell(row=work_sheet.max_row, column=1), work.color)
    else:
        for name, color in (("業務A", "#2563eb"), ("業務B", "#16a34a"), ("業務C", "#f97316")):
            work_sheet.append([name, 1, "有効"])
            _apply_work_name_fill(work_sheet.cell(row=work_sheet.max_row, column=1), color)

    for column_index, width in enumerate((24, 16, 12), start=1):
        work_sheet.column_dimensions[get_column_letter(column_index)].width = width

    _add_previous_shift_example_sheet(
        workbook,
        header_fill,
        header_font,
        [
            ["S001", "スタッフ01", "業務A", "業務B", "公休", "業務A", "有給", "業務B", "公休"],
            ["S002", "スタッフ02", "公休", "業務A", "業務B", "公休", "業務A", "業務B", "有給"],
            ["S003", "スタッフ03", "業務C", "公休", "業務A", "業務B", "公休", "業務C", "業務A"],
        ],
    )

    guide = workbook.create_sheet("入力ルール")
    guide_rows = [
        ["項目", "入力内容"],
        ["社員番号", "スタッフを照合するキーです。必須です。"],
        ["氏名", "スタッフ名です。必須です。"],
        ["公休数", "スタッフごとの月公休数です。スタッフ管理へ反映します。"],
        ["備考", "勤務ルールにしたい条件を書きます。複数ある場合は ; で区切れます。"],
        ["業務マスタ", "業務名・必要人数・有効を入力します。業務名セルの塗りつぶし色がシフト表の色として反映されます。"],
        ["業務列", "スキル表の公休数・備考の後ろには、業務マスタと同じ業務名を見出しとして追加します。"],
        ["スキル区分", "スキル区分シートの記号・意味・優先度・アサイン可を取込時に自動設定します。"],
        ["先月シフト実績", "マスタ取込後にこのシートを使って、前月の月末7日分の勤務・公休・有給を取り込めます。"],
        ["備考例", "ベース2勤1休 / 3勤1休 / 可能な限り4勤不可"],
        ["備考例", "4勤以上不可 / 4連勤不可 / 単休不可"],
        ["備考例", "業務Aと業務B交互 / 業務A連続不可 / 業務B禁止"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
    guide.column_dimensions["A"].width = 16
    guide.column_dimensions["B"].width = 72

    _add_skill_sheet_comments(sheet)
    _add_work_sheet_comments(work_sheet)
    _add_level_sheet_comments(level_sheet)

    return workbook


def _add_skill_sheet_comments(sheet):
    comments = {
        "A1": "スタッフを照合するキーです。",
        "B1": "スタッフ名です。",
        "C1": "スタッフごとの月公休数です。",
        "D1": "勤務ルールにしたい条件を書きます。例：2勤1休、単休不可、ロールとエーカス交互",
        "E1": "ここから右側が業務スキル欄です。業務マスタと同じ業務名を見出しにしてください。",
    }
    for address, text in comments.items():
        if sheet[address].value is not None:
            sheet[address].comment = Comment(text, "ShiftFlow")


def _add_work_sheet_comments(sheet):
    comments = {
        "A1": "スキル表の業務列・先月実績の勤務名と同じ名前にしてください。セルの塗りつぶし色がシフト表の色になります。",
        "B1": "1日に必要な人数です。",
        "C1": "有効/無効を書きます。",
    }
    for address, text in comments.items():
        sheet[address].comment = Comment(text, "ShiftFlow")


def _apply_work_name_fill(cell, color):
    color_code = _excel_color(color)
    if not color_code:
        return
    cell.fill = PatternFill("solid", fgColor=color_code)
    cell.font = Font(color=_excel_font_color_for_background(color_code), bold=True)


def _add_level_sheet_comments(sheet):
    comments = {
        "A1": "スキル表のセルに書くマークです。例：◎、○、△、×",
        "B1": "マークの意味です。",
        "C1": "数字が小さいほど優先されます。",
        "D1": "このマークのスタッフをアサイン可能にする場合は可、不可にする場合は不可。",
    }
    for address, text in comments.items():
        sheet[address].comment = Comment(text, "ShiftFlow")


def _add_previous_shift_example_sheet(workbook, header_fill, header_font, rows):
    sheet = workbook.create_sheet("先月シフト実績")
    headers = [
        "社員番号",
        "氏名",
        "2026/6/24",
        "2026/6/25",
        "2026/6/26",
        "2026/6/27",
        "2026/6/28",
        "2026/6/29",
        "2026/6/30",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=9):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for column_index, width in enumerate((14, 18, 12, 12, 12, 12, 12, 12, 12), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "C2"


SAMPLE_WORK_HEADERS = ["受付", "ロール", "エーカス", "検品", "出荷", "仕分け"]

SAMPLE_STAFF_ROWS = [
    [f"S{index:03}", f"スタッフ{index:02}", 8 + ((index - 1) % 4), note, *skills]
    for index, (note, skills) in enumerate(
        [
            ("ベース2勤1休;可能な限り4勤不可", ("◎", "○", "△", "○", "◎", "○")),
            ("ロールとエーカス交互;ロール連続不可", ("○", "◎", "◎", "△", "○", "○")),
            ("受付禁止", ("×", "○", "◎", "○", "△", "◎")),
            ("4勤以上不可", ("◎", "△", "○", "◎", "○", "△")),
            ("エーカス連続不可;単休不可", ("○", "◎", "○", "△", "◎", "○")),
            ("土日祝は公休", ("◎", "×", "○", "○", "◎", "△")),
            ("2勤1休", ("△", "◎", "○", "◎", "○", "○")),
            ("単休不可;可能な限り5勤不可", ("○", "○", "◎", "○", "△", "◎")),
            ("受付とロール交互;受付連続不可", ("◎", "◎", "△", "○", "○", "◎")),
            ("ベース3勤1休", ("○", "△", "◎", "◎", "○", "○")),
            ("4勤不可;単休不可", ("◎", "○", "×", "○", "◎", "△")),
            ("ロール禁止;土日祝は公休", ("○", "×", "◎", "△", "○", "◎")),
            ("2勤1休", ("△", "○", "◎", "◎", "○", "○")),
            ("受付連続不可;可能な限り4勤不可", ("◎", "○", "○", "△", "◎", "○")),
            ("エーカス禁止", ("○", "◎", "×", "○", "△", "◎")),
            ("土日祝は公休", ("○", "◎", "○", "◎", "○", "△")),
            ("単休不可", ("◎", "△", "○", "○", "◎", "○")),
            ("ロールとエーカス交互;ベース2勤1休", ("○", "◎", "◎", "△", "○", "○")),
            ("4勤不可", ("△", "○", "◎", "◎", "○", "○")),
            ("受付禁止", ("×", "◎", "○", "○", "△", "◎")),
            ("ベース3勤1休;可能な限り5勤不可", ("◎", "○", "○", "△", "◎", "○")),
            ("2勤1休;単休不可", ("○", "◎", "△", "◎", "○", "○")),
            ("ロール連続不可", ("◎", "○", "◎", "○", "△", "◎")),
            ("エーカス連続不可;土日祝は公休", ("○", "△", "◎", "◎", "○", "○")),
            ("4勤不可", ("◎", "◎", "○", "△", "○", "○")),
            ("ベース2勤1休", ("○", "○", "◎", "○", "◎", "△")),
            ("単休不可", ("△", "◎", "○", "◎", "○", "○")),
            ("受付とロール交互", ("◎", "◎", "△", "○", "○", "◎")),
            ("ロール禁止", ("○", "×", "◎", "△", "◎", "○")),
            ("2勤1休", ("◎", "○", "○", "○", "△", "◎")),
        ],
        start=1,
    )
]


def _skill_import_sample_workbook():
    # 実運用前にそのまま取り込んで試せるサンプル。
    # 対応画面: templates/shifts/import.html の「サンプルデータをダウンロード」
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    skill_fill = PatternFill("solid", fgColor="EAF4FF")

    sheet = workbook.active
    sheet.title = "スキル表"
    headers = ["社員番号", "氏名", "公休数", "備考", *SAMPLE_WORK_HEADERS]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in SAMPLE_STAFF_ROWS:
        sheet.append(row)

    for cell in sheet["D"][1:]:
        cell.fill = note_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=len(headers)):
        for cell in row:
            cell.fill = skill_fill
            cell.alignment = Alignment(horizontal="center")

    for column_index, width in enumerate((14, 18, 12, 42, 12, 12, 12, 12, 12, 12), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"

    level_sheet = workbook.create_sheet("スキル区分")
    level_sheet.append(["記号", "意味", "優先度", "アサイン可"])
    for cell in level_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in (
        ["◎", "主担当・指導可", 1, "可"],
        ["○", "対応可能", 2, "可"],
        ["△", "補助・訓練中", 3, "可"],
        ["×", "対応不可", 99, "不可"],
    ):
        level_sheet.append(row)
    for column_index, width in enumerate((12, 28, 12, 14), start=1):
        level_sheet.column_dimensions[get_column_letter(column_index)].width = width

    work_sheet = workbook.create_sheet("業務マスタ")
    work_sheet.append(["業務名", "必要人数", "有効"])
    for cell in work_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in (
        ["受付", 3, "有効", "#2563eb"],
        ["ロール", 3, "有効", "#16a34a"],
        ["エーカス", 3, "有効", "#f97316"],
        ["検品", 3, "有効", "#9333ea"],
        ["出荷", 3, "有効", "#dc2626"],
        ["仕分け", 3, "有効", "#0f766e"],
    ):
        work_sheet.append(row[:3])
        _apply_work_name_fill(work_sheet.cell(row=work_sheet.max_row, column=1), row[3])
    for column_index, width in enumerate((24, 16, 12), start=1):
        work_sheet.column_dimensions[get_column_letter(column_index)].width = width

    _add_previous_shift_example_sheet(
        workbook,
        header_fill,
        header_font,
        [
            ["S001", "スタッフ01", "受付", "公休", "公休", "ロール", "エーカス", "公休", "公休"],
            ["S002", "スタッフ02", "ロール", "エーカス", "公休", "ロール", "エーカス", "ロール", "エーカス"],
            ["S003", "スタッフ03", "公休", "ロール", "エーカス", "公休", "ロール", "エーカス", "ロール"],
            ["S004", "スタッフ04", "受付", "ロール", "エーカス", "公休", "受付", "ロール", "エーカス"],
            ["S005", "スタッフ05", "エーカス", "ロール", "公休", "公休", "受付", "エーカス", "ロール"],
            ["S006", "スタッフ06", "公休", "受付", "エーカス", "公休", "受付", "エーカス", "受付"],
            ["S007", "スタッフ07", "ロール", "公休", "受付", "エーカス", "公休", "ロール", "エーカス"],
            ["S008", "スタッフ08", "受付", "公休", "公休", "エーカス", "ロール", "公休", "公休"],
            ["S009", "スタッフ09", "受付", "ロール", "受付", "ロール", "公休", "受付", "ロール"],
            ["S010", "スタッフ10", "公休", "エーカス", "受付", "ロール", "公休", "受付", "ロール"],
            ["S011", "スタッフ11", "受付", "ロール", "公休", "エーカス", "受付", "公休", "ロール"],
            ["S012", "スタッフ12", "公休", "エーカス", "ロール", "受付", "公休", "エーカス", "受付"],
            ["S013", "スタッフ13", "ロール", "受付", "エーカス", "公休", "ロール", "受付", "公休"],
            ["S014", "スタッフ14", "受付", "公休", "ロール", "エーカス", "受付", "ロール", "公休"],
            ["S015", "スタッフ15", "エーカス", "ロール", "公休", "受付", "エーカス", "公休", "受付"],
            ["S016", "スタッフ16", "ロール", "受付", "エーカス", "ロール", "公休", "受付", "エーカス"],
            ["S017", "スタッフ17", "公休", "受付", "ロール", "公休", "エーカス", "受付", "ロール"],
            ["S018", "スタッフ18", "エーカス", "ロール", "受付", "エーカス", "ロール", "受付", "公休"],
            ["S019", "スタッフ19", "受付", "エーカス", "公休", "ロール", "受付", "エーカス", "ロール"],
            ["S020", "スタッフ20", "公休", "ロール", "受付", "エーカス", "公休", "ロール", "受付"],
            ["S021", "スタッフ21", "受付", "ロール", "エーカス", "受付", "ロール", "公休", "エーカス"],
            ["S022", "スタッフ22", "ロール", "公休", "受付", "ロール", "エーカス", "受付", "公休"],
            ["S023", "スタッフ23", "エーカス", "受付", "ロール", "公休", "エーカス", "受付", "ロール"],
            ["S024", "スタッフ24", "受付", "エーカス", "公休", "ロール", "受付", "公休", "エーカス"],
            ["S025", "スタッフ25", "ロール", "受付", "エーカス", "ロール", "受付", "エーカス", "公休"],
            ["S026", "スタッフ26", "公休", "エーカス", "受付", "公休", "ロール", "エーカス", "受付"],
            ["S027", "スタッフ27", "受付", "ロール", "公休", "エーカス", "受付", "ロール", "公休"],
            ["S028", "スタッフ28", "エーカス", "受付", "ロール", "エーカス", "公休", "受付", "ロール"],
            ["S029", "スタッフ29", "ロール", "公休", "エーカス", "受付", "ロール", "エーカス", "受付"],
            ["S030", "スタッフ30", "受付", "ロール", "エーカス", "公休", "受付", "ロール", "エーカス"],
        ],
    )

    guide = workbook.create_sheet("入力ルール")
    guide_rows = [
        ["項目", "入力内容"],
        ["このファイルの目的", "取込テスト用のサンプルです。スタッフ30人・業務6つ・1日18枠を登録できます。"],
        ["社員番号", "スタッフを照合するキーです。"],
        ["公休数", "スタッフごとの月公休数です。スタッフ管理へ反映します。"],
        ["備考", "勤務ルールへ自動変換される条件の例を入れています。不要なら空欄で問題ありません。"],
        ["業務マスタ", "業務名・必要人数・有効を業務管理へ反映します。業務色は業務名セルの塗りつぶし色から読み取ります。"],
        ["スキル表", "公休数・備考の後ろの業務名とセルの記号から、スタッフごとのスキルを登録します。"],
        ["スキル区分", "記号の意味・優先度・アサイン可否を登録します。"],
        ["先月シフト実績", "マスタ取込後にこのシートを使って、前月の月末7日分の勤務・公休・有給を取り込めます。"],
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 78

    _add_skill_sheet_comments(sheet)
    _add_work_sheet_comments(work_sheet)
    _add_level_sheet_comments(level_sheet)

    return workbook


@login_required
@admin_required
def download_import_template(request):
    workbook = _skill_import_template_workbook(request.company)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="shift_import_template.xlsx"'
    )
    return response


@login_required
@admin_required
def download_import_sample(request):
    workbook = _skill_import_sample_workbook()
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="shift_import_sample.xlsx"'
    return response


CSV_TEMPLATE_ROWS = [
    ["社員番号", "氏名", "公休数", "備考", "業務A", "業務B", "業務C"],
]


CSV_SAMPLE_ROWS = [
    ["社員番号", "氏名", "公休数", "備考", *SAMPLE_WORK_HEADERS],
    *SAMPLE_STAFF_ROWS,
]


def _csv_download_response(rows, filename):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerows(rows)
    return response


@login_required
@admin_required
def download_csv_template(request):
    return _csv_download_response(CSV_TEMPLATE_ROWS, "shift_import_template.csv")


@login_required
@admin_required
def download_csv_sample(request):
    return _csv_download_response(CSV_SAMPLE_ROWS, "shift_import_sample.csv")


PUBLIC_HOLIDAY_TOKENS = {"休", "公休", "公", "休日", "公休日"}
PAID_LEAVE_TOKENS = {"有休", "有給", "有", "年休"}
BLANK_SHIFT_TOKENS = {"", "-", "ー", "－", "未入力"}


def _previous_shift_day_from_header(header, month):
    if isinstance(header, datetime):
        return header.date()
    if isinstance(header, date):
        return header
    if isinstance(header, int) or (isinstance(header, float) and header.is_integer()):
        try:
            return month.replace(day=int(header))
        except ValueError:
            return None
    text = str(header or "").strip()
    if not text:
        return None
    if text.isdigit() or (text.endswith(".0") and text[:-2].isdigit()):
        try:
            return month.replace(day=int(float(text)))
        except ValueError:
            return None

    numbers = [int(value) for value in re.findall(r"\d+", text)]
    if len(numbers) == 1:
        try:
            return month.replace(day=numbers[0])
        except ValueError:
            return None
    if len(numbers) == 2:
        try:
            return date(month.year, numbers[0], numbers[1])
        except ValueError:
            return None
    if len(numbers) >= 3:
        try:
            return date(numbers[0], numbers[1], numbers[2])
        except ValueError:
            return None
    return None


def _previous_shift_rows_from_file(file_obj):
    suffix = Path(file_obj.name).suffix.lower()
    file_obj.seek(0)
    if suffix == ".csv":
        text = file_obj.read().decode("utf-8-sig")
        return list(csv.reader(StringIO(text)))
    if suffix == ".xlsx":
        workbook = load_workbook(file_obj, data_only=True)
        if "先月シフト実績" not in workbook.sheetnames:
            raise ValueError("Excel内に「先月シフト実績」シートが必要です。")
        sheet = workbook["先月シフト実績"]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError("先月シフト実績は .xlsx または .csv で取り込んでください。")


def _rows_from_excel_or_csv(file_obj, *, sheet_name=None, label="ファイル"):
    suffix = Path(file_obj.name).suffix.lower()
    file_obj.seek(0)
    if suffix == ".csv":
        text = file_obj.read().decode("utf-8-sig")
        return list(csv.reader(StringIO(text)))
    if suffix == ".xlsx":
        workbook = load_workbook(file_obj, data_only=True)
        sheet = (
            workbook[sheet_name]
            if sheet_name and sheet_name in workbook.sheetnames
            else workbook.active
        )
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError(f"{label}は .xlsx または .csv で取り込んでください。")


def _classify_previous_shift_cell(raw_value, work_map):
    value = str(raw_value or "").strip()
    compact = value.replace(" ", "").replace("　", "")
    if compact in BLANK_SHIFT_TOKENS:
        return PreviousMonthShiftDay.Status.BLANK, None
    if compact in PUBLIC_HOLIDAY_TOKENS:
        return PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY, None
    if compact in PAID_LEAVE_TOKENS:
        return PreviousMonthShiftDay.Status.PAID_LEAVE, None
    work = work_map.get(value) or work_map.get(compact)
    if work:
        return PreviousMonthShiftDay.Status.WORK, work
    raise ValueError(f"業務名「{value}」が業務管理に登録されていません。")


def _import_previous_shift_days(company, month, file_obj):
    rows = _previous_shift_rows_from_file(file_obj)
    if not rows:
        raise ValueError("取込対象の行がありません。")

    last_day_number = calendar.monthrange(month.year, month.month)[1]
    import_from = month.replace(day=last_day_number - 6)
    import_to = month.replace(day=last_day_number)
    headers = [str(value or "").strip() for value in rows[0]]
    if "社員番号" not in headers:
        raise ValueError("見出しに「社員番号」が必要です。")
    employee_index = headers.index("社員番号")
    day_columns = [
        (index, day)
        for index, header in enumerate(rows[0])
        if index != employee_index
        for day in [_previous_shift_day_from_header(header, month)]
        if day and import_from <= day <= import_to
    ]
    if not day_columns:
        raise ValueError(
            f"月末7日分の日付列が見つかりません。例：{month.month}/{import_from.day}〜{month.month}/{import_to.day}"
        )

    staff_map = {
        staff.employee_number: staff
        for staff in Staff.objects.filter(company=company, active=True)
    }
    work_map = {}
    for work in WorkType.objects.filter(company=company, active=True):
        work_map[work.name] = work
        work_map[work.name.replace(" ", "").replace("　", "")] = work

    imported_items = []
    skipped = 0
    for raw_row in rows[1:]:
        row = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
        employee_number = str(row[employee_index] or "").strip()
        if not employee_number:
            skipped += 1
            continue
        staff = staff_map.get(employee_number)
        if not staff:
            raise ValueError(f"社員番号「{employee_number}」のスタッフが登録されていません。")
        for index, day in day_columns:
            raw_value = row[index] if index < len(row) else ""
            status, work = _classify_previous_shift_cell(raw_value, work_map)
            imported_items.append(
                PreviousMonthShiftDay(
                    company=company,
                    staff=staff,
                    day=day,
                    status=status,
                    work_type=work,
                    raw_value=str(raw_value or "").strip(),
                )
            )

    PreviousMonthShiftDay.objects.filter(
        company=company, day__range=(import_from, import_to)
    ).delete()
    PreviousMonthShiftDay.objects.bulk_create(imported_items)
    return {
        "days": len(imported_items),
        "staff": len({item.staff_id for item in imported_items}),
        "skipped": skipped,
        "from": import_from.isoformat(),
        "to": import_to.isoformat(),
    }


def _availability_rows_from_file(file_obj):
    rows = _rows_from_excel_or_csv(file_obj, sheet_name="公休申請", label="公休申請")
    if not rows:
        raise ValueError("取込対象の行がありません。")
    return rows


def _employee_number_column(headers):
    candidates = ("社員番号", "従業員番号", "スタッフ番号", "コード", "ID", "id")
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate)
    raise ValueError("見出しに「社員番号」列が必要です。")


def _classify_availability_cell(raw_value):
    value = str(raw_value or "").strip()
    compact = value.replace(" ", "").replace("　", "")
    if compact in PAID_LEAVE_TOKENS:
        return "paid"
    if compact in PUBLIC_HOLIDAY_TOKENS:
        return "off"
    return "available"


def _import_availability_submissions(company, month, file_obj):
    rows = _availability_rows_from_file(file_obj)
    headers = [str(value or "").strip() for value in rows[0]]
    employee_index = _employee_number_column(headers)
    day_columns = [
        (index, day)
        for index, header in enumerate(rows[0])
        if index != employee_index
        for day in [_previous_shift_day_from_header(header, month)]
        if day and day.year == month.year and day.month == month.month
    ]

    staff_map = {
        staff.employee_number: staff
        for staff in Staff.objects.filter(company=company, active=True)
    }
    imported_staff_ids = set()
    preferred_off_count = 0
    paid_leave_count = 0
    over_limit_errors = []
    parsed_submissions = []

    for raw_row in rows[1:]:
        row = list(raw_row) + [""] * max(0, len(headers) - len(raw_row))
        employee_number = str(row[employee_index] or "").strip()
        if not employee_number:
            continue
        staff = staff_map.get(employee_number)
        if not staff:
            raise ValueError(f"社員番号「{employee_number}」のスタッフが登録されていません。")

        requested_days = []
        for index, day in day_columns:
            state = _classify_availability_cell(row[index] if index < len(row) else "")
            if state != "available":
                requested_days.append((day, state))
        if not requested_days:
            continue
        if len(requested_days) > staff.desired_off_limit:
            over_limit_errors.append(
                f"{staff.employee_number} {staff.name}: {len(requested_days)}日 / 上限{staff.desired_off_limit}日"
            )
            continue
        parsed_submissions.append((staff, requested_days))

    if over_limit_errors:
        raise ValueError(
            "公休・有休の申請上限を超えているスタッフがいます。"
            + " / ".join(over_limit_errors[:10])
        )
    if not parsed_submissions:
        raise ValueError("公休または有休が入力されているスタッフが見つかりませんでした。")

    for staff, requested_days in parsed_submissions:
        submission, _ = AvailabilitySubmission.objects.update_or_create(
            staff=staff,
            month=month,
            defaults={
                "status": AvailabilitySubmission.Status.SUBMITTED,
                "submitted_at": timezone.now(),
            },
        )
        submission.days.all().delete()
        imported_staff_ids.add(staff.id)

        for day, state in requested_days:
            preferred_off = state == "off"
            paid_leave = state == "paid"
            preferred_off_count += int(preferred_off)
            paid_leave_count += int(paid_leave)
            AvailabilityDay.objects.create(
                submission=submission,
                day=day,
                available=False,
                preferred_off=preferred_off,
                paid_leave=paid_leave,
            )

    return {
        "staff": len(imported_staff_ids),
        "preferred_off": preferred_off_count,
        "paid_leave": paid_leave_count,
    }


@login_required
@admin_required
def import_skill_map(request):
    form = ImportForm(request.POST or None, request.FILES or None)
    previous_shift_form = PreviousShiftImportForm()
    if request.method == "POST" and request.POST.get("action") == "previous_shift":
        previous_shift_form = PreviousShiftImportForm(request.POST, request.FILES)
        if previous_shift_form.is_valid():
            file = previous_shift_form.cleaned_data["file"]
            month = previous_shift_form.cleaned_data["month"].replace(day=1)
            try:
                result = _import_previous_shift_days(request.company, month, file)
            except ValueError as exc:
                previous_shift_form.add_error("file", str(exc))
            else:
                ImportJob.objects.create(
                    company=request.company,
                    uploaded_by=request.user,
                    filename=f"先月実績: {file.name}",
                    result=result,
                )
                messages.success(
                    request,
                    f"{month.year}年{month.month}月の月末7日分の先月シフト実績を{result['days']}件取り込みました。",
                )
                return redirect("import_skill_map")
    elif request.method == "POST" and form.is_valid():
        file = form.cleaned_data["file"]
        try:
            result = ImportSkillMap(
                SkillMapFileReader(), DjangoMasterRepository()
            ).execute(request.company.id, file.name, file)
        except (SkillMapReadError, ValueError) as exc:
            form.add_error("file", str(exc))
        else:
            ImportJob.objects.create(
                company=request.company,
                uploaded_by=request.user,
                filename=file.name,
                result=result,
            )
            messages.success(
                request,
                f"スタッフ{result['staff']}件、スキル{result['skills']}件を取り込みました。"
                f"スキル区分{result.get('levels', 0)}件を設定しました。"
                f"備考から勤務ルール{result.get('constraints', 0)}件を反映しました。",
            )
            return redirect("skill_map")
    return render(
        request,
        "shifts/import.html",
        {
            "form": form,
            "previous_shift_form": previous_shift_form,
            "jobs": ImportJob.objects.filter(company=request.company).order_by(
                "-created_at"
            )[:10],
        },
    )


@login_required
@admin_required
def previous_shift_list(request):
    if request.GET.get("month"):
        month = _month(request.GET.get("month"))
    else:
        month = _latest_previous_shift_month(request.company) or _add_months(
            timezone.localdate().replace(day=1), -1
        )

    first_day, last_day, days = _previous_shift_month_end_days(month)
    period = (
        ShiftPeriod.objects.filter(
            company=request.company,
            month=month,
            assignments__isnull=False,
        )
        .distinct()
        .first()
    )
    if period:
        display = _previous_shift_display_from_period(
            request.company, period, first_day, last_day, days
        )
    else:
        display = _previous_shift_display_from_import(
            request.company, first_day, last_day, days
        )

    return render(
        request,
        "shifts/previous_shift_list.html",
        {
            "month": month,
            "prev_month": _add_months(month, -1),
            "next_month": _add_months(month, 1),
            "first_day": first_day,
            "last_day": last_day,
            "days": days,
            **display,
        },
    )


def _latest_previous_shift_month(company):
    internal_month = (
        ShiftPeriod.objects.filter(company=company, assignments__isnull=False)
        .distinct()
        .order_by("-month")
        .values_list("month", flat=True)
        .first()
    )
    imported_day = (
        PreviousMonthShiftDay.objects.filter(company=company)
        .order_by("-day")
        .values_list("day", flat=True)
        .first()
    )
    imported_month = imported_day.replace(day=1) if imported_day else None
    months = [month for month in (internal_month, imported_month) if month]
    return max(months) if months else None


def _previous_shift_month_end_days(month):
    last_day_number = calendar.monthrange(month.year, month.month)[1]
    first_day = month.replace(day=last_day_number - 6)
    last_day = month.replace(day=last_day_number)
    days = [
        month.replace(day=number)
        for number in range(first_day.day, last_day.day + 1)
    ]
    return first_day, last_day, days


def _previous_shift_display_from_period(company, period, first_day, last_day, days):
    staff_list = list(
        Staff.objects.filter(company=company, active=True).order_by("employee_number")
    )
    assignment_map = {
        (item.staff_id, item.day): item
        for item in ShiftAssignment.objects.filter(
            period=period,
            day__range=(first_day, last_day),
        ).select_related("work_type")
    }
    paid_leave_days = {
        (item.submission.staff_id, item.day)
        for item in AvailabilityDay.objects.filter(
            submission__staff__company=company,
            submission__month=period.month,
            submission__status=AvailabilitySubmission.Status.SUBMITTED,
            paid_leave=True,
            day__range=(first_day, last_day),
        ).select_related("submission")
    }
    rows = []
    totals = defaultdict(int)
    for staff in staff_list:
        cells = []
        for day in days:
            assignment = assignment_map.get((staff.id, day))
            if assignment and assignment.work_type:
                status = PreviousMonthShiftDay.Status.WORK
                label = assignment.work_type.name
                raw_value = "作成済みシフト"
            elif (staff.id, day) in paid_leave_days:
                status = PreviousMonthShiftDay.Status.PAID_LEAVE
                label = "有給"
                raw_value = "作成済みシフト"
            else:
                status = PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY
                label = "公休"
                raw_value = "作成済みシフト"
            totals[status] += 1
            cells.append(
                {
                    "day": day,
                    "label": label,
                    "status": status,
                    "raw_value": raw_value,
                }
            )
        rows.append({"staff": staff, "cells": cells})

    return {
        "source": "internal",
        "source_label": "作成済みシフト表",
        "source_note": "この月のシフト表がアプリ内にあるため、Excel取込より優先して表示しています。",
        "item_count": len(staff_list) * len(days),
        "staff_count": len(staff_list),
        "latest_imported_at": None,
        "rows": rows,
        "totals": {
            "work": totals[PreviousMonthShiftDay.Status.WORK],
            "public_holiday": totals[PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY],
            "paid_leave": totals[PreviousMonthShiftDay.Status.PAID_LEAVE],
            "blank": 0,
        },
    }


def _previous_shift_display_from_import(company, first_day, last_day, days):
    items = list(
        PreviousMonthShiftDay.objects.filter(
            company=company,
            day__range=(first_day, last_day),
        )
        .select_related("staff", "work_type")
        .order_by("staff__employee_number", "day")
    )
    items_by_staff_day = {(item.staff_id, item.day): item for item in items}
    staff_list = []
    seen_staff_ids = set()
    for item in items:
        if item.staff_id in seen_staff_ids:
            continue
        seen_staff_ids.add(item.staff_id)
        staff_list.append(item.staff)

    rows = []
    totals = defaultdict(int)
    for staff in staff_list:
        cells = []
        for day in days:
            item = items_by_staff_day.get((staff.id, day))
            if not item:
                cells.append(
                    {
                        "day": day,
                        "label": "未取込",
                        "status": "missing",
                        "raw_value": "",
                    }
                )
                continue
            label = item.get_status_display()
            if item.status == PreviousMonthShiftDay.Status.WORK and item.work_type:
                label = item.work_type.name
            cells.append(
                {
                    "day": day,
                    "label": label,
                    "status": item.status,
                    "raw_value": item.raw_value,
                }
            )
            totals[item.status] += 1
        rows.append({"staff": staff, "cells": cells})

    latest_imported_at = max((item.imported_at for item in items), default=None)
    return {
        "source": "excel",
        "source_label": "Excel取込データ",
        "source_note": "作成済みシフト表がないため、Excelから取り込んだ先月実績を表示しています。",
        "item_count": len(items),
        "staff_count": len(staff_list),
        "latest_imported_at": latest_imported_at,
        "rows": rows,
        "totals": {
            "work": totals[PreviousMonthShiftDay.Status.WORK],
            "public_holiday": totals[PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY],
            "paid_leave": totals[PreviousMonthShiftDay.Status.PAID_LEAVE],
            "blank": totals[PreviousMonthShiftDay.Status.BLANK],
        },
    }


@login_required
@admin_required
def shift_manage(request):
    return render(
        request,
        "shifts/shift_manage.html",
        {
            "form": GenerateForm(initial={"month": _month()}),
            "periods": ShiftPeriod.objects.filter(company=request.company).annotate(
                assignment_count=Count("assignments")
            ),
        },
    )


@login_required
@admin_required
@require_POST
def generate_shift(request):
    form = GenerateForm(request.POST)
    if not form.is_valid():
        messages.error(request, "対象月を確認してください。")
        return redirect("shift_manage")
    month = form.cleaned_data["month"].replace(day=1)
    previous_month = _add_months(month, -1)
    previous_period = (
        ShiftPeriod.objects.filter(
            company=request.company,
            month=previous_month,
            assignments__isnull=False,
        )
        .distinct()
        .first()
    )
    if previous_period:
        previous_shift_count = previous_period.assignments.count()
        carryover_message = (
            f"作成済みの前月シフト{previous_shift_count}件を加味しました。"
        )
    else:
        previous_shift_count = PreviousMonthShiftDay.objects.filter(
            company=request.company,
            day__year=previous_month.year,
            day__month=previous_month.month,
        ).count()
        carryover_message = (
            f"Excel取込の先月実績{previous_shift_count}件を加味しました。"
            if previous_shift_count
            else "先月実績が未取込のため、当月データのみで作成しました。"
        )
    output = GenerateMonthlyShift(DjangoShiftRepository()).execute(
        request.company.id, month
    )
    period = ShiftPeriod.objects.get(pk=output.period_id, company=request.company)
    _refresh_period_constraint_warnings(period)
    period.refresh_from_db()
    messages.success(
        request,
        f"{output.assignment_count}件を配置しました。警告は{period.warning_count}件です。"
        f"{carryover_message}",
    )
    return redirect("shift_detail", pk=output.period_id)


def _period_rows(period, works=None):
    # 管理者シフト表専用の行データ。
    # days は表の横軸、rows はスタッフごとの縦軸。
    # row.cells[*].day の class 情報が shift_detail.html 経由で CSS に渡る。
    days = _calendar_days(period.month)
    works = list(
        works
        or WorkType.objects.filter(company=period.company, active=True).order_by(
            "display_order", "name"
        )
    )
    staff_list = list(
        Staff.objects.filter(company=period.company, active=True).order_by(
            "employee_number", "name"
        )
    )
    work_options_by_staff = _assignable_work_options_by_staff(
        period.company, staff_list, works
    )
    skill_map = {
        (item.staff_id, item.work_type_id): item
        for item in StaffSkill.objects.filter(
            staff__company=period.company,
            staff__in=staff_list,
            work_type__in=works,
        ).select_related("level")
    }
    assignment_map = {
        (item.staff_id, item.day.day): item
        for item in period.assignments.select_related("staff", "work_type")
    }
    leave_request_map = {
        (item.staff_id, item.day.day): item
        for item in period.leave_requests.exclude(
            status=ShiftLeaveRequest.Status.REJECTED
        ).select_related("staff", "work_type")
    }
    paid_leave_days = defaultdict(set)
    for item in AvailabilityDay.objects.filter(
        submission__staff__company=period.company,
        submission__month=period.month,
        paid_leave=True,
    ).select_related("submission"):
        paid_leave_days[item.submission.staff_id].add(item.day)
    requested_public_holiday_days = defaultdict(set)
    for item in AvailabilityDay.objects.filter(
        submission__staff__company=period.company,
        submission__month=period.month,
        preferred_off=True,
    ).select_related("submission"):
        requested_public_holiday_days[item.submission.staff_id].add(item.day)
    day_count = len(days)
    rows = []
    for staff in staff_list:
        work_options = work_options_by_staff[staff.id]
        work_option_ids = {work.id for work in work_options}
        staff_paid_leave_days = paid_leave_days[staff.id]
        staff_requested_public_holiday_days = requested_public_holiday_days[staff.id]
        cells = []
        for day in days:
            assignment = assignment_map.get((staff.id, day["number"]))
            assignment_skill = (
                skill_map.get((staff.id, assignment.work_type_id))
                if assignment and assignment.work_type_id
                else None
            )
            day_date = day["date"]
            rest_kind = ""
            rest_label = ""
            if not assignment:
                if day_date in staff_paid_leave_days:
                    rest_kind = "paid-leave"
                    rest_label = "有給"
                elif day_date in staff_requested_public_holiday_days:
                    rest_kind = "requested-public-holiday"
                    rest_label = "申請公休"
                else:
                    rest_kind = "inserted-public-holiday"
                    rest_label = "公休"
            cells.append(
                {
                    "day": day,
                    "assignment": assignment,
                    "leave_request": leave_request_map.get((staff.id, day["number"])),
                    "field_name": f"assignment_{staff.id}_{day_date:%Y%m%d}",
                    "rest_kind": rest_kind,
                    "rest_label": rest_label,
                    "assignment_work_is_option": (
                        not assignment or assignment.work_type_id in work_option_ids
                    ),
                    "assignment_is_trainee": (
                        bool(assignment_skill)
                        and _skill_level_is_trainee(assignment_skill.level)
                    ),
                    "assignment_is_instructor": (
                        bool(assignment_skill)
                        and _skill_level_is_instructor(assignment_skill.level)
                    ),
                }
            )
        work_count = sum(1 for cell in cells if cell["assignment"])
        rest_count = day_count - work_count
        paid_leave_count = sum(
            1
            for cell in cells
            if not cell["assignment"] and cell["day"]["date"] in staff_paid_leave_days
        )
        public_holiday_count = max(rest_count - paid_leave_count, 0)
        public_holiday_target = staff.monthly_public_holidays
        expected_total = work_count + public_holiday_count + paid_leave_count
        public_holiday_status = (
            "ok"
            if public_holiday_count == public_holiday_target
            else "under"
            if public_holiday_count < public_holiday_target
            else "over"
        )
        expected_total_status = (
            "ok"
            if expected_total == day_count
            else "under"
            if expected_total < day_count
            else "over"
        )
        rows.append(
            {
                "staff": staff,
                "work_options": work_options,
                "cells": cells,
                "work_count": work_count,
                "rest_count": rest_count,
                "public_holiday_count": public_holiday_count,
                "public_holiday_target": public_holiday_target,
                "public_holiday_status": public_holiday_status,
                "public_holiday_status_label": (
                    "不足" if public_holiday_status == "under" else "超過"
                    if public_holiday_status == "over"
                    else ""
                ),
                "paid_leave_count": paid_leave_count,
                "expected_total": expected_total,
                "expected_total_status": expected_total_status,
                "expected_total_status_label": (
                    "不足" if expected_total_status == "under" else "超過"
                    if expected_total_status == "over"
                    else ""
                ),
            }
        )
    return days, rows


def _previous_week_context_for_shift_detail(period, rows):
    # 管理者シフト表の左側に、前月末7日分を参考表示する。
    # 編集・Excel/CSV出力には含めない画面専用データ。
    previous_month = _add_months(period.month, -1)
    _first_day, _last_day, raw_days = _previous_shift_month_end_days(previous_month)
    day_context = {day["date"]: day for day in _calendar_days(previous_month)}
    previous_days = [day_context[day] for day in raw_days if day in day_context]

    work_by_id = {
        work.id: work
        for work in WorkType.objects.filter(company=period.company, active=True)
    }
    previous_items = DjangoShiftRepository().previous_shift_days_for_generation(
        period.company_id,
        period.month,
    )
    previous_map = {
        (item.staff_id, item.day): item
        for item in previous_items
    }

    for row in rows:
        staff = row["staff"]
        previous_cells = []
        for day in previous_days:
            item = previous_map.get((staff.id, day["date"]))
            work = work_by_id.get(item.work_id) if item and item.work_id else None
            status = item.status if item else PreviousMonthShiftDay.Status.BLANK
            if work:
                label = work.name
                cell_class = "assigned"
            elif status == PreviousMonthShiftDay.Status.PAID_LEAVE:
                label = "有給"
                cell_class = "rest paid-leave"
            elif status == PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY:
                label = "公休"
                cell_class = "rest inserted-public-holiday"
            else:
                label = "—"
                cell_class = "rest blank"
            previous_cells.append(
                {
                    "day": day,
                    "label": label,
                    "work": work,
                    "status": status,
                    "cell_class": cell_class,
                }
            )
        row["previous_cells"] = previous_cells

    return previous_days


def _assignable_work_options_by_staff(company, staff_list, works):
    disallowed_work_ids = defaultdict(set)
    for skill in StaffSkill.objects.filter(
        staff__company=company,
        staff__in=staff_list,
        work_type__in=works,
        level__assignable=False,
    ).select_related("level"):
        disallowed_work_ids[skill.staff_id].add(skill.work_type_id)

    return {
        staff.id: [
            work
            for work in works
            if work.id not in disallowed_work_ids[staff.id]
        ]
        for staff in staff_list
    }


def _attach_shift_statistics(rows, days, works):
    # シフト表の右側・下部に出す集計値を作る。
    # 右側: スタッフごとの業務別回数
    # 下部: 日ごとの業務別人数
    daily_stats = [
        {
            "work": work,
            "counts": [0 for _day in days],
            "total": 0,
        }
        for work in works
    ]
    daily_stat_map = {item["work"].id: item for item in daily_stats}

    for row in rows:
        counts = {work.id: 0 for work in works}
        for index, cell in enumerate(row["cells"]):
            assignment = cell["assignment"]
            if not assignment or not assignment.work_type_id:
                continue
            if assignment.work_type_id not in counts:
                continue
            counts[assignment.work_type_id] += 1
            daily_stat_map[assignment.work_type_id]["counts"][index] += 1
            daily_stat_map[assignment.work_type_id]["total"] += 1
        row["work_summary"] = [
            {
                "work": work,
                "count": counts[work.id],
            }
            for work in works
        ]

    return daily_stats


def _shift_export_day_label(day):
    # ダウンロード用の日付見出し。
    # 集計値は含めず、画面のシフト本体だけをファイル化する。
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    label = f"{day['number']}日({weekdays[day['date'].weekday()]})"
    if day["holiday"]:
        label = f"{label} {day['holiday']}"
    return label


def _shift_export_cell_label(staff, cell):
    assignment = cell["assignment"]
    if assignment and assignment.work_type:
        return assignment.work_type.name
    if staff.is_employee and cell["rest_kind"] != "paid-leave":
        return ""
    return cell["rest_label"] or "休"


def _shift_export_rows(period):
    days, rows = _period_rows(period)
    header = ["社員番号", "氏名"] + [_shift_export_day_label(day) for day in days]
    body = []
    for row in rows:
        staff = row["staff"]
        body.append(
            [
                staff.employee_number,
                staff.name,
                *[_shift_export_cell_label(staff, cell) for cell in row["cells"]],
            ]
        )
    return days, rows, header, body


def _excel_color(value):
    if not value:
        return None
    color = str(value).strip().replace("#", "")
    if len(color) == 6 and re.fullmatch(r"[0-9A-Fa-f]{6}", color):
        return color.upper()
    return None


def _excel_font_color_for_background(hex_color):
    if not hex_color:
        return "111827"
    red = int(hex_color[0:2], 16)
    green = int(hex_color[2:4], 16)
    blue = int(hex_color[4:6], 16)
    brightness = (red * 299 + green * 587 + blue * 114) / 1000
    return "111827" if brightness >= 165 else "FFFFFF"


def _shift_export_header_text(day):
    weekday = ["月", "火", "水", "木", "金", "土", "日"][day["date"].weekday()]
    return f"{day['holiday']}\n{weekday}" if day["holiday"] else weekday


def _write_shift_excel_sheet(sheet, period, days, rows):
    # Excelは印刷・配布用の見た目を優先する。
    # CSVは _shift_export_rows() の素データ形式を維持し、Excelだけ2段見出しにする。
    max_column = 2 + len(days)
    last_column = get_column_letter(max_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    sheet.cell(row=1, column=1, value=f"{period.month.year}年{period.month.month}月 シフト表")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    sheet.cell(
        row=2,
        column=1,
        value="凡例：土曜=青 / 日曜・祝日=赤 / 申請公休=赤字 / 有給=紫 / 業務色=業務管理の色",
    )

    sheet.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    sheet.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    sheet.cell(row=4, column=1, value="社員番号")
    sheet.cell(row=4, column=2, value="氏名")
    for column_index, day in enumerate(days, start=3):
        sheet.cell(row=4, column=column_index, value=day["number"])
        sheet.cell(row=5, column=column_index, value=_shift_export_header_text(day))

    for row_index, row in enumerate(rows, start=6):
        staff = row["staff"]
        sheet.cell(row=row_index, column=1, value=staff.employee_number)
        sheet.cell(row=row_index, column=2, value=staff.name)
        for column_index, cell_data in enumerate(row["cells"], start=3):
            sheet.cell(
                row=row_index,
                column=column_index,
                value=_shift_export_cell_label(staff, cell_data) or None,
            )

    _apply_shift_export_styles(sheet, period, days, rows)
    sheet.print_area = f"A1:{last_column}{sheet.max_row}"


def _apply_shift_export_styles(sheet, period, days, rows):
    title_fill = PatternFill("solid", fgColor="1E3A8A")
    note_fill = PatternFill("solid", fgColor="EFF6FF")
    staff_header_fill = PatternFill("solid", fgColor="E2E8F0")
    weekday_fill = PatternFill("solid", fgColor="F8FAFC")
    saturday_fill = PatternFill("solid", fgColor="DBEAFE")
    non_workday_fill = PatternFill("solid", fgColor="FEE2E2")
    requested_public_holiday_fill = PatternFill("solid", fgColor="FEE2E2")
    inserted_public_holiday_fill = PatternFill("solid", fgColor="F8FAFC")
    paid_leave_fill = PatternFill("solid", fgColor="EDE9FE")
    thin_gray = Side(style="thin", color="CBD5E1")
    medium_gray = Side(style="medium", color="94A3B8")
    table_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    max_column = sheet.max_column
    title_cell = sheet.cell(row=1, column=1)
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = center

    note_cell = sheet.cell(row=2, column=1)
    note_cell.font = Font(size=9, color="334155")
    note_cell.fill = note_fill
    note_cell.alignment = left

    for row_index in (4, 5):
        for cell in sheet[row_index]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = staff_header_fill
            cell.alignment = center
            cell.border = table_border

    for index, day in enumerate(days, start=3):
        date_cell = sheet.cell(row=4, column=index)
        weekday_cell = sheet.cell(row=5, column=index)
        date_cell.fill = weekday_fill
        weekday_cell.fill = weekday_fill
        if day["is_saturday"]:
            date_cell.fill = saturday_fill
            weekday_cell.fill = saturday_fill
            date_cell.font = Font(bold=True, color="1D4ED8")
            weekday_cell.font = Font(bold=True, color="1D4ED8")
        elif day["is_non_workday"]:
            date_cell.fill = non_workday_fill
            weekday_cell.fill = non_workday_fill
            date_cell.font = Font(bold=True, color="B91C1C")
            weekday_cell.font = Font(bold=True, color="B91C1C")

    for row_index, row in enumerate(rows, start=6):
        staff = row["staff"]
        for column_index in (1, 2):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.alignment = left if column_index == 2 else center
            cell.border = table_border
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
        for day_index, cell_data in enumerate(row["cells"], start=3):
            cell = sheet.cell(row=row_index, column=day_index)
            cell.alignment = center
            cell.border = table_border
            day = cell_data["day"]
            if day["is_saturday"]:
                cell.fill = PatternFill("solid", fgColor="F8FBFF")
            elif day["is_non_workday"]:
                cell.fill = PatternFill("solid", fgColor="FFF7F7")
            assignment = cell_data["assignment"]
            if assignment and assignment.work_type:
                work_color = _excel_color(assignment.work_type.color)
                if work_color:
                    cell.fill = PatternFill("solid", fgColor=work_color)
                    cell.font = Font(
                        color=_excel_font_color_for_background(work_color),
                        bold=True,
                    )
                else:
                    cell.font = Font(color="111827", bold=True)
                continue

            if staff.is_employee and cell_data["rest_kind"] != "paid-leave":
                continue

            if cell_data["rest_kind"] == "requested-public-holiday":
                cell.fill = requested_public_holiday_fill
                cell.font = Font(color="B4233A", bold=True)
            elif cell_data["rest_kind"] == "paid-leave":
                cell.fill = paid_leave_fill
                cell.font = Font(color="6D28D9", bold=True)
            elif cell_data["rest_kind"] == "inserted-public-holiday":
                cell.fill = inserted_public_holiday_fill

    for row_index in range(4, sheet.max_row + 1):
        for column_index in range(1, max_column + 1):
            sheet.cell(row=row_index, column=column_index).border = table_border

    for column_index in range(3, max_column + 1):
        sheet.cell(row=4, column=column_index).border = Border(
            left=thin_gray,
            right=thin_gray,
            top=medium_gray,
            bottom=thin_gray,
        )
        sheet.cell(row=5, column=column_index).border = Border(
            left=thin_gray,
            right=thin_gray,
            top=thin_gray,
            bottom=medium_gray,
        )

    sheet.freeze_panes = "C6"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 6
    sheet.row_dimensions[4].height = 22
    sheet.row_dimensions[5].height = 38
    for row_index in range(6, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 23
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 18
    for column_index in range(3, max_column + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 9.5

    sheet.print_title_rows = "1:5"
    sheet.print_title_cols = "A:B"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45
    sheet.page_margins.header = 0.2
    sheet.page_margins.footer = 0.2
    sheet.oddFooter.center.text = "Page &P / &N"


@login_required
@admin_required
def download_shift_csv(request, pk):
    period = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    _days, _rows, header, body = _shift_export_rows(period)

    response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
    response["Content-Disposition"] = (
        f'attachment; filename="shift_{period.month:%Y_%m}.csv"'
    )
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow([f"{period.month.year}年{period.month.month}月 シフト表"])
    writer.writerow(header)
    writer.writerows(body)
    return response


@login_required
@admin_required
def download_shift_excel(request, pk):
    period = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    days, rows, _header, _body = _shift_export_rows(period)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "シフト表"
    _write_shift_excel_sheet(sheet, period, days, rows)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="shift_{period.month:%Y_%m}.xlsx"'
    )
    return response


def _shift_edit_support(period, days, rows, works):
    # 下書き編集で崩れやすいポイントを検知する。
    # 対応HTML: templates/shifts/shift_detail.html の「下書き編集サポート」
    assignments = list(
        period.assignments.select_related("staff", "work_type").order_by(
            "day", "work_type__display_order", "staff__employee_number"
        )
    )
    skills = {
        (item.staff_id, item.work_type_id): item
        for item in StaffSkill.objects.filter(
            staff__company=period.company,
            work_type__company=period.company,
        ).select_related("level")
    }
    assignment_counts = defaultdict(int)
    for assignment in assignments:
        skill = skills.get((assignment.staff_id, assignment.work_type_id))
        if skill and _skill_level_is_trainee(skill.level):
            continue
        assignment_counts[(assignment.day, assignment.work_type_id)] += 1

    daily_shortages = []
    for day in days:
        for work in works:
            count = assignment_counts[(day["date"], work.id)]
            shortage = max(0, work.required_staff_per_day - count)
            if shortage:
                daily_shortages.append(
                    {
                        "day": day,
                        "work": work,
                        "count": count,
                        "required": work.required_staff_per_day,
                        "shortage": shortage,
                    }
                )

    availability = {
        (item.submission.staff_id, item.day): item
        for item in AvailabilityDay.objects.filter(
            submission__staff__company=period.company,
            submission__month=period.month,
        ).select_related("submission")
    }
    pending_leave_map = {
        (item.staff_id, item.day): item
        for item in ShiftLeaveRequest.objects.filter(
            period=period,
            status=ShiftLeaveRequest.Status.PENDING,
        )
    }

    assignment_issues = []
    for assignment in assignments:
        work_name = assignment.work_type.name if assignment.work_type else "業務未設定"
        pending_leave = pending_leave_map.get((assignment.staff_id, assignment.day))
        if pending_leave:
            assignment_issues.append(
                {
                    "level": "danger",
                    "day": assignment.day,
                    "staff": assignment.staff,
                    "message": f"{work_name}に急な休み申請が出ています。",
                }
            )

        request_day = availability.get((assignment.staff_id, assignment.day))
        if request_day and not request_day.available:
            assignment_issues.append(
                {
                    "level": "danger",
                    "day": assignment.day,
                    "staff": assignment.staff,
                    "message": f"{work_name}に入っていますが、勤務不可で提出されています。",
                }
            )
        elif request_day and request_day.paid_leave:
            assignment_issues.append(
                {
                    "level": "danger",
                    "day": assignment.day,
                    "staff": assignment.staff,
                    "message": f"{work_name}に入っていますが、有給希望の日です。",
                }
            )
        elif request_day and request_day.preferred_off:
            assignment_issues.append(
                {
                    "level": "warning",
                    "day": assignment.day,
                    "staff": assignment.staff,
                    "message": f"{work_name}に入っていますが、休み希望の日です。",
                }
            )

        skill = skills.get((assignment.staff_id, assignment.work_type_id))
        if not skill:
            if not assignment.staff.is_employee:
                assignment_issues.append(
                    {
                        "level": "warning",
                        "day": assignment.day,
                        "staff": assignment.staff,
                        "message": f"{work_name}のスキルが未設定です。",
                    }
                )
            continue
        if not skill.level.assignable:
            assignment_issues.append(
                {
                    "level": "danger",
                    "day": assignment.day,
                    "staff": assignment.staff,
                    "message": f"{work_name}はスキル区分「{skill.level.symbol}」のためアサイン不可です。",
                }
            )

    public_holiday_rows = []
    for row in rows:
        difference = row["public_holiday_count"] - row["public_holiday_target"]
        if difference == 0:
            continue
        public_holiday_rows.append(
            {
                "staff": row["staff"],
                "actual": row["public_holiday_count"],
                "target": row["public_holiday_target"],
                "difference": difference,
                "status": row["public_holiday_status"],
                "status_label": row["public_holiday_status_label"],
            }
        )
    public_holiday_has_issues = any(
        row["status"] != "ok" for row in public_holiday_rows
    )

    return {
        "daily_shortages": daily_shortages,
        "assignment_issues": assignment_issues,
        "public_holiday_rows": public_holiday_rows,
        "public_holiday_has_issues": public_holiday_has_issues,
        "has_issues": bool(
            daily_shortages or assignment_issues or public_holiday_has_issues
        ),
    }


def _skill_level_is_trainee(level) -> bool:
    text = f"{level.symbol} {level.meaning}".replace(" ", "").replace("　", "")
    return any(word in text for word in ("研修", "訓練"))


def _skill_level_is_instructor(level) -> bool:
    text = f"{level.symbol} {level.meaning}".replace(" ", "").replace("　", "")
    return any(word in text for word in ("指導", "教官", "主担当", "リーダー"))


def _refresh_period_constraint_warnings(period):
    GenerationWarning.objects.filter(
        period=period,
        message__startswith="制約違反：",
    ).delete()
    warnings = _constraint_violation_warnings(period)
    GenerationWarning.objects.bulk_create(
        [
            GenerationWarning(
                period=period,
                day=item["day"],
                work_type=item.get("work"),
                message=item["message"],
            )
            for item in warnings
        ]
    )
    period.warning_count = period.warnings.count()
    period.save(update_fields=["warning_count"])
    return len(warnings)


def _constraint_violation_warnings(period):
    assignments = list(
        period.assignments.select_related("staff", "work_type").order_by(
            "day", "staff__employee_number", "work_type__display_order"
        )
    )
    if not assignments:
        return []

    previous_days = sorted(
        DjangoShiftRepository().previous_shift_days_for_generation(
            period.company_id, period.month
        ),
        key=lambda item: (item.staff_id, item.day),
    )
    previous_days_by_staff = defaultdict(list)
    for item in previous_days:
        previous_days_by_staff[item.staff_id].append(item)

    previous_work_days = [
        item
        for item in previous_days
        if item.status == PreviousMonthShiftDay.Status.WORK and item.work_id
    ]
    previous_assignments = [
        {
            "staff_id": item.staff_id,
            "day": item.day,
            "work_id": item.work_id,
            "staff": None,
            "work": None,
        }
        for item in previous_work_days
    ]
    current_assignments = [
        {
            "staff_id": item.staff_id,
            "day": item.day,
            "work_id": item.work_type_id,
            "staff": item.staff,
            "work": item.work_type,
        }
        for item in assignments
        if item.work_type_id
    ]
    all_assignments = sorted(
        [*previous_assignments, *current_assignments],
        key=lambda item: (item["day"], item["staff_id"]),
    )
    current_by_staff = defaultdict(list)
    all_by_staff = defaultdict(list)
    current_by_day_work = defaultdict(list)
    for item in all_assignments:
        all_by_staff[item["staff_id"]].append(item)
        if item["day"].year == period.month.year and item["day"].month == period.month.month:
            current_by_staff[item["staff_id"]].append(item)
            current_by_day_work[(item["day"], item["work_id"])].append(item)

    rules = DjangoShiftRepository().rules_for_generation(period.company_id)
    warnings = []
    seen = set()
    active_rule = None

    def add_warning(day, staff, work, message):
        detail = _constraint_warning_detail(active_rule, staff, work, message)
        key = (day, staff.id if staff else None, work.id if work else None, detail)
        if key in seen:
            return
        seen.add(key)
        warnings.append(
            {
                "day": day,
                "work": work,
                "message": detail,
            }
        )

    for rule in rules:
        active_rule = rule
        if rule.operator == "incompatible_same_work":
            _append_incompatible_same_work_warnings(
                rule, current_by_day_work, add_warning
            )
            continue
        if rule.staff_id is None:
            continue
        staff_assignments = all_by_staff[rule.staff_id]
        current_staff_assignments = current_by_staff[rule.staff_id]
        if not current_staff_assignments:
            continue
        if rule.operator == "max_consecutive":
            _append_max_consecutive_warnings(rule, staff_assignments, add_warning)
        elif rule.operator == "work_alternation":
            _append_work_alternation_warnings(rule, staff_assignments, add_warning)
        elif rule.operator == "avoid_same_work":
            _append_avoid_same_work_warnings(rule, staff_assignments, add_warning)
        elif rule.operator == "avoid_specific_work":
            _append_avoid_specific_work_warnings(rule, staff_assignments, add_warning)
        elif rule.operator == "forbid_specific_work":
            _append_forbid_specific_work_warnings(
                rule, current_staff_assignments, add_warning
            )
        elif rule.operator == "forbid_works_on_weekdays":
            _append_forbid_works_on_weekdays_warnings(
                rule, current_staff_assignments, add_warning
            )
        elif rule.operator == "no_single_rest":
            _append_no_single_rest_warnings(
                rule, staff_assignments, period.month, add_warning
            )
        elif rule.operator == "work_rest_pattern":
            _append_work_rest_pattern_warnings(
                rule,
                current_staff_assignments,
                period.month,
                previous_days_by_staff[rule.staff_id],
                add_warning,
            )
    return warnings


def _constraint_warning_detail(rule, staff, work, message):
    # 管理者が修正に動けるよう、違反内容を「誰・業務・ルール・理由」に分けて表示する。
    detail = str(message or "").strip()
    for prefix in ("さんが", "さんの"):
        if detail.startswith(prefix):
            detail = detail.removeprefix(prefix)
            break

    staff_label = staff.name if staff else "全体"
    work_label = work.name if work else "業務なし"
    rule_label = ""
    contexts = []
    if rule:
        rule_label = rule.name or rule.rule_type_name or rule.operator
        if rule.related_staff_name:
            contexts.append(f"相手：{rule.related_staff_name}")
        if rule.work_names:
            contexts.append(f"対象業務：{'・'.join(rule.work_names)}")
        if rule.weekdays:
            weekday_labels = ["月", "火", "水", "木", "金", "土", "日"]
            contexts.append(
                "対象曜日：" + "・".join(weekday_labels[index] for index in rule.weekdays)
            )
    if not rule_label:
        rule_label = "勤務ルール"
    context_label = f"（{' / '.join(contexts)}）" if contexts else ""
    return (
        f"制約違反：スタッフ：{staff_label} / 業務：{work_label} / "
        f"ルール：{rule_label}{context_label} / 違反内容：{detail}"
    )


def _append_incompatible_same_work_warnings(rule, current_by_day_work, add_warning):
    pair = {rule.staff_id, rule.related_staff_id}
    if None in pair:
        return
    for (day, work_id), items in current_by_day_work.items():
        if rule.work_ids and work_id not in rule.work_ids:
            continue
        staff_ids = {item["staff_id"] for item in items}
        if pair.issubset(staff_ids):
            target = next(item for item in items if item["staff_id"] == rule.staff_id)
            add_warning(day, target["staff"], target["work"], "さんが同時配置禁止の相手と同じ業務に入っています。")


def _append_max_consecutive_warnings(rule, staff_assignments, add_warning):
    limit = rule.numeric_value or 6
    streak = []
    previous_day = None
    for item in staff_assignments:
        if previous_day and item["day"] == previous_day + timedelta(days=1):
            streak.append(item)
        else:
            streak = [item]
        previous_day = item["day"]
        if len(streak) > limit and item["staff"]:
            add_warning(item["day"], item["staff"], item["work"], f"さんが最大連続勤務{limit}日を超えています。")


def _append_work_alternation_warnings(rule, staff_assignments, add_warning):
    filtered = [item for item in staff_assignments if item["work_id"] in rule.work_ids]
    for previous, current in zip(filtered, filtered[1:]):
        if current["work_id"] == previous["work_id"] and current["staff"]:
            add_warning(current["day"], current["staff"], current["work"], "さんの交互配置が崩れています。")


def _append_avoid_same_work_warnings(rule, staff_assignments, add_warning):
    for previous, current in zip(staff_assignments, staff_assignments[1:]):
        if current["work_id"] == previous["work_id"] and current["staff"]:
            add_warning(current["day"], current["staff"], current["work"], "さんが同じ業務に連続で入っています。")


def _append_avoid_specific_work_warnings(rule, staff_assignments, add_warning):
    for previous, current in zip(staff_assignments, staff_assignments[1:]):
        if (
            current["work_id"] == previous["work_id"]
            and current["work_id"] in rule.work_ids
            and current["staff"]
        ):
            add_warning(current["day"], current["staff"], current["work"], "さんが連続回避対象の業務に続けて入っています。")


def _append_forbid_specific_work_warnings(rule, current_staff_assignments, add_warning):
    for item in current_staff_assignments:
        if item["work_id"] in rule.work_ids and item["staff"]:
            add_warning(item["day"], item["staff"], item["work"], "さんが禁止業務に入っています。")


def _append_forbid_works_on_weekdays_warnings(rule, current_staff_assignments, add_warning):
    for item in current_staff_assignments:
        if item["day"].weekday() not in rule.weekdays:
            continue
        if rule.work_ids and item["work_id"] not in rule.work_ids:
            continue
        if item["staff"]:
            add_warning(item["day"], item["staff"], item["work"], "さんが禁止曜日の業務に入っています。")


def _append_no_single_rest_warnings(rule, staff_assignments, month, add_warning):
    worked_days = {item["day"] for item in staff_assignments}
    staff_item = next((item for item in staff_assignments if item["staff"]), None)
    if not staff_item:
        return
    start = min(worked_days)
    end = max(worked_days)
    current = start + timedelta(days=1)
    while current < end:
        if (
            current.year == month.year
            and current.month == month.month
            and current not in worked_days
            and current - timedelta(days=1) in worked_days
            and current + timedelta(days=1) in worked_days
        ):
            add_warning(current, staff_item["staff"], None, "さんが単休になっています。")
        current += timedelta(days=1)


def _append_work_rest_pattern_warnings(
    rule, current_staff_assignments, month, previous_days, add_warning
):
    pattern = work_rest_pattern_from_text(rule.text_value)
    if not pattern:
        return
    pattern_index = next_pattern_index_from_previous(
        pattern,
        previous_work_statuses(previous_days),
    )
    rest_days = set()
    day_count = calendar.monthrange(month.year, month.month)[1]
    for day_number in range(1, day_count + 1):
        current = month.replace(day=day_number)
        if not pattern[pattern_index % len(pattern)]:
            rest_days.add(current)
        pattern_index += 1

    for item in current_staff_assignments:
        if item["day"] in rest_days and item["staff"]:
            add_warning(item["day"], item["staff"], item["work"], "さんが勤務・休みパターンの休み日に入っています。")

def _replacement_candidates(period, leave_request):
    # 急な休み申請の代替候補。
    # すでに同日に勤務が入っている人、勤務不可の人、対象業務にアサイン不可の人は候補から外す。
    if not leave_request.work_type:
        return []

    assigned_staff_ids = set(
        ShiftAssignment.objects.filter(period=period, day=leave_request.day).values_list(
            "staff_id", flat=True
        )
    )
    unavailable_staff_ids = set(
        AvailabilityDay.objects.filter(
            submission__staff__company=period.company,
            submission__month=period.month,
            day=leave_request.day,
        ).filter(
            Q(available=False) | Q(preferred_off=True) | Q(paid_leave=True)
        ).values_list("submission__staff_id", flat=True)
    )
    assignable_skill_staff_ids = set(
        StaffSkill.objects.filter(
            staff__company=period.company,
            work_type=leave_request.work_type,
            level__assignable=True,
        ).values_list("staff_id", flat=True)
    )
    unassignable_skill_staff_ids = set(
        StaffSkill.objects.filter(
            staff__company=period.company,
            work_type=leave_request.work_type,
            level__assignable=False,
        ).values_list("staff_id", flat=True)
    )
    monthly_work_counts = dict(
        ShiftAssignment.objects.filter(period=period)
        .values_list("staff_id")
        .annotate(count=Count("id"))
    )
    candidates = []
    for staff in Staff.objects.filter(company=period.company, active=True):
        if staff.id == leave_request.staff_id:
            continue
        if staff.id in assigned_staff_ids:
            continue
        if staff.id in unavailable_staff_ids:
            continue
        if staff.id in unassignable_skill_staff_ids:
            continue
        if staff.id not in assignable_skill_staff_ids and not staff.is_employee:
            continue
        candidates.append(
            {
                "staff": staff,
                "work_count": monthly_work_counts.get(staff.id, 0),
            }
        )
    return sorted(candidates, key=lambda item: (item["work_count"], item["staff"].employee_number))


def _leave_request_context(period):
    requests = list(
        period.leave_requests.select_related("staff", "work_type").order_by(
            "status", "day", "staff__employee_number"
        )
    )
    pending = []
    for item in requests:
        if item.status == ShiftLeaveRequest.Status.PENDING:
            pending.append(
                {
                    "request": item,
                    "candidates": _replacement_candidates(period, item),
                }
            )
    return {
        "leave_requests": requests,
        "pending_leave_requests": pending,
    }


@login_required
@admin_required
def shift_detail(request, pk):
    # 管理者の「シフト表」画面。
    # templates/shifts/shift_detail.html に days / rows を渡す。
    period = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    works = list(WorkType.objects.filter(company=request.company, active=True))
    days, rows = _period_rows(period, works)
    previous_days = _previous_week_context_for_shift_detail(period, rows)
    daily_work_stats = _attach_shift_statistics(rows, days, works)
    can_edit_shift = period.status in {
        ShiftPeriod.Status.DRAFT,
        ShiftPeriod.Status.PUBLISHED,
    }
    leave_context = _leave_request_context(period)
    show_shift_support = can_edit_shift or bool(leave_context["leave_requests"])
    constraint_warnings = period.warnings.filter(
        message__startswith="制約違反："
    ).select_related("work_type")
    return render(
        request,
        "shifts/shift_detail.html",
        {
            "period": period,
            "days": days,
            "previous_days": previous_days,
            "rows": rows,
            "works": works,
            "daily_work_stats": daily_work_stats,
            "can_edit_shift": can_edit_shift,
            "constraint_warning_count": constraint_warnings.count(),
            "show_shift_support": show_shift_support,
            "edit_support": _shift_edit_support(period, days, rows, works),
            **leave_context,
            "warnings": constraint_warnings,
        },
    )


@login_required
@admin_required
@require_POST
def update_shift_draft(request, pk):
    period = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    if period.status not in {
        ShiftPeriod.Status.DRAFT,
        ShiftPeriod.Status.PUBLISHED,
    }:
        messages.error(request, "まだ作成されていないシフトは編集できません。")
        return redirect("shift_detail", pk=pk)

    works = {
        str(work.id): work
        for work in WorkType.objects.filter(company=request.company, active=True)
    }
    staff_list = list(Staff.objects.filter(company=request.company, active=True))
    disallowed_work_ids = defaultdict(set)
    for skill in StaffSkill.objects.filter(
        staff__company=request.company,
        staff__in=staff_list,
        work_type_id__in=[work.id for work in works.values()],
        level__assignable=False,
    ).select_related("level"):
        disallowed_work_ids[skill.staff_id].add(str(skill.work_type_id))
    day_count = calendar.monthrange(period.month.year, period.month.month)[1]
    days = [period.month.replace(day=number) for number in range(1, day_count + 1)]
    changed_count = 0
    skipped_count = 0

    for staff in staff_list:
        for current in days:
            field_name = f"assignment_{staff.id}_{current:%Y%m%d}"
            if field_name not in request.POST:
                continue

            selected_work_id = request.POST.get(field_name, "")
            assignment = ShiftAssignment.objects.filter(
                period=period,
                staff=staff,
                day=current,
            ).first()

            if not selected_work_id:
                if assignment:
                    assignment.delete()
                    changed_count += 1
                continue

            work = works.get(selected_work_id)
            if not work:
                continue
            if selected_work_id in disallowed_work_ids[staff.id]:
                skipped_count += 1
                continue

            if assignment:
                # 既存割当は、実際に業務を変更して保存した時だけ「手動」にする。
                # 同じ内容で保存しただけなら、自動生成の表示を維持する。
                if assignment.work_type_id != work.id:
                    assignment.work_type = work
                    assignment.manually_edited = True
                    assignment.save(update_fields=["work_type", "manually_edited"])
                    changed_count += 1
            else:
                ShiftAssignment.objects.create(
                    period=period,
                    staff=staff,
                    day=current,
                    work_type=work,
                    manually_edited=True,
                )
                changed_count += 1

    constraint_warning_count = _refresh_period_constraint_warnings(period)
    message = f"シフトを保存しました。（変更{changed_count}件）"
    if skipped_count:
        message += f" 不可スキルの業務{skipped_count}件は反映しませんでした。"
    if constraint_warning_count:
        message += f" 制約違反の警告{constraint_warning_count}件を確認してください。"
    messages.success(request, message)
    return redirect("shift_detail", pk=pk)


@login_required
@admin_required
@require_POST
def resolve_leave_request(request, pk):
    leave_request = get_object_or_404(
        ShiftLeaveRequest.objects.select_related("period", "staff", "work_type"),
        pk=pk,
        period__company=request.company,
    )
    period = leave_request.period
    action = request.POST.get("action")

    if leave_request.status != ShiftLeaveRequest.Status.PENDING:
        messages.error(request, "この休み申請はすでに対応済みです。")
        return redirect("shift_detail", pk=period.pk)

    if action == "reject":
        leave_request.status = ShiftLeaveRequest.Status.REJECTED
        leave_request.admin_note = request.POST.get("admin_note", "")
        leave_request.resolved_at = timezone.now()
        leave_request.resolved_by = request.user
        leave_request.save(
            update_fields=["status", "admin_note", "resolved_at", "resolved_by"]
        )
        messages.success(request, "急な休み申請を却下しました。")
        return redirect("shift_detail", pk=period.pk)

    if action != "approve":
        messages.error(request, "対応内容を選択してください。")
        return redirect("shift_detail", pk=period.pk)

    assignment = ShiftAssignment.objects.filter(
        period=period,
        staff=leave_request.staff,
        day=leave_request.day,
    ).first()
    work_type = leave_request.work_type or (assignment.work_type if assignment else None)
    replacement_staff_id = request.POST.get("replacement_staff")
    replacement_staff = None
    if replacement_staff_id:
        replacement_staff = Staff.objects.filter(
            pk=replacement_staff_id,
            company=request.company,
            active=True,
        ).first()
        if not replacement_staff or replacement_staff == leave_request.staff:
            messages.error(request, "代替スタッフを確認してください。")
            return redirect("shift_detail", pk=period.pk)
        if ShiftAssignment.objects.filter(
            period=period,
            staff=replacement_staff,
            day=leave_request.day,
        ).exists():
            messages.error(request, "選択した代替スタッフは同日にすでに勤務があります。")
            return redirect("shift_detail", pk=period.pk)
        valid_candidate_ids = {
            item["staff"].id for item in _replacement_candidates(period, leave_request)
        }
        if replacement_staff.id not in valid_candidate_ids:
            messages.error(request, "選択した代替スタッフは条件に合いません。")
            return redirect("shift_detail", pk=period.pk)

    if assignment:
        assignment.delete()

    if replacement_staff and work_type:
        ShiftAssignment.objects.create(
            period=period,
            staff=replacement_staff,
            day=leave_request.day,
            work_type=work_type,
            note=f"{leave_request.staff.name}さん急休の代替",
            manually_edited=True,
        )

    leave_request.status = ShiftLeaveRequest.Status.APPROVED
    leave_request.admin_note = request.POST.get("admin_note", "")
    leave_request.resolved_at = timezone.now()
    leave_request.resolved_by = request.user
    leave_request.save(
        update_fields=["status", "admin_note", "resolved_at", "resolved_by"]
    )
    if replacement_staff:
        messages.success(
            request,
            f"急な休み申請を承認し、{replacement_staff.name}さんを代替配置しました。",
        )
    else:
        messages.warning(request, "急な休み申請を承認しました。代替は未配置です。")
    return redirect("shift_detail", pk=period.pk)


@login_required
@admin_required
@require_POST
def publish_shift(request, pk):
    period = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    period.status = ShiftPeriod.Status.PUBLISHED
    period.published_at = timezone.now()
    period.save(update_fields=["status", "published_at"])
    messages.success(request, "シフトを公開しました。")
    return redirect("shift_detail", pk=pk)


@login_required
@admin_required
def shift_delete(request, pk):
    item = get_object_or_404(ShiftPeriod, pk=pk, company=request.company)
    return _delete_confirmation(
        request,
        item,
        f"{item.month:%Y年%m月}のシフト",
        reverse("shift_detail", kwargs={"pk": pk}),
        "shift_manage",
    )


def _previous_shift_days_for_staff(staff, month):
    previous_days = DjangoShiftRepository().previous_shift_days_for_generation(
        staff.company_id, month
    )
    return sorted(
        [item for item in previous_days if item.staff_id == staff.id],
        key=lambda item: item.day,
    )


def _rest_rule_inputs_for_staff(staff):
    rules = IndividualConstraint.objects.filter(
        company=staff.company,
        staff=staff,
        active=True,
    ).select_related("rule_type")
    return [
        RestRuleInput(
            name=rule.name,
            operator=rule.rule_type.operator if rule.rule_type else "",
            kind=rule.kind,
            numeric_value=rule.numeric_value,
            text_value=rule.text_value,
            strength=rule.strength,
            strength_label=rule.strength_label,
            strength_class=rule.strength_class,
        )
        for rule in rules
    ]


def _suggested_off_day_map(staff, month):
    # スタッフの勤務ルールと前月末の実績から、提出画面の休み希望候補を作る。
    # 対応HTML: templates/shifts/submit.html の「候補から休み希望を自動入力」
    return build_suggested_off_day_map(
        month,
        _rest_rule_inputs_for_staff(staff),
        _previous_shift_days_for_staff(staff, month),
    )


def _suggested_off_days_from_constraints(staff, month):
    return set(_suggested_off_day_map(staff, month))


def _limited_suggested_off_days(suggested_off_days, limit):
    if not limit:
        return set()
    return set(sorted(suggested_off_days)[:limit])


def _staff_rest_constraint_notes(staff):
    # シフト提出時にスタッフ本人へ見せる「休み方」だけの勤務ルール。
    # 業務交互・特定業務禁止など、管理者側の割当判断に近い制約はここでは出さない。
    return build_staff_rest_constraint_notes(_rest_rule_inputs_for_staff(staff))


@login_required
@staff_required
def submit_availability(request):
    # スタッフの「シフト提出」画面。
    # _calendar_days() の日付情報に、提出済み状態(state)を足して submit.html に渡す。
    requested_month = request.POST.get("month") or request.GET.get("month")
    month = _month(requested_month) if requested_month else _next_month()
    submission, _ = AvailabilitySubmission.objects.get_or_create(
        staff=request.staff, month=month
    )
    day_count = calendar.monthrange(month.year, month.month)[1]
    if request.method == "POST":
        requested_off_count = sum(
            1
            for number in range(1, day_count + 1)
            if request.POST.get(f"day_{number}", "available") in {"off", "paid"}
        )
        if requested_off_count > request.staff.desired_off_limit:
            messages.error(
                request,
                f"公休希望と有給希望は合計{request.staff.desired_off_limit}日までです。",
            )
            return redirect(f"{request.path}?month={month:%Y-%m}")
        for number in range(1, day_count + 1):
            state = request.POST.get(f"day_{number}", "available")
            AvailabilityDay.objects.update_or_create(
                submission=submission,
                day=month.replace(day=number),
                defaults={
                    "available": True,
                    "preferred_off": state == "off",
                    "paid_leave": state == "paid",
                },
            )
        submission.status = AvailabilitySubmission.Status.SUBMITTED
        submission.submitted_at = timezone.now()
        submission.save(update_fields=["status", "submitted_at"])
        messages.success(request, f"{month.year}年{month.month}月分を提出しました。")
        return redirect(f"{request.path}?month={month:%Y-%m}")
    saved = {
        item.day.day: (
            "paid" if item.paid_leave else "off" if item.preferred_off else "available"
        )
        for item in submission.days.all()
    }
    suggested_off_day_map = _suggested_off_day_map(request.staff, month)
    suggested_off_days = set(suggested_off_day_map)
    auto_fill_off_days = _limited_suggested_off_days(
        suggested_off_days,
        request.staff.desired_off_limit,
    )
    apply_constraints = request.GET.get("apply_constraints") == "1"

    days = []
    requested_off_count = 0
    for day in _calendar_days(month):
        suggested_state = "off" if day["number"] in auto_fill_off_days else "available"
        state = suggested_state if apply_constraints else saved.get(
            day["number"], suggested_state
        )
        if state in {"off", "paid"}:
            requested_off_count += 1
        suggestion_reasons = suggested_off_day_map.get(day["number"], [])
        days.append(
            {
                **day,
                "suggested_off": day["number"] in suggested_off_days,
                "auto_fill_off": day["number"] in auto_fill_off_days,
                "suggestion_reasons": suggestion_reasons,
                "suggestion_label": "・".join(suggestion_reasons),
                "state": state,
            }
        )
    return render(
        request,
        "shifts/submit.html",
        {
            "month": month,
            "submission": submission,
            "days": days,
            "suggested_off_count": len(suggested_off_days),
            "auto_fill_off_count": len(auto_fill_off_days),
            "desired_off_limit": request.staff.desired_off_limit,
            "requested_off_count": requested_off_count,
            "rest_constraint_notes": _staff_rest_constraint_notes(request.staff),
        },
    )


@login_required
@staff_required
def my_shift(request):
    # スタッフの「シフト表確認」画面。
    # _calendar_days() の日付情報に、公開済みの割当(assignment)を足して my_shift.html に渡す。
    requested_month = request.GET.get("month")
    today_month = timezone.localdate().replace(day=1)
    min_month = _add_months(today_month, -1)
    max_month = _add_months(today_month, 1)
    month = _month(requested_month) if requested_month else max_month
    if month < min_month:
        month = min_month
    elif month > max_month:
        month = max_month
    prev_month = _add_months(month, -1) if month > min_month else None
    next_month = _add_months(month, 1) if month < max_month else None
    period = ShiftPeriod.objects.filter(
        company=request.company, month=month, status=ShiftPeriod.Status.PUBLISHED
    ).first()
    assignments = (
        period.assignments.filter(staff=request.staff).select_related("work_type")
        if period
        else []
    )
    assignment_map = {item.day.day: item for item in assignments}
    leave_request_map = {
        item.day.day: item
        for item in (
            period.leave_requests.filter(staff=request.staff).select_related(
                "assignment", "work_type"
            )
            if period
            else ShiftLeaveRequest.objects.none()
        )
    }
    days = [
        {
            **day,
            "assignment": assignment_map.get(day["number"]),
            "leave_request": leave_request_map.get(day["number"]),
        }
        for day in _calendar_days(month)
    ]
    return render(
        request,
        "shifts/my_shift.html",
        {
            "month": month,
            "period": period,
            "days": days,
            "min_month": min_month,
            "max_month": max_month,
            "prev_month": prev_month,
            "next_month": next_month,
        },
    )


@login_required
@staff_required
@require_POST
def request_shift_leave(request):
    assignment = get_object_or_404(
        ShiftAssignment.objects.select_related("period", "work_type"),
        pk=request.POST.get("assignment_id"),
        staff=request.staff,
        period__company=request.company,
        period__status=ShiftPeriod.Status.PUBLISHED,
    )
    reason = request.POST.get("reason", "").strip()
    leave_request, created = ShiftLeaveRequest.objects.update_or_create(
        period=assignment.period,
        staff=request.staff,
        day=assignment.day,
        defaults={
            "assignment": assignment,
            "work_type": assignment.work_type,
            "reason": reason,
            "status": ShiftLeaveRequest.Status.PENDING,
            "admin_note": "",
            "resolved_at": None,
            "resolved_by": None,
        },
    )
    if created:
        messages.success(request, "急な休み申請を送信しました。")
    else:
        messages.success(request, "急な休み申請を更新して再送信しました。")
    return redirect(f"{reverse('my_shift')}?month={assignment.period.month:%Y-%m}")


@login_required
@staff_required
def staff_change_password(request):
    form = PasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        messages.success(request, "パスワードを変更しました。")
        return redirect("staff_change_password")
    return render(request, "shifts/change_password.html", {"form": form})
