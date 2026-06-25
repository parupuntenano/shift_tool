from datetime import date
from io import BytesIO
from unittest import TestCase

from django.contrib.auth import get_user_model
from django.test import TestCase as DjangoTestCase
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from shifts.domain.import_data import (
    ImportedSkillLevel,
    ImportedSkillMap,
    ImportedStaffRow,
    ImportedWorkType,
)
from shifts.infrastructure.importers import SkillMapFileReader
from shifts.infrastructure.master_repository import DjangoMasterRepository
from shifts.infrastructure.repositories import DjangoShiftRepository
from shifts.infrastructure.models import (
    AvailabilityDay,
    AvailabilitySubmission,
    Company,
    CompanyMembership,
    ConstraintType,
    IndividualConstraint,
    PreviousMonthShiftDay,
    ShiftAssignment,
    ShiftPeriod,
    Staff,
    StaffSkill,
    SkillLevel,
    WorkType,
)


class SkillMapFileReaderTests(TestCase):
    def test_reads_utf8_csv_to_domain_data(self):
        raw = "社員番号,氏名,備考,受付\nS001,青木,4勤不可,◎\n".encode("utf-8")
        result = SkillMapFileReader().read("skills.csv", BytesIO(raw))
        self.assertEqual(result.rows[0].employee_number, "S001")
        self.assertEqual(result.rows[0].skills, {"受付": "◎"})

    def test_reads_skill_level_sheet_from_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "スキル表"
        sheet.append(["社員番号", "氏名", "公休数", "備考", "受付"])
        sheet.append(["S001", "青木", 9, "", "A"])
        level_sheet = workbook.create_sheet("スキル区分")
        level_sheet.append(["記号", "意味", "優先度", "アサイン可"])
        level_sheet.append(["A", "主担当", 1, "可"])
        work_sheet = workbook.create_sheet("業務マスタ")
        work_sheet.append(["業務名", "最低必要人数", "色", "有効"])
        work_sheet.append(["受付", 2, "#2563eb", "有効"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = SkillMapFileReader().read("skills.xlsx", stream)

        self.assertEqual(result.rows[0].skills, {"受付": "A"})
        self.assertEqual(result.rows[0].monthly_public_holidays, 9)
        self.assertIsNone(result.rows[0].desired_off_limit)
        self.assertEqual(result.skill_levels[0].symbol, "A")
        self.assertEqual(result.skill_levels[0].meaning, "主担当")
        self.assertEqual(result.skill_levels[0].priority, 1)
        self.assertTrue(result.skill_levels[0].assignable)
        self.assertEqual(result.work_types[0].name, "受付")
        self.assertEqual(result.work_types[0].minimum_staff_per_day, 2)
        self.assertEqual(result.work_types[0].color, "#2563eb")

    def test_reads_work_color_from_filled_work_name_cell(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "スキル表"
        sheet.append(["社員番号", "氏名", "備考", "受付"])
        sheet.append(["S001", "青木", "", "A"])
        work_sheet = workbook.create_sheet("業務マスタ")
        work_sheet.append(["業務名", "必要人数", "有効"])
        work_sheet.append(["受付", 2, "有効"])
        work_sheet["A2"].fill = PatternFill("solid", fgColor="22C55E")
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = SkillMapFileReader().read("skills.xlsx", stream)

        self.assertEqual(result.work_types[0].color, "#22c55e")


class MasterImportTests(DjangoTestCase):
    def test_import_creates_constraints_from_note(self):
        company = Company.objects.create(name="テスト", code="import-test")
        WorkType.objects.create(company=company, name="受付")
        data = ImportedSkillMap(
            (ImportedStaffRow("S001", "青木", "2勤1休;4勤不可;単休不可", {"受付": "○"}),)
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company)
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(staff.note, "2勤1休;4勤不可;単休不可")
        self.assertEqual(result["constraints"], 3)
        self.assertEqual(constraints.count(), 3)
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.WORK_REST_PATTERN,
                text_value="2,1",
                parameters__source="staff_note",
            ).exists()
        )
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.MAX_CONSECUTIVE,
                numeric_value=3,
                parameters__source="staff_note",
            ).exists()
        )
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.NO_SINGLE_REST,
                parameters__source="staff_note",
            ).exists()
        )

    def test_import_creates_staff_login_with_initial_password(self):
        company = Company.objects.create(name="テスト", code="account-import-test")
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}),))

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.select_related("user").get(
            company=company, employee_number="50592"
        )
        self.assertEqual(staff.user.username, "50592")
        self.assertTrue(staff.user.check_password("0000"))
        self.assertEqual(result["accounts"], 1)
        self.assertTrue(
            CompanyMembership.objects.filter(
                company=company, user=staff.user, role=CompanyMembership.Role.STAFF
            ).exists()
        )

    def test_import_updates_staff_public_holidays_and_uses_company_request_limit(self):
        company = Company.objects.create(
            name="テスト", code="staff-off-import-test", default_desired_off_limit=6
        )
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}, 9),))

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="50592")
        self.assertEqual(staff.monthly_public_holidays, 9)
        self.assertEqual(staff.desired_off_limit, 6)

    def test_import_updates_skill_levels_from_excel_definition(self):
        company = Company.objects.create(name="テスト", code="level-import-test")
        data = ImportedSkillMap(
            rows=(ImportedStaffRow("S001", "青木", "", {"受付": "A"}),),
            skill_levels=(
                ImportedSkillLevel("A", "主担当", 1, True),
            ),
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        level = SkillLevel.objects.get(company=company, symbol="A")
        self.assertEqual(result["levels"], 1)
        self.assertEqual((level.meaning, level.priority, level.assignable), ("主担当", 1, True))

    def test_import_updates_work_types_from_excel_definition(self):
        company = Company.objects.create(name="テスト", code="work-import-test")
        data = ImportedSkillMap(
            rows=(),
            work_types=(ImportedWorkType("受付", 2, True, "#16a34a"),),
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        work = WorkType.objects.get(company=company, name="受付")
        self.assertEqual(result["works"], 1)
        self.assertEqual(work.required_staff_per_day, 2)
        self.assertEqual(work.color, "#16a34a")
        self.assertTrue(work.active)

    def test_generation_skill_flags_detect_instructor_and_trainee_levels(self):
        company = Company.objects.create(name="テスト", code="skill-flag-test")
        instructor = Staff.objects.create(
            company=company, employee_number="S001", name="指導者"
        )
        trainee = Staff.objects.create(
            company=company, employee_number="S002", name="研修中"
        )
        work = WorkType.objects.create(company=company, name="受付")
        instructor_level = SkillLevel.objects.create(
            company=company,
            symbol="◎",
            meaning="主担当・指導可",
            priority=1,
            assignable=True,
        )
        trainee_level = SkillLevel.objects.create(
            company=company,
            symbol="△",
            meaning="補助・訓練中",
            priority=3,
            assignable=True,
        )
        StaffSkill.objects.create(
            staff=instructor, work_type=work, level=instructor_level
        )
        StaffSkill.objects.create(staff=trainee, work_type=work, level=trainee_level)

        ratings = {
            item.staff_id: item
            for item in DjangoShiftRepository().skills_for_generation(company.id)
        }

        self.assertTrue(ratings[instructor.id].instructor_capable)
        self.assertFalse(ratings[instructor.id].trainee)
        self.assertFalse(ratings[trainee.id].instructor_capable)
        self.assertTrue(ratings[trainee.id].trainee)

    def test_generation_uses_existing_previous_shift_before_imported_result(self):
        company = Company.objects.create(name="テスト", code="previous-source-test")
        staff = Staff.objects.create(
            company=company, employee_number="S001", name="青木"
        )
        generated_work = WorkType.objects.create(company=company, name="ロール")
        imported_work = WorkType.objects.create(company=company, name="受付")
        period = ShiftPeriod.objects.create(
            company=company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 6, 30),
            work_type=generated_work,
        )
        PreviousMonthShiftDay.objects.create(
            company=company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=imported_work,
            raw_value="受付",
        )

        previous_days = DjangoShiftRepository().previous_shift_days_for_generation(
            company.id, date(2026, 7, 1)
        )

        last_day = next(item for item in previous_days if item.day == date(2026, 6, 30))
        self.assertEqual(last_day.work_id, generated_work.id)
        self.assertEqual(last_day.status, PreviousMonthShiftDay.Status.WORK)
        self.assertTrue(
            any(
                item.day == date(2026, 6, 29)
                and item.status == PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY
                for item in previous_days
            )
        )

    def test_reimport_does_not_reset_existing_password(self):
        company = Company.objects.create(name="テスト", code="password-import-test")
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}),))
        repository = DjangoMasterRepository()
        repository.save_skill_map(company.id, data)
        user = get_user_model().objects.get(username="50592")
        user.set_password("changed-password")
        user.save()

        result = repository.save_skill_map(company.id, data)

        user.refresh_from_db()
        self.assertTrue(user.check_password("changed-password"))
        self.assertEqual(result["accounts"], 0)

    def test_import_creates_work_constraints_from_note(self):
        company = Company.objects.create(name="テスト", code="constraint-note-test")
        work_a = WorkType.objects.create(company=company, name="ロール")
        work_b = WorkType.objects.create(company=company, name="エーカス")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "ロールとエーカス交互;ロール連続不可;エーカス禁止",
                    {},
                ),
            )
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(result["constraints"], 3)
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.WORK_ALTERNATION,
                work_type_a=work_a,
                work_type_b=work_b,
            ).exists()
        )
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.AVOID_SPECIFIC_WORK,
                work_type_a=work_a,
            ).exists()
        )
        self.assertTrue(
            constraints.filter(
                rule_type__operator=ConstraintType.Operator.FORBID_SPECIFIC_WORK,
                work_type_a=work_b,
            ).exists()
        )

    def test_reimport_replaces_only_note_constraints(self):
        company = Company.objects.create(name="テスト", code="constraint-reimport-test")
        rule_type = ConstraintType.objects.create(
            company=company,
            name="手入力用",
            operator=ConstraintType.Operator.CUSTOM,
            default_strength=5,
        )
        repository = DjangoMasterRepository()
        repository.save_skill_map(
            company.id,
            ImportedSkillMap((ImportedStaffRow("S001", "青木", "2勤1休", {}),)),
        )
        staff = Staff.objects.get(company=company, employee_number="S001")
        IndividualConstraint.objects.create(
            company=company,
            staff=staff,
            rule_type=rule_type,
            name="手入力メモ",
            kind=rule_type.operator,
            strength=5,
            parameters={"source": "manual"},
        )

        result = repository.save_skill_map(
            company.id,
            ImportedSkillMap((ImportedStaffRow("S001", "青木", "3勤1休", {}),)),
        )

        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(result["constraints"], 1)
        self.assertEqual(constraints.filter(parameters__source="staff_note").count(), 1)
        self.assertTrue(
            constraints.filter(
                parameters__source="staff_note",
                text_value="3,1",
            ).exists()
        )
        self.assertTrue(constraints.filter(name="手入力メモ").exists())
