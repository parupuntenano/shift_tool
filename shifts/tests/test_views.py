import csv
from datetime import date
from io import BytesIO, StringIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import NoReverseMatch, reverse
from openpyxl import Workbook, load_workbook

from shifts.infrastructure.models import (
    AvailabilityDay,
    AvailabilitySubmission,
    Company,
    CompanyMembership,
    ConstraintType,
    GenerationWarning,
    IndividualConstraint,
    PreviousMonthShiftDay,
    ShiftAssignment,
    ShiftLeaveRequest,
    ShiftPeriod,
    SkillLevel,
    Staff,
    StaffSkill,
    WorkType,
)


class LoginRoutingTests(TestCase):
    def test_admin_is_routed_to_manager_dashboard(self):
        company = Company.objects.create(name="テスト", code="test")
        user = get_user_model().objects.create_user("admin-test", password="pass")
        CompanyMembership.objects.create(company=company, user=user, role="admin")
        self.client.login(username="admin-test", password="pass")
        response = self.client.get(reverse("home"))
        self.assertRedirects(response, reverse("manager_dashboard"))


class StaffLoginClosedTests(TestCase):
    def test_staff_only_urls_are_closed(self):
        for url_name in (
            "submit_availability",
            "my_shift",
            "request_shift_leave",
            "staff_change_password",
        ):
            with self.subTest(url_name=url_name):
                with self.assertRaises(NoReverseMatch):
                    reverse(url_name)

    def test_staff_user_is_not_routed_to_staff_screen(self):
        company = Company.objects.create(name="テスト", code="staff-closed-test")
        user = get_user_model().objects.create_user("S001", password="0000")
        CompanyMembership.objects.create(company=company, user=user, role="staff")
        self.client.login(username="S001", password="0000")

        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "管理者のみ利用できます")


class ManagerCrudTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(name="テスト", code="crud-test")
        self.other_company = Company.objects.create(name="他社", code="other")
        self.user = get_user_model().objects.create_user("crud-admin", password="pass")
        CompanyMembership.objects.create(
            company=self.company, user=self.user, role="admin"
        )
        self.client.login(username="crud-admin", password="pass")

    def test_can_edit_work_in_own_company(self):
        work = WorkType.objects.create(
            company=self.company, name="旧業務", display_order=1
        )
        response = self.client.post(
            reverse("work_edit", args=[work.pk]),
            {
                "name": "新業務",
                "display_order": 2,
                "required_staff_per_day": 3,
                "color": "#22c55e",
                "active": "on",
            },
        )
        self.assertRedirects(response, reverse("work_manage"))
        work.refresh_from_db()
        self.assertEqual((work.name, work.required_staff_per_day), ("新業務", 3))
        self.assertEqual(work.color, "#22c55e")

    def test_staff_manage_hides_login_and_request_limit_fields(self):
        Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="青木 太郎",
            desired_off_limit=4,
        )

        response = self.client.get(reverse("staff_manage"))

        self.assertContains(response, "スタッフ管理")
        self.assertNotContains(response, "ログインID")
        self.assertNotContains(response, "パスワード")
        self.assertNotContains(response, "希望上限")
        self.assertNotContains(response, "公有給希望上限")

    def test_staff_registration_does_not_create_login_account(self):
        response = self.client.post(
            reverse("staff_manage"),
            {
                "employee_number": "S300",
                "name": "後入 太郎",
                "monthly_public_holidays": "8",
                "is_employee": "on",
                "note": "",
                "active": "on",
            },
        )
        self.assertRedirects(response, reverse("staff_manage"))
        added = Staff.objects.get(company=self.company, employee_number="S300")
        self.assertIsNone(added.user)
        self.assertEqual(added.desired_off_limit, self.company.default_desired_off_limit)
        self.assertTrue(added.is_employee)

    def test_staff_bulk_delete_removes_staff_related_data_only(self):
        staff_user = get_user_model().objects.create_user("S100", password="pass")
        staff = Staff.objects.create(
            company=self.company,
            user=staff_user,
            employee_number="S100",
            name="削除 太郎",
        )
        other_staff = Staff.objects.create(
            company=self.other_company,
            employee_number="O100",
            name="他社 花子",
        )
        CompanyMembership.objects.create(
            company=self.company,
            user=staff_user,
            role=CompanyMembership.Role.STAFF,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company,
            symbol="○",
            meaning="対応可",
            assignable=True,
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=level)
        IndividualConstraint.objects.create(
            company=self.company,
            staff=staff,
            name="単休不可",
            kind=IndividualConstraint.Kind.NO_SINGLE_REST,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 1),
            preferred_off=True,
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )
        ShiftLeaveRequest.objects.create(
            period=period,
            staff=staff,
            assignment=assignment,
            day=date(2026, 7, 1),
            work_type=work,
        )
        GenerationWarning.objects.create(
            period=period,
            day=date(2026, 7, 1),
            work_type=work,
            message="警告",
        )

        response = self.client.post(reverse("staff_bulk_delete"))

        self.assertRedirects(response, reverse("staff_manage"))
        self.assertFalse(Staff.objects.filter(company=self.company).exists())
        self.assertTrue(Staff.objects.filter(pk=other_staff.pk).exists())
        self.assertFalse(
            CompanyMembership.objects.filter(
                company=self.company,
                user=staff_user,
                role=CompanyMembership.Role.STAFF,
            ).exists()
        )
        self.assertTrue(
            CompanyMembership.objects.filter(
                company=self.company,
                user=self.user,
                role=CompanyMembership.Role.ADMIN,
            ).exists()
        )
        self.assertFalse(StaffSkill.objects.filter(staff__company=self.company).exists())
        self.assertFalse(
            IndividualConstraint.objects.filter(company=self.company).exists()
        )
        self.assertFalse(
            AvailabilitySubmission.objects.filter(staff__company=self.company).exists()
        )
        self.assertFalse(
            PreviousMonthShiftDay.objects.filter(company=self.company).exists()
        )
        self.assertFalse(ShiftPeriod.objects.filter(company=self.company).exists())
        self.assertTrue(WorkType.objects.filter(pk=work.pk).exists())
        self.assertTrue(SkillLevel.objects.filter(pk=level.pk).exists())

    def test_cannot_edit_other_company_data(self):
        work = WorkType.objects.create(company=self.other_company, name="他社業務")
        response = self.client.get(reverse("work_edit", args=[work.pk]))
        self.assertEqual(response.status_code, 404)

    def test_can_download_import_template(self):
        response = self.client.get(reverse("download_import_template"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "shift_import_template.xlsx",
            response["Content-Disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["スキル表"]
        levels = workbook["スキル区分"]
        works = workbook["業務マスタ"]
        previous = workbook["先月シフト実績"]
        guide = workbook["入力ルール"]

        self.assertNotIn("業務スキル記入例", workbook.sheetnames)
        self.assertEqual(
            [sheet.cell(row=1, column=index).value for index in range(1, 8)],
            ["社員番号", "氏名", "公休数", "備考", "業務A", "業務B", "業務C"],
        )
        self.assertEqual(sheet["A2"].value, "S001")
        self.assertEqual(sheet["C2"].value, 8)
        self.assertEqual(sheet["D2"].value, "4勤不可;単休不可")
        self.assertEqual(sheet["E2"].value, "◎")
        self.assertEqual(
            [levels.cell(row=1, column=index).value for index in range(1, 5)],
            ["記号", "意味", "優先度", "アサイン可"],
        )
        self.assertIsNone(levels["E1"].value)
        self.assertEqual(levels["A2"].value, "◎")
        self.assertEqual(
            [works.cell(row=1, column=index).value for index in range(1, 4)],
            ["業務名", "必要人数", "有効"],
        )
        self.assertEqual(works["B2"].value, 1)
        self.assertEqual(works["C2"].value, "有効")
        self.assertEqual(works["A2"].fill.fgColor.rgb[-6:], "2563EB")
        self.assertEqual(
            [previous.cell(row=1, column=index).value for index in range(1, 5)],
            ["社員番号", "氏名", "6/20", "6/21"],
        )
        self.assertEqual(previous["A2"].value, "S001")
        self.assertEqual(previous["C2"].value, "業務A")
        self.assertEqual(guide["A1"].value, "項目")

    def test_can_download_import_sample(self):
        response = self.client.get(reverse("download_import_sample"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "shift_import_sample.xlsx",
            response["Content-Disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["スキル表"]
        works = workbook["業務マスタ"]
        levels = workbook["スキル区分"]
        previous = workbook["先月シフト実績"]

        self.assertNotIn("業務スキル記入例", workbook.sheetnames)
        self.assertEqual(sheet.max_row, 31)
        self.assertEqual(
            [sheet.cell(row=1, column=index).value for index in range(1, 13)],
            ["社員番号", "氏名", "公休数", "備考", "A", "B", "C", "D", "E", "F", "G", "H"],
        )
        self.assertEqual(sheet["A2"].value, "S001")
        self.assertEqual(sheet["C2"].value, 8)
        self.assertEqual(sheet["D2"].value, "単休不可")
        self.assertEqual(sheet["E2"].value, "◎")
        self.assertEqual(
            [sheet.cell(row=row, column=3).value for row in range(2, 32)],
            [8, 9] * 15,
        )
        self.assertLessEqual(
            max(
                len(
                    [
                        token
                        for token in str(sheet.cell(row=row, column=4).value or "").split(";")
                        if token
                    ]
                )
                for row in range(2, 32)
            ),
            4,
        )
        self.assertEqual(works.max_row, 9)
        self.assertEqual(works["A2"].value, "A")
        self.assertEqual(
            [works.cell(row=1, column=index).value for index in range(1, 4)],
            ["業務名", "必要人数", "有効"],
        )
        self.assertEqual(works["B2"].value, 4)
        self.assertEqual(
            [works.cell(row=row, column=2).value for row in range(2, 10)],
            [4, 3, 3, 2, 2, 2, 2, 1],
        )
        self.assertEqual(works["C2"].value, "有効")
        self.assertEqual(works["A2"].fill.fgColor.rgb[-6:], "2563EB")
        self.assertEqual(levels["A2"].value, "◎")
        self.assertEqual(previous["A2"].value, "S001")
        self.assertEqual(previous["C2"].value, "A")
        self.assertNotIn(
            "A",
            [previous.cell(row=4, column=column).value for column in range(3, 10)],
        )
        self.assertNotIn(
            "B",
            [previous.cell(row=7, column=column).value for column in range(3, 10)],
        )
        self.assertEqual(previous["L8"].value, "G")
        self.assertEqual(previous["M8"].value, "H")

    def test_can_download_csv_template_and_sample(self):
        template_response = self.client.get(reverse("download_csv_template"))
        sample_response = self.client.get(reverse("download_csv_sample"))

        self.assertEqual(template_response.status_code, 200)
        self.assertEqual(sample_response.status_code, 200)
        self.assertEqual(
            template_response["Content-Type"], "text/csv; charset=utf-8"
        )
        self.assertEqual(
            sample_response["Content-Type"], "text/csv; charset=utf-8"
        )
        self.assertIn(
            'filename="shift_import_template.csv"',
            template_response["Content-Disposition"],
        )
        self.assertIn(
            'filename="shift_import_sample.csv"',
            sample_response["Content-Disposition"],
        )
        template_rows = list(
            csv.reader(StringIO(template_response.content.decode("utf-8-sig")))
        )
        sample_rows = list(
            csv.reader(StringIO(sample_response.content.decode("utf-8-sig")))
        )
        self.assertEqual(
            template_rows[0],
            ["社員番号", "氏名", "公休数", "備考", "業務A", "業務B", "業務C"],
        )
        self.assertEqual(sample_rows[0][4:], ["A", "B", "C", "D", "E", "F", "G", "H"])
        self.assertEqual(sample_rows[1][0], "S001")

    def test_previous_shift_import_reads_saturday_week_context_days(self):
        from shifts.presentation.views import _import_previous_shift_days

        Staff.objects.create(
            company=self.company, employee_number="S001", name="青木 太郎"
        )
        WorkType.objects.create(company=self.company, name="受付")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "先月シフト実績"
        sheet.append(
            [
                "社員番号",
                "氏名",
                "2026/6/19",
                "6/20",
                "6/21",
                "6/22",
                "6/23",
                "6/24",
                "6/25",
                "6/26",
                "6/27",
                "6/28",
                "6/29",
                "6/30",
            ]
        )
        sheet.append(
            [
                "S001",
                "青木 太郎",
                "受付",
                "受付",
                "公休",
                "受付",
                "有給",
                "受付",
                "公休",
                "有給",
                "受付",
                "受付",
                "公休",
                "受付",
            ]
        )
        file_obj = BytesIO()
        file_obj.name = "previous.xlsx"
        workbook.save(file_obj)

        result = _import_previous_shift_days(
            self.company, date(2026, 6, 1), file_obj
        )

        self.assertEqual(result["days"], 11)
        self.assertFalse(
            PreviousMonthShiftDay.objects.filter(day=date(2026, 6, 19)).exists()
        )
        self.assertEqual(
            PreviousMonthShiftDay.objects.filter(
                company=self.company,
                day__range=(date(2026, 6, 20), date(2026, 6, 30)),
            ).count(),
            11,
        )

    def test_previous_shift_import_accepts_inactive_work_name(self):
        from shifts.presentation.views import _import_previous_shift_days

        staff = Staff.objects.create(
            company=self.company,
            employee_number="S001",
            name="青木 太郎",
        )
        work = WorkType.objects.create(
            company=self.company,
            name="前月だけの業務",
            active=False,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "先月シフト実績"
        sheet.append(["社員番号", "氏名", "6/30"])
        sheet.append(["S001", "青木 太郎", "前月だけの業務"])
        file_obj = BytesIO()
        file_obj.name = "previous.xlsx"
        workbook.save(file_obj)

        _import_previous_shift_days(self.company, date(2026, 6, 1), file_obj)

        item = PreviousMonthShiftDay.objects.get(staff=staff, day=date(2026, 6, 30))
        self.assertEqual(item.status, PreviousMonthShiftDay.Status.WORK)
        self.assertEqual(item.work_type, work)

    def test_previous_shift_month_form_accepts_browser_month_value(self):
        from shifts.presentation.forms import PreviousShiftImportForm

        cleaned = PreviousShiftImportForm.base_fields["month"].clean("2026-06")

        self.assertEqual(cleaned, date(2026, 6, 1))

    def test_previous_shift_header_reads_japanese_month_end_text(self):
        from shifts.presentation.views import _previous_shift_day_from_header

        month = date(2026, 6, 1)

        self.assertEqual(
            _previous_shift_day_from_header("2026年6月30日(火)", month),
            date(2026, 6, 30),
        )
        self.assertEqual(
            _previous_shift_day_from_header("6/30", month),
            date(2026, 6, 30),
        )
        self.assertEqual(
            _previous_shift_day_from_header("6月30日", month),
            date(2026, 6, 30),
        )
        self.assertEqual(
            _previous_shift_day_from_header("30日", month),
            date(2026, 6, 30),
        )

    def test_previous_shift_import_requires_named_sheet(self):
        from shifts.presentation.views import _import_previous_shift_days

        workbook = Workbook()
        workbook.active.title = "別シート"
        file_obj = BytesIO()
        file_obj.name = "previous.xlsx"
        workbook.save(file_obj)

        with self.assertRaisesMessage(
            ValueError, "Excel内に「先月シフト実績」シートが必要です。"
        ):
            _import_previous_shift_days(self.company, date(2026, 6, 1), file_obj)

    def test_previous_shift_list_shows_imported_results(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S001", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 24),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
            raw_value="受付",
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 25),
            status=PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY,
            raw_value="公休",
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 26),
            status=PreviousMonthShiftDay.Status.PAID_LEAVE,
            raw_value="有給",
        )

        response = self.client.get(
            reverse("previous_shift_list"), {"month": "2026-06"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "先月シフト実績確認")
        self.assertContains(response, "2026年6月 / 6/20〜6/30")
        self.assertContains(response, "青木 太郎")
        self.assertContains(response, "S001")
        self.assertContains(response, "受付")
        self.assertContains(response, "公休")
        self.assertContains(response, "有給")

    def test_previous_shift_list_prefers_generated_shift_over_imported_results(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S001", name="青木 太郎"
        )
        generated_work = WorkType.objects.create(company=self.company, name="ロール")
        imported_work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 6, 24),
            work_type=generated_work,
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 24),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=imported_work,
            raw_value="受付",
        )

        response = self.client.get(
            reverse("previous_shift_list"), {"month": "2026-06"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "作成済みシフト表")
        self.assertContains(response, "ロール")
        self.assertNotContains(response, "受付")

    def test_previous_shift_list_default_prefers_latest_imported_month(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S001", name="青木 太郎"
        )
        generated_work = WorkType.objects.create(company=self.company, name="当月業務")
        imported_work = WorkType.objects.create(company=self.company, name="前月業務")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=generated_work,
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=imported_work,
            raw_value="前月業務",
        )

        response = self.client.get(reverse("previous_shift_list"))

        self.assertEqual(response.context["month"], date(2026, 6, 1))
        self.assertContains(response, "前月業務")
        self.assertNotContains(response, "当月業務")

    def test_draft_shift_detail_shows_edit_controls_and_support(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(
            company=self.company, name="受付", required_staff_per_day=1
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertContains(response, "シフト編集モード")
        self.assertContains(response, f'name="assignment_{staff.pk}_20260701"')
        self.assertContains(response, "シフト調整サポート")

    def test_shift_detail_shows_staff_and_daily_work_statistics(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work_a = WorkType.objects.create(
            company=self.company, name="受付", display_order=1
        )
        work_b = WorkType.objects.create(
            company=self.company, name="ロール", display_order=2
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work_a,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 2),
            work_type=work_b,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertContains(response, "受付")
        self.assertContains(response, "ロール")
        self.assertContains(response, "daily-work-stat-row")
        self.assertEqual(response.context["rows"][0]["work_summary"][0]["count"], 1)
        self.assertEqual(response.context["rows"][0]["work_summary"][1]["count"], 1)
        self.assertEqual(response.context["daily_work_stats"][0]["counts"][0], 1)

    def test_shift_detail_shows_previous_week_as_readonly_reference(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        previous_work = WorkType.objects.create(
            company=self.company, name="前月業務", display_order=1
        )
        current_work = WorkType.objects.create(
            company=self.company, name="当月業務", display_order=2
        )
        previous_period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=previous_period,
            staff=staff,
            day=date(2026, 6, 30),
            work_type=previous_work,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=current_work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertEqual(len(response.context["previous_days"]), 11)
        self.assertEqual(
            response.context["rows"][0]["previous_cells"][-1]["label"],
            "前月業務",
        )
        self.assertContains(response, "前月")
        self.assertContains(response, "readonly-previous-cell")
        self.assertContains(response, 'id="current-month-start"')
        self.assertNotContains(response, f'name="assignment_{staff.pk}_20260630"')

    def test_shift_detail_shows_inactive_previous_work_reference(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        previous_work = WorkType.objects.create(
            company=self.company, name="前月だけの業務", active=False
        )
        current_work = WorkType.objects.create(company=self.company, name="当月業務")
        previous_period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=previous_period,
            staff=staff,
            day=date(2026, 6, 30),
            work_type=previous_work,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=current_work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertEqual(
            response.context["rows"][0]["previous_cells"][-1]["label"],
            "前月だけの業務",
        )
        self.assertContains(response, "前月だけの業務")

    def test_shift_detail_has_download_buttons(self):
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertContains(response, reverse("download_shift_excel", args=[period.pk]))
        self.assertContains(response, reverse("download_shift_csv", args=[period.pk]))

    def test_can_download_shift_csv_without_summary_columns(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("download_shift_csv", args=[period.pk]))
        content = response.content.decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertIn('filename="shift_2026_07.csv"', response["Content-Disposition"])
        self.assertIn("青木 太郎", content)
        self.assertIn("受付", content)
        self.assertNotIn("出勤", content)
        self.assertNotIn("合計", content)

    def test_can_download_shift_excel_without_summary_columns(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付", color="#2563eb")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("download_shift_excel", args=[period.pk]))
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["シフト表"]
        values = [
            cell.value
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]

        self.assertEqual(response.status_code, 200)
        self.assertIn('filename="shift_2026_07.xlsx"', response["Content-Disposition"])
        self.assertEqual(sheet["A1"].value, "2026年7月 シフト表")
        self.assertEqual(sheet["A4"].value, "社員番号")
        self.assertEqual(sheet["C4"].value, 1)
        self.assertEqual(sheet["C6"].value, "受付")
        self.assertEqual(sheet.freeze_panes, "C6")
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertNotIn("出勤", values)
        self.assertNotIn("合計", values)

    def test_shift_download_leaves_employee_rest_days_blank(self):
        employee = Staff.objects.create(
            company=self.company,
            employee_number="E100",
            name="社員 太郎",
            is_employee=True,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=employee,
            day=date(2026, 7, 1),
            work_type=work,
        )

        csv_response = self.client.get(reverse("download_shift_csv", args=[period.pk]))
        csv_rows = list(
            csv.reader(StringIO(csv_response.content.decode("utf-8-sig")))
        )
        excel_response = self.client.get(reverse("download_shift_excel", args=[period.pk]))
        workbook = load_workbook(BytesIO(excel_response.content))
        sheet = workbook["シフト表"]

        self.assertEqual(csv_rows[2][2], "受付")
        self.assertEqual(csv_rows[2][3], "")
        self.assertEqual(sheet["C6"].value, "受付")
        self.assertIsNone(sheet["D6"].value)

    def test_shift_excel_download_does_not_include_previous_week_reference(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        previous_work = WorkType.objects.create(company=self.company, name="前月業務")
        current_work = WorkType.objects.create(company=self.company, name="当月業務")
        previous_period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=previous_period,
            staff=staff,
            day=date(2026, 6, 30),
            work_type=previous_work,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=current_work,
        )

        response = self.client.get(reverse("download_shift_excel", args=[period.pk]))
        workbook = load_workbook(BytesIO(response.content))
        values = [
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]

        self.assertIn("当月業務", values)
        self.assertNotIn("前月業務", values)

    def test_shift_detail_warns_when_public_holiday_count_misses_target(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="青木 太郎",
            monthly_public_holidays=8,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 2),
            paid_leave=True,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertContains(response, "警告")
        self.assertContains(response, "設定8")
        self.assertContains(response, "公休数の超過・不足")
        self.assertContains(response, "0/8")
        self.assertEqual(response.context["rows"][0]["public_holiday_count"], 0)
        self.assertEqual(response.context["rows"][0]["public_holiday_target"], 8)
        self.assertEqual(response.context["rows"][0]["public_holiday_status"], "under")
        self.assertEqual(
            response.context["edit_support"]["public_holiday_rows"][0]["actual"], 0
        )
        self.assertEqual(
            response.context["edit_support"]["public_holiday_rows"][0]["target"], 8
        )
        self.assertEqual(
            response.context["edit_support"]["public_holiday_rows"][0]["difference"], -8
        )
        self.assertTrue(response.context["edit_support"]["public_holiday_has_issues"])
        self.assertEqual(response.context["rows"][0]["paid_leave_count"], 1)
        self.assertEqual(response.context["rows"][0]["assignment_pending_count"], 29)
        self.assertEqual(response.context["rows"][0]["expected_total"], 2)
        self.assertEqual(response.context["rows"][0]["expected_total_status"], "under")

    def test_shift_detail_shortage_count_ignores_trainee_assignment(self):
        instructor = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="指導者",
        )
        trainee = Staff.objects.create(
            company=self.company,
            employee_number="S101",
            name="研修中",
        )
        work = WorkType.objects.create(
            company=self.company,
            name="受付",
            required_staff_per_day=2,
        )
        instructor_level = SkillLevel.objects.create(
            company=self.company,
            symbol="◎",
            meaning="主担当・指導可",
            priority=1,
            assignable=True,
        )
        trainee_level = SkillLevel.objects.create(
            company=self.company,
            symbol="△",
            meaning="補助・訓練中",
            priority=3,
            assignable=True,
        )
        StaffSkill.objects.create(
            staff=instructor,
            work_type=work,
            level=instructor_level,
        )
        StaffSkill.objects.create(staff=trainee, work_type=work, level=trainee_level)
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=instructor,
            day=date(2026, 7, 1),
            work_type=work,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=trainee,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        shortage = response.context["edit_support"]["daily_shortages"][0]
        self.assertEqual(shortage["count"], 1)
        self.assertEqual(shortage["required"], 2)
        self.assertEqual(shortage["shortage"], 1)

    def test_shift_detail_distinguishes_requested_holidays_and_assignment_pending(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="青木 太郎",
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 2),
            preferred_off=True,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        cells = response.context["rows"][0]["cells"]
        self.assertEqual(cells[1]["rest_kind"], "requested-public-holiday")
        self.assertEqual(cells[1]["rest_label"], "申請公休")
        self.assertEqual(cells[2]["rest_kind"], "assignment-pending")
        self.assertEqual(cells[2]["rest_label"], "割当待ち")
        self.assertContains(response, "requested-public-holiday")
        self.assertContains(response, "assignment-pending")

    def test_shift_detail_limits_preferred_holidays_to_two_per_week_excluding_paid_leave(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="青木 太郎",
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 1),
            preferred_off=True,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 2),
            paid_leave=True,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 3),
            preferred_off=True,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 4),
            preferred_off=True,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        cells = response.context["rows"][0]["cells"]
        self.assertEqual(cells[0]["rest_kind"], "requested-public-holiday")
        self.assertEqual(cells[1]["rest_kind"], "paid-leave")
        self.assertEqual(cells[2]["rest_kind"], "requested-public-holiday")
        self.assertEqual(cells[3]["rest_kind"], "assignment-pending")
        self.assertEqual(response.context["rows"][0]["public_holiday_count"], 2)
        self.assertEqual(response.context["rows"][0]["paid_leave_count"], 1)

    def test_can_update_draft_shift_assignment(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        old_work = WorkType.objects.create(company=self.company, name="受付")
        new_work = WorkType.objects.create(company=self.company, name="ロール")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=old_work,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(new_work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        assignment.refresh_from_db()
        self.assertEqual(assignment.work_type, new_work)
        self.assertTrue(assignment.manually_edited)

    def test_shift_detail_save_controls_are_bound_to_draft_form(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))
        html = response.content.decode()

        self.assertIn('id="draft-shift-form"', html)
        self.assertIn('type="submit" form="draft-shift-form"', html)
        self.assertIn('name="assignment_', html)
        self.assertIn('form="draft-shift-form" class="draft-shift-select"', html)
        self.assertIn('select.dataset.initialValue = select.value;', html)
        self.assertIn('select.disabled = select.value === select.dataset.initialValue;', html)

    def test_can_mark_unassigned_day_as_public_holiday(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": "public_holiday"},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        day = AvailabilityDay.objects.get(
            submission__staff=staff,
            submission__month=date(2026, 7, 1),
            day=date(2026, 7, 1),
        )
        self.assertTrue(day.preferred_off)
        self.assertFalse(day.paid_leave)

    def test_unchanged_shift_save_keeps_assignment_auto(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        assignment.refresh_from_db()
        self.assertFalse(assignment.manually_edited)

        detail_response = self.client.get(reverse("shift_detail", args=[period.pk]))
        self.assertContains(detail_response, "assignment-source-mark auto")
        self.assertContains(detail_response, "自動")

    def test_shift_edit_dropdown_hides_unassignable_staff_work(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        allowed_work = WorkType.objects.create(company=self.company, name="受付")
        blocked_work = WorkType.objects.create(company=self.company, name="ロール")
        allowed_level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可", assignable=True
        )
        blocked_level = SkillLevel.objects.create(
            company=self.company, symbol="×", meaning="対応不可", assignable=False
        )
        StaffSkill.objects.create(
            staff=staff, work_type=allowed_work, level=allowed_level
        )
        StaffSkill.objects.create(
            staff=staff, work_type=blocked_work, level=blocked_level
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertEqual(
            [work.id for work in response.context["rows"][0]["work_options"]],
            [allowed_work.id],
        )
        self.assertContains(response, f'value="{allowed_work.id}"')
        self.assertNotContains(response, f'value="{blocked_work.id}"')

    def test_update_shift_draft_ignores_unassignable_staff_work(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        blocked_work = WorkType.objects.create(company=self.company, name="ロール")
        blocked_level = SkillLevel.objects.create(
            company=self.company, symbol="×", meaning="対応不可", assignable=False
        )
        StaffSkill.objects.create(
            staff=staff, work_type=blocked_work, level=blocked_level
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(blocked_work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(
            ShiftAssignment.objects.filter(
                period=period,
                staff=staff,
                day=date(2026, 7, 1),
            ).exists()
        )

    def test_update_shift_draft_ignores_constraints(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        rule_type = ConstraintType.objects.create(
            company=self.company,
            name="受付禁止",
            operator=ConstraintType.Operator.FORBID_SPECIFIC_WORK,
        )
        IndividualConstraint.objects.create(
            company=self.company,
            rule_type=rule_type,
            staff=staff,
            work_type_a=work,
            name="青木受付禁止",
            kind=IndividualConstraint.Kind.CUSTOM,
            strength=10,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        period.refresh_from_db()
        self.assertEqual(period.warning_count, 0)
        self.assertFalse(period.warnings.exists())
        self.assertTrue(
            ShiftAssignment.objects.filter(
                period=period,
                staff=staff,
                day=date(2026, 7, 1),
                work_type=work,
            ).exists()
        )

    def test_update_shift_draft_does_not_warn_for_work_rest_pattern(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 29),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
            raw_value="受付",
        )
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
            raw_value="受付",
        )
        rule_type = ConstraintType.objects.create(
            company=self.company,
            name="2勤1休",
            operator=ConstraintType.Operator.WORK_REST_PATTERN,
        )
        IndividualConstraint.objects.create(
            company=self.company,
            rule_type=rule_type,
            staff=staff,
            name="青木2勤1休",
            kind=IndividualConstraint.Kind.CUSTOM,
            text_value="2,1",
            strength=10,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(period.warnings.exists())

    def test_update_shift_draft_does_not_warn_for_no_single_rest(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        PreviousMonthShiftDay.objects.create(
            company=self.company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
            raw_value="受付",
        )
        rule_type = ConstraintType.objects.create(
            company=self.company,
            name="単休禁止",
            operator=ConstraintType.Operator.NO_SINGLE_REST,
        )
        IndividualConstraint.objects.create(
            company=self.company,
            rule_type=rule_type,
            staff=staff,
            name="青木単休禁止",
            kind=IndividualConstraint.Kind.NO_SINGLE_REST,
            strength=10,
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260702": str(work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(period.warnings.exists())

    def test_shift_detail_warning_section_hides_constraint_violations(self):
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
            warning_count=2,
        )
        GenerationWarning.objects.create(
            period=period,
            day=date(2026, 7, 1),
            work_type=work,
            message="制約違反：青木禁止業務に入っています。",
        )
        GenerationWarning.objects.create(
            period=period,
            day=date(2026, 7, 2),
            work_type=work,
            message="受付の必要人数が不足しています。",
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertEqual(response.context["visible_warning_count"], 1)
        self.assertNotContains(response, "制約違反：青木禁止業務に入っています。")
        self.assertContains(response, "警告 1件")
        self.assertContains(response, "受付の必要人数が不足しています。")

    def test_generate_shift_ignores_soft_constraint(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(
            company=self.company, name="受付", required_staff_per_day=1
        )
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可", assignable=True
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=level)
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 1),
            available=True,
        )
        rule_type = ConstraintType.objects.create(
            company=self.company,
            name="受付回避",
            operator=ConstraintType.Operator.FORBID_SPECIFIC_WORK,
            default_strength=5,
            default_is_hard=False,
        )
        IndividualConstraint.objects.create(
            company=self.company,
            rule_type=rule_type,
            staff=staff,
            work_type_a=work,
            name="青木受付回避",
            kind=IndividualConstraint.Kind.CUSTOM,
            strength=5,
            is_hard=False,
        )

        response = self.client.post(
            reverse("generate_shift"),
            {"month": "2026-07"},
        )

        period = ShiftPeriod.objects.get(company=self.company, month=date(2026, 7, 1))
        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(period.warnings.filter(message__contains="制約違反").exists())
        self.assertTrue(period.assignments.filter(staff=staff, work_type=work).exists())

    def test_draft_edit_can_assign_on_requested_public_or_paid_leave(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 1),
            preferred_off=True,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 2),
            paid_leave=True,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {
                f"assignment_{staff.pk}_20260701": str(work.pk),
                f"assignment_{staff.pk}_20260702": str(work.pk),
            },
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertEqual(
            ShiftAssignment.objects.filter(period=period, staff=staff).count(),
            2,
        )

    def test_shift_edit_support_ignores_missing_skill_for_employee_tag(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="社員 太郎",
            is_employee=True,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        messages_text = [
            issue["message"]
            for issue in response.context["edit_support"]["assignment_issues"]
        ]
        self.assertFalse(
            any("スキルが未設定" in message for message in messages_text)
        )

    def test_shift_edit_support_still_warns_unassignable_employee_skill(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="社員 太郎",
            is_employee=True,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        blocked_level = SkillLevel.objects.create(
            company=self.company, symbol="×", meaning="対応不可", assignable=False
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=blocked_level)
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        messages_text = [
            issue["message"]
            for issue in response.context["edit_support"]["assignment_issues"]
        ]
        self.assertTrue(any("アサイン不可" in message for message in messages_text))

    def test_can_change_draft_shift_assignment_to_rest(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": ""},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(
            ShiftAssignment.objects.filter(
                period=period,
                staff=staff,
                day=date(2026, 7, 1),
            ).exists()
        )

    def test_shift_edit_dropdown_can_change_assignment_to_paid_leave(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": "paid_leave"},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        self.assertFalse(
            ShiftAssignment.objects.filter(
                period=period,
                staff=staff,
                day=date(2026, 7, 1),
            ).exists()
        )
        self.assertTrue(
            AvailabilityDay.objects.filter(
                submission__staff=staff,
                submission__month=date(2026, 7, 1),
                day=date(2026, 7, 1),
                paid_leave=True,
            ).exists()
        )

    def test_published_shift_can_be_edited(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        old_work = WorkType.objects.create(company=self.company, name="受付")
        new_work = WorkType.objects.create(company=self.company, name="ロール")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=old_work,
        )

        response = self.client.post(
            reverse("update_shift_draft", args=[period.pk]),
            {f"assignment_{staff.pk}_20260701": str(new_work.pk)},
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        assignment.refresh_from_db()
        self.assertEqual(assignment.work_type, new_work)
        self.assertTrue(assignment.manually_edited)

    def test_shift_edit_support_flags_unavailable_assignment(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.DRAFT,
        )
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 1),
            available=False,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        self.assertContains(response, "勤務不可で提出されています")

    def test_manager_dashboard_highlights_pending_sudden_leave(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 7, 1),
            work_type=work,
        )
        ShiftLeaveRequest.objects.create(
            period=period,
            staff=staff,
            assignment=assignment,
            day=date(2026, 7, 1),
            work_type=work,
            reason="体調不良",
        )

        response = self.client.get(reverse("manager_dashboard"), {"month": "2026-07"})

        self.assertContains(response, "急な休み申請が未対応です")
        self.assertContains(response, "急休 未対応1件")

    def test_manager_dashboard_excludes_employee_tag_from_submission_metrics(self):
        submitted_staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="提出済み 太郎",
        )
        Staff.objects.create(
            company=self.company,
            employee_number="S200",
            name="未提出 花子",
        )
        Staff.objects.create(
            company=self.company,
            employee_number="E100",
            name="社員 応援",
            is_employee=True,
        )
        AvailabilitySubmission.objects.create(
            staff=submitted_staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )

        response = self.client.get(reverse("manager_dashboard"), {"month": "2026-07"})

        self.assertEqual(response.context["staff_count"], 2)
        self.assertEqual(response.context["submitted_count"], 1)
        self.assertEqual(response.context["missing_count"], 1)
        self.assertEqual(response.context["submission_rate"], 50)
        self.assertContains(response, "社員タグ除外")

    def test_missing_submissions_page_lists_only_non_employee_unsubmitted_staff(self):
        submitted_staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="提出済み 太郎",
        )
        missing_staff = Staff.objects.create(
            company=self.company,
            employee_number="S200",
            name="未提出 花子",
        )
        draft_staff = Staff.objects.create(
            company=self.company,
            employee_number="S300",
            name="下書き 次郎",
        )
        Staff.objects.create(
            company=self.company,
            employee_number="E100",
            name="社員 応援",
            is_employee=True,
        )
        AvailabilitySubmission.objects.create(
            staff=submitted_staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilitySubmission.objects.create(
            staff=draft_staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.DRAFT,
        )

        response = self.client.get(reverse("missing_submissions"), {"month": "2026-07"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["target_staff_count"], 3)
        self.assertEqual(response.context["submitted_count"], 1)
        self.assertEqual(response.context["missing_count"], 2)
        self.assertContains(response, missing_staff.name)
        self.assertContains(response, draft_staff.name)
        self.assertContains(response, "下書き")
        self.assertNotContains(response, submitted_staff.name)
        self.assertNotContains(response, "社員 応援")

    def test_availability_import_marks_staff_as_submitted_by_employee_number(self):
        submitted_staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="提出済み 太郎",
        )
        missing_staff = Staff.objects.create(
            company=self.company,
            employee_number="S200",
            name="未提出 花子",
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "公休申請"
        sheet.append(["社員番号", "氏名", "2026/7/1", "2026/7/2"])
        sheet.append(["S100", submitted_staff.name, "公休", "有給"])
        sheet.append(["S200", missing_staff.name, "", ""])
        file_obj = BytesIO()
        file_obj.name = "availability.xlsx"
        workbook.save(file_obj)
        file_obj.seek(0)

        response = self.client.post(
            reverse("missing_submissions"),
            {
                "month": "2026-07",
                "file": file_obj,
            },
        )

        self.assertRedirects(response, f"{reverse('missing_submissions')}?month=2026-07")
        submission = AvailabilitySubmission.objects.get(
            staff=submitted_staff,
            month=date(2026, 7, 1),
        )
        self.assertEqual(submission.status, AvailabilitySubmission.Status.SUBMITTED)
        self.assertTrue(
            AvailabilityDay.objects.filter(
                submission=submission,
                day=date(2026, 7, 1),
                preferred_off=True,
            ).exists()
        )
        self.assertTrue(
            AvailabilityDay.objects.filter(
                submission=submission,
                day=date(2026, 7, 2),
                paid_leave=True,
            ).exists()
        )

        response = self.client.get(reverse("missing_submissions"), {"month": "2026-07"})

        self.assertEqual(response.context["submitted_count"], 1)
        self.assertEqual(response.context["missing_count"], 1)
        self.assertContains(response, missing_staff.name)
        self.assertNotContains(response, submitted_staff.name)

    def test_availability_import_rejects_staff_over_request_limit(self):
        staff = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="提出済み 太郎",
            desired_off_limit=1,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "公休申請"
        sheet.append(["社員番号", "氏名", "2026/7/1", "2026/7/2"])
        sheet.append(["S100", staff.name, "公休", "有給"])
        file_obj = BytesIO()
        file_obj.name = "availability.xlsx"
        workbook.save(file_obj)
        file_obj.seek(0)

        response = self.client.post(
            reverse("missing_submissions"),
            {
                "month": "2026-07",
                "file": file_obj,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "申請上限を超えているスタッフ")
        self.assertFalse(AvailabilitySubmission.objects.filter(staff=staff).exists())

    def test_availability_request_limit_can_be_bulk_updated_from_missing_page(self):
        first = Staff.objects.create(
            company=self.company,
            employee_number="S100",
            name="青木 太郎",
            desired_off_limit=4,
        )
        second = Staff.objects.create(
            company=self.company,
            employee_number="S200",
            name="田中 花子",
            desired_off_limit=5,
        )

        response = self.client.post(
            reverse("missing_submissions"),
            {
                "action": "bulk_limit",
                "desired_off_limit": "6",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("missing_submissions"), response["Location"])
        first.refresh_from_db()
        second.refresh_from_db()
        self.company.refresh_from_db()
        self.assertEqual(self.company.default_desired_off_limit, 6)
        self.assertEqual(first.desired_off_limit, 6)
        self.assertEqual(second.desired_off_limit, 6)

    def test_admin_can_approve_sudden_leave_with_replacement(self):
        original = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        replacement = Staff.objects.create(
            company=self.company, employee_number="S200", name="田中 花子"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可", assignable=True
        )
        StaffSkill.objects.create(staff=replacement, work_type=work, level=level)
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=original,
            day=date(2026, 7, 1),
            work_type=work,
        )
        leave_request = ShiftLeaveRequest.objects.create(
            period=period,
            staff=original,
            assignment=assignment,
            day=date(2026, 7, 1),
            work_type=work,
            reason="体調不良",
        )

        detail_response = self.client.get(reverse("shift_detail", args=[period.pk]))
        self.assertContains(detail_response, "急な休み申請")
        self.assertContains(detail_response, "田中 花子")

        response = self.client.post(
            reverse("resolve_leave_request", args=[leave_request.pk]),
            {
                "action": "approve",
                "replacement_staff": str(replacement.pk),
                "admin_note": "代替済み",
            },
        )

        self.assertRedirects(response, reverse("shift_detail", args=[period.pk]))
        leave_request.refresh_from_db()
        self.assertEqual(leave_request.status, ShiftLeaveRequest.Status.APPROVED)
        self.assertEqual(leave_request.admin_note, "代替済み")
        self.assertFalse(
            ShiftAssignment.objects.filter(
                period=period,
                staff=original,
                day=date(2026, 7, 1),
            ).exists()
        )
        self.assertTrue(
            ShiftAssignment.objects.filter(
                period=period,
                staff=replacement,
                day=date(2026, 7, 1),
                work_type=work,
                manually_edited=True,
            ).exists()
        )

    def test_employee_tag_appears_as_replacement_without_skill_setting(self):
        original = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        employee = Staff.objects.create(
            company=self.company,
            employee_number="S200",
            name="社員 花子",
            is_employee=True,
        )
        blocked_employee = Staff.objects.create(
            company=self.company,
            employee_number="S300",
            name="不可 社員",
            is_employee=True,
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        blocked_level = SkillLevel.objects.create(
            company=self.company, symbol="×", meaning="不可", assignable=False
        )
        StaffSkill.objects.create(
            staff=blocked_employee, work_type=work, level=blocked_level
        )
        period = ShiftPeriod.objects.create(
            company=self.company,
            month=date(2026, 7, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        assignment = ShiftAssignment.objects.create(
            period=period,
            staff=original,
            day=date(2026, 7, 1),
            work_type=work,
        )
        ShiftLeaveRequest.objects.create(
            period=period,
            staff=original,
            assignment=assignment,
            day=date(2026, 7, 1),
            work_type=work,
        )

        response = self.client.get(reverse("shift_detail", args=[period.pk]))

        candidate_names = [
            candidate["staff"].name
            for item in response.context["pending_leave_requests"]
            for candidate in item["candidates"]
        ]
        self.assertIn("社員 花子", candidate_names)
        self.assertNotIn("不可 社員", candidate_names)

    def test_constraint_routes_are_disabled(self):
        with self.assertRaises(NoReverseMatch):
            reverse("constraint_manage")
        with self.assertRaises(NoReverseMatch):
            reverse("constraint_delete", args=[1])

    def test_used_skill_level_is_protected_from_delete(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S1", name="A"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="可"
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=level)
        response = self.client.post(reverse("skill_delete", args=[level.pk]))
        self.assertRedirects(response, reverse("skill_manage"))
        self.assertTrue(SkillLevel.objects.filter(pk=level.pk).exists())

    def test_skill_map_no_longer_shows_registered_skill_delete_section(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="検索 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="検索対象業務")
        level = SkillLevel.objects.create(
            company=self.company, symbol="◎", meaning="リーダー"
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=level)

        response = self.client.get(reverse("skill_map"))

        self.assertNotContains(response, "登録済みスキルの削除")
        self.assertNotContains(response, "選択した項目を削除")
        self.assertNotContains(response, 'class="skill-checkbox"')
        self.assertContains(response, "検索対象業務")

    def test_skill_map_select_options_show_only_symbols(self):
        Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        WorkType.objects.create(company=self.company, name="受付")
        SkillLevel.objects.create(
            company=self.company, symbol="◎", meaning="リーダー"
        )

        response = self.client.get(reverse("skill_map"))

        self.assertContains(response, ">◎</option>")
        self.assertNotContains(response, "◎ リーダー")

    def test_skill_map_staff_search_filters_edit_matrix(self):
        target = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        other = Staff.objects.create(
            company=self.company, employee_number="S200", name="田中 花子"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可"
        )
        StaffSkill.objects.create(staff=target, work_type=work, level=level)
        StaffSkill.objects.create(staff=other, work_type=work, level=level)

        response = self.client.get(reverse("skill_map"), {"matrix_q": "青木"})

        self.assertContains(response, "青木 太郎")
        self.assertContains(response, f'name="skill_{target.pk}_{work.pk}"')
        self.assertNotContains(response, f'name="skill_{other.pk}_{work.pk}"')

    def test_skill_map_ignores_old_delete_search_query(self):
        target = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        other = Staff.objects.create(
            company=self.company, employee_number="S200", name="田中 花子"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可"
        )
        StaffSkill.objects.create(staff=target, work_type=work, level=level)
        StaffSkill.objects.create(staff=other, work_type=work, level=level)

        response = self.client.get(reverse("skill_map"), {"delete_q": "田中"})

        self.assertContains(response, f'name="skill_{target.pk}_{work.pk}"')
        self.assertContains(response, f'name="skill_{other.pk}_{work.pk}"')
        self.assertContains(response, "青木 太郎")
        self.assertContains(response, "田中 花子")
        self.assertNotContains(response, "登録済みスキルの削除")

    def test_can_update_staff_skills_from_skill_map_matrix(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可"
        )

        response = self.client.post(
            reverse("skill_map"),
            {
                "action": "update_matrix",
                f"skill_{staff.pk}_{work.pk}": str(level.pk),
            },
        )

        self.assertRedirects(response, reverse("skill_map"))
        self.assertTrue(
            StaffSkill.objects.filter(
                staff=staff, work_type=work, level=level
            ).exists()
        )

    def test_can_clear_staff_skill_from_skill_map_matrix(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="青木 太郎"
        )
        work = WorkType.objects.create(company=self.company, name="受付")
        level = SkillLevel.objects.create(
            company=self.company, symbol="○", meaning="対応可"
        )
        StaffSkill.objects.create(staff=staff, work_type=work, level=level)

        response = self.client.post(
            reverse("skill_map"),
            {
                "action": "update_matrix",
                f"skill_{staff.pk}_{work.pk}": "",
            },
        )

        self.assertRedirects(response, reverse("skill_map"))
        self.assertFalse(StaffSkill.objects.filter(staff=staff, work_type=work).exists())

    def test_bulk_delete_only_removes_own_company_skills(self):
        staff = Staff.objects.create(
            company=self.company, employee_number="S100", name="自社"
        )
        work = WorkType.objects.create(company=self.company, name="自社業務")
        level = SkillLevel.objects.create(
            company=self.company, symbol="◎", meaning="可"
        )
        own = StaffSkill.objects.create(staff=staff, work_type=work, level=level)
        other_staff = Staff.objects.create(
            company=self.other_company, employee_number="O100", name="他社"
        )
        other_work = WorkType.objects.create(
            company=self.other_company, name="他社業務"
        )
        other_level = SkillLevel.objects.create(
            company=self.other_company, symbol="◎", meaning="可"
        )
        other = StaffSkill.objects.create(
            staff=other_staff, work_type=other_work, level=other_level
        )
        response = self.client.post(
            reverse("staff_skill_bulk_delete"),
            {"skill_ids": [own.pk, other.pk], "confirmed": "1"},
        )
        self.assertRedirects(response, reverse("skill_map"))
        self.assertFalse(StaffSkill.objects.filter(pk=own.pk).exists())
        self.assertTrue(StaffSkill.objects.filter(pk=other.pk).exists())

