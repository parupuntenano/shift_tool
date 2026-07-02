from datetime import date
from io import BytesIO
from unittest import TestCase

from django.test import TestCase as DjangoTestCase
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from shifts.domain.import_data import (
    ImportedSkillLevel,
    ImportedSkillMap,
    ImportedStaffRow,
    ImportedWorkType,
)
from shifts.infrastructure.importers import SkillMapFileReader
from shifts.infrastructure.master_repository import DjangoMasterRepository
from shifts.infrastructure.repositories import DjangoShiftRepository
from shifts.infrastructure.models import (
    AvailabilityDay,
    AvailabilitySubmission,
    Company,
    ConstraintType,
    IndividualConstraint,
    PreviousMonthShiftDay,
    ShiftAssignment,
    ShiftPeriod,
    Staff,
    StaffSkill,
    SkillLevel,
    WorkType,
)


class SkillMapFileReaderTests(TestCase):
    def test_reads_utf8_csv_to_domain_data(self):
        raw = "社員番号,氏名,備考,受付\nS001,青木,4勤不可,◎\n".encode("utf-8")
        result = SkillMapFileReader().read("skills.csv", BytesIO(raw))
        self.assertEqual(result.rows[0].employee_number, "S001")
        self.assertEqual(result.rows[0].skills, {"受付": "◎"})

    def test_reads_skill_level_sheet_from_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "スキル表"
        sheet.append(["社員番号", "氏名", "公休数", "備考", "受付"])
        sheet.append(["S001", "青木", 9, "", "A"])
        level_sheet = workbook.create_sheet("スキル区分")
        level_sheet.append(["記号", "意味", "優先度", "アサイン可"])
        level_sheet.append(["A", "主担当", 1, "可"])
        work_sheet = workbook.create_sheet("業務マスタ")
        work_sheet.append(["業務名", "最低必要人数", "色", "有効"])
        work_sheet.append(["受付", 2, "#2563eb", "有効"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = SkillMapFileReader().read("skills.xlsx", stream)

        self.assertEqual(result.rows[0].skills, {"受付": "A"})
        self.assertEqual(result.rows[0].monthly_public_holidays, 9)
        self.assertIsNone(result.rows[0].desired_off_limit)
        self.assertEqual(result.skill_levels[0].symbol, "A")
        self.assertEqual(result.skill_levels[0].meaning, "主担当")
        self.assertEqual(result.skill_levels[0].priority, 1)
        self.assertTrue(result.skill_levels[0].assignable)
        self.assertEqual(result.work_types[0].name, "受付")
        self.assertEqual(result.work_types[0].minimum_staff_per_day, 2)
        self.assertEqual(result.work_types[0].color, "#2563eb")

    def test_reads_work_color_from_filled_work_name_cell(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "スキル表"
        sheet.append(["社員番号", "氏名", "備考", "受付"])
        sheet.append(["S001", "青木", "", "A"])
        work_sheet = workbook.create_sheet("業務マスタ")
        work_sheet.append(["業務名", "必要人数", "有効"])
        work_sheet.append(["受付", 2, "有効"])
        work_sheet["A2"].fill = PatternFill("solid", fgColor="22C55E")
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        result = SkillMapFileReader().read("skills.xlsx", stream)

        self.assertEqual(result.work_types[0].color, "#22c55e")

class MasterImportTests(DjangoTestCase):
    def test_import_keeps_note_without_creating_constraints(self):
        company = Company.objects.create(name="テスト", code="import-test")
        WorkType.objects.create(company=company, name="受付")
        data = ImportedSkillMap(
            (ImportedStaffRow("S001", "青木", "2勤1休;4勤不可;単休不可", {"受付": "○"}),)
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company)
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(staff.note, "2勤1休;4勤不可;単休不可")
        self.assertEqual(result["constraints"], 0)
        self.assertEqual(constraints.count(), 0)

    def test_import_does_not_create_soft_max_and_weekend_rest_constraints_from_note(self):
        company = Company.objects.create(name="テスト", code="complex-note-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "可能な限り4勤不可;土日祝は公休",
                    {},
                ),
            )
        )

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(constraints.count(), 0)

    def test_import_does_not_save_base_and_outside_base_rest_patterns(self):
        company = Company.objects.create(name="テスト", code="base-pattern-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "ベース2勤1休;3勤1休",
                    {},
                ),
            )
        )

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        pattern_constraints = constraints.filter(
            rule_type__operator=ConstraintType.Operator.WORK_REST_PATTERN,
        )
        self.assertEqual(pattern_constraints.count(), 0)

    def test_base_pattern_does_not_auto_add_when_candidates_are_written(self):
        company = Company.objects.create(name="テスト", code="manual-candidate-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "ベース3勤1休;2勤1休",
                    {},
                ),
            )
        )

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        patterns = set(
            IndividualConstraint.objects.filter(
                company=company,
                staff=staff,
                rule_type__operator=ConstraintType.Operator.WORK_REST_PATTERN,
            ).values_list("text_value", "strength")
        )
        self.assertEqual(patterns, set())

    def test_rest_pattern_without_base_does_not_auto_add_candidates(self):
        company = Company.objects.create(name="テスト", code="single-pattern-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "3勤1休",
                    {},
                ),
            )
        )

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        patterns = set(
            IndividualConstraint.objects.filter(
                company=company,
                staff=staff,
                rule_type__operator=ConstraintType.Operator.WORK_REST_PATTERN,
            ).values_list("text_value", "strength")
        )
        self.assertEqual(patterns, set())

    def test_generation_rules_add_default_rest_pattern_candidates_when_note_has_no_pattern(self):
        company = Company.objects.create(name="テスト", code="default-pattern-test")
        staff = Staff.objects.create(
            company=company,
            employee_number="S001",
            name="青木",
        )

        rules = DjangoShiftRepository().rules_for_generation(
            company.id,
            include_default_patterns=True,
        )

        patterns = {
            rule.text_value
            for rule in rules
            if rule.staff_id == staff.id and rule.operator == "work_rest_pattern"
        }
        self.assertEqual(patterns, {"1,1", "2,1", "3,1", "4,1", "5,2"})

    def test_generation_rules_do_not_add_default_patterns_when_note_has_pattern(self):
        company = Company.objects.create(name="テスト", code="no-default-pattern-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "3勤1休",
                    {},
                ),
            )
        )
        DjangoMasterRepository().save_skill_map(company.id, data)
        staff = Staff.objects.get(company=company, employee_number="S001")

        rules = DjangoShiftRepository().rules_for_generation(
            company.id,
            include_default_patterns=True,
        )

        patterns = {
            rule.text_value
            for rule in rules
            if rule.staff_id == staff.id and rule.operator == "work_rest_pattern"
        }
        self.assertEqual(patterns, {"1,1", "2,1", "3,1", "4,1", "5,2"})

    def test_base_pattern_candidates_respect_specific_limits(self):
        company = Company.objects.create(name="テスト", code="base-limit-test")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "ベース3勤1休;4勤以上禁止;単日禁止",
                    {},
                ),
            )
        )

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        patterns = set(
            IndividualConstraint.objects.filter(
                company=company,
                staff=staff,
                rule_type__operator=ConstraintType.Operator.WORK_REST_PATTERN,
            ).values_list("text_value", "strength")
        )
        self.assertEqual(patterns, set())

    def test_import_does_not_create_staff_login_account(self):
        company = Company.objects.create(name="テスト", code="account-import-test")
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}),))

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.select_related("user").get(
            company=company, employee_number="50592"
        )
        self.assertIsNone(staff.user)
        self.assertEqual(result["accounts"], 0)

    def test_import_updates_staff_public_holidays_and_uses_company_request_limit(self):
        company = Company.objects.create(
            name="テスト", code="staff-off-import-test", default_desired_off_limit=6
        )
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}, 9),))

        DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="50592")
        self.assertEqual(staff.monthly_public_holidays, 9)
        self.assertEqual(staff.desired_off_limit, 6)

    def test_import_updates_skill_levels_from_excel_definition(self):
        company = Company.objects.create(name="テスト", code="level-import-test")
        data = ImportedSkillMap(
            rows=(ImportedStaffRow("S001", "青木", "", {"受付": "A"}),),
            skill_levels=(
                ImportedSkillLevel("A", "主担当", 1, True),
            ),
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        level = SkillLevel.objects.get(company=company, symbol="A")
        self.assertEqual(result["levels"], 1)
        self.assertEqual((level.meaning, level.priority, level.assignable), ("主担当", 1, True))

    def test_import_updates_work_types_from_excel_definition(self):
        company = Company.objects.create(name="テスト", code="work-import-test")
        data = ImportedSkillMap(
            rows=(),
            work_types=(ImportedWorkType("受付", 2, True, "#16a34a"),),
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        work = WorkType.objects.get(company=company, name="受付")
        self.assertEqual(result["works"], 1)
        self.assertEqual(work.required_staff_per_day, 2)
        self.assertEqual(work.color, "#16a34a")
        self.assertTrue(work.active)

    def test_generation_skill_flags_detect_instructor_and_trainee_levels(self):
        company = Company.objects.create(name="テスト", code="skill-flag-test")
        instructor = Staff.objects.create(
            company=company, employee_number="S001", name="指導者"
        )
        trainee = Staff.objects.create(
            company=company, employee_number="S002", name="研修中"
        )
        work = WorkType.objects.create(company=company, name="受付")
        instructor_level = SkillLevel.objects.create(
            company=company,
            symbol="◎",
            meaning="主担当・指導可",
            priority=1,
            assignable=True,
        )
        trainee_level = SkillLevel.objects.create(
            company=company,
            symbol="△",
            meaning="補助・訓練中",
            priority=3,
            assignable=True,
        )
        StaffSkill.objects.create(
            staff=instructor, work_type=work, level=instructor_level
        )
        StaffSkill.objects.create(staff=trainee, work_type=work, level=trainee_level)

        ratings = {
            item.staff_id: item
            for item in DjangoShiftRepository().skills_for_generation(company.id)
        }

        self.assertTrue(ratings[instructor.id].instructor_capable)
        self.assertFalse(ratings[instructor.id].trainee)
        self.assertFalse(ratings[trainee.id].instructor_capable)
        self.assertTrue(ratings[trainee.id].trainee)

    def test_generation_uses_existing_previous_shift_before_imported_result(self):
        company = Company.objects.create(name="テスト", code="previous-source-test")
        staff = Staff.objects.create(
            company=company, employee_number="S001", name="青木"
        )
        generated_work = WorkType.objects.create(company=company, name="ロール")
        imported_work = WorkType.objects.create(company=company, name="受付")
        period = ShiftPeriod.objects.create(
            company=company,
            month=date(2026, 6, 1),
            status=ShiftPeriod.Status.PUBLISHED,
        )
        ShiftAssignment.objects.create(
            period=period,
            staff=staff,
            day=date(2026, 6, 30),
            work_type=generated_work,
        )
        PreviousMonthShiftDay.objects.create(
            company=company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=imported_work,
            raw_value="受付",
        )

        previous_days = DjangoShiftRepository().previous_shift_days_for_generation(
            company.id, date(2026, 7, 1)
        )

        last_day = next(item for item in previous_days if item.day == date(2026, 6, 30))
        self.assertEqual(last_day.work_id, generated_work.id)
        self.assertEqual(last_day.status, PreviousMonthShiftDay.Status.WORK)
        self.assertTrue(
            any(
                item.day == date(2026, 6, 29)
                and item.status == PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY
                for item in previous_days
            )
        )

    def test_reimport_does_not_create_staff_login_account(self):
        company = Company.objects.create(name="テスト", code="account-reimport-test")
        data = ImportedSkillMap((ImportedStaffRow("50592", "青木", "", {}),))
        repository = DjangoMasterRepository()
        repository.save_skill_map(company.id, data)

        result = repository.save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="50592")
        self.assertIsNone(staff.user)
        self.assertEqual(result["accounts"], 0)

    def test_import_does_not_create_work_constraints_from_note(self):
        company = Company.objects.create(name="テスト", code="constraint-note-test")
        WorkType.objects.create(company=company, name="ロール")
        WorkType.objects.create(company=company, name="エーカス")
        data = ImportedSkillMap(
            (
                ImportedStaffRow(
                    "S001",
                    "青木",
                    "ロールとエーカス交互;ロール連続不可;エーカス禁止",
                    {},
                ),
            )
        )

        result = DjangoMasterRepository().save_skill_map(company.id, data)

        staff = Staff.objects.get(company=company, employee_number="S001")
        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(result["constraints"], 0)
        self.assertEqual(constraints.count(), 0)

    def test_reimport_leaves_manual_constraints_and_does_not_create_note_constraints(self):
        company = Company.objects.create(name="テスト", code="constraint-reimport-test")
        rule_type = ConstraintType.objects.create(
            company=company,
            name="手入力用",
            operator=ConstraintType.Operator.CUSTOM,
            default_strength=5,
        )
        repository = DjangoMasterRepository()
        repository.save_skill_map(
            company.id,
            ImportedSkillMap((ImportedStaffRow("S001", "青木", "2勤1休", {}),)),
        )
        staff = Staff.objects.get(company=company, employee_number="S001")
        IndividualConstraint.objects.create(
            company=company,
            staff=staff,
            rule_type=rule_type,
            name="手入力メモ",
            kind=rule_type.operator,
            strength=5,
            parameters={"source": "manual"},
        )

        result = repository.save_skill_map(
            company.id,
            ImportedSkillMap((ImportedStaffRow("S001", "青木", "3勤1休", {}),)),
        )

        constraints = IndividualConstraint.objects.filter(company=company, staff=staff)
        self.assertEqual(result["constraints"], 0)
        self.assertEqual(constraints.filter(parameters__source="staff_note").count(), 0)
        self.assertTrue(constraints.filter(name="手入力メモ").exists())


class ShiftRepositoryAvailabilityTests(DjangoTestCase):
    def test_unsubmitted_staff_is_available_without_auto_public_holidays(self):
        company = Company.objects.create(name="テスト", code="auto-availability-test")
        staff = Staff.objects.create(company=company, employee_number="S001", name="青木")

        rows = DjangoShiftRepository().availability_for_generation(
            company.id, date(2026, 7, 1)
        )

        staff_rows = [row for row in rows if row.staff_id == staff.id]
        weekly_off_numbers = [row.day.day for row in staff_rows if row.preferred_off]
        expected_first_off = (staff.id - 1) % 7 + 1
        self.assertEqual(len(staff_rows), 31)
        self.assertTrue(all(row.available for row in staff_rows))
        self.assertEqual(weekly_off_numbers, [])

    def test_unsubmitted_staff_public_holidays_do_not_follow_previous_month_result(self):
        company = Company.objects.create(name="テスト", code="previous-off-test")
        staff = Staff.objects.create(company=company, employee_number="S001", name="青木")
        PreviousMonthShiftDay.objects.create(
            company=company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.PUBLIC_HOLIDAY,
        )

        rows = DjangoShiftRepository().availability_for_generation(
            company.id, date(2026, 7, 1)
        )

        weekly_off_numbers = [
            row.day.day for row in rows if row.staff_id == staff.id and row.preferred_off
        ]
        self.assertEqual(weekly_off_numbers, [])

    def test_unsubmitted_staff_rest_pattern_does_not_create_auto_public_holidays(self):
        company = Company.objects.create(name="テスト", code="previous-pattern-test")
        staff = Staff.objects.create(company=company, employee_number="S001", name="青木")
        rule_type = ConstraintType.objects.create(
            company=company,
            name="勤休パターン",
            operator=ConstraintType.Operator.WORK_REST_PATTERN,
            default_strength=5,
            default_is_hard=False,
        )
        IndividualConstraint.objects.create(
            company=company,
            staff=staff,
            rule_type=rule_type,
            name="青木：2勤1休",
            kind=rule_type.operator,
            text_value="2,1",
            strength=5,
            is_hard=False,
        )
        work = WorkType.objects.create(company=company, name="受付")
        PreviousMonthShiftDay.objects.create(
            company=company,
            staff=staff,
            day=date(2026, 6, 29),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
        )
        PreviousMonthShiftDay.objects.create(
            company=company,
            staff=staff,
            day=date(2026, 6, 30),
            status=PreviousMonthShiftDay.Status.WORK,
            work_type=work,
        )

        rows = DjangoShiftRepository().availability_for_generation(
            company.id, date(2026, 7, 1)
        )

        off_numbers = [
            row.day.day for row in rows if row.staff_id == staff.id and row.preferred_off
        ]
        self.assertEqual(off_numbers, [])

    def test_submitted_staff_request_is_not_overwritten_by_auto_holidays(self):
        company = Company.objects.create(name="テスト", code="submitted-availability-test")
        staff = Staff.objects.create(company=company, employee_number="S001", name="青木")
        submission = AvailabilitySubmission.objects.create(
            staff=staff,
            month=date(2026, 7, 1),
            status=AvailabilitySubmission.Status.SUBMITTED,
        )
        AvailabilityDay.objects.create(
            submission=submission,
            day=date(2026, 7, 2),
            available=True,
            preferred_off=True,
        )

        rows = DjangoShiftRepository().availability_for_generation(
            company.id, date(2026, 7, 1)
        )

        off_numbers = [
            row.day.day for row in rows if row.staff_id == staff.id and row.preferred_off
        ]
        self.assertEqual(off_numbers, [2])
